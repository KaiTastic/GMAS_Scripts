"""
当前日期文件处理模块

处理指定日期的所有图幅文件的集合，包括报告生成和统计功能
"""

import os
import logging
import functools
import threading
import warnings
from datetime import datetime
from typing import Dict, List, Optional, Any, Tuple, Union
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment

from ..data_models.observation_data import ObservationData
from ..data_models.date_types import DateType
from ..file_handlers.kmz_handler import KMZFile
from .mapsheet_daily import MapsheetDailyFile

# 使用系统配置模块
from config.config_manager import ConfigManager

# 移除模块级的配置获取，改为动态获取以避免初始化问题

# 创建 logger 实例
logger = logging.getLogger('Current Date Files')
if not logger.handlers:  # 避免重复添加处理器
    logger.setLevel(logging.INFO)  # 改为INFO级别以看到详细日志
    handler = logging.StreamHandler()
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    handler.setFormatter(formatter)
    logger.addHandler(handler)
    # 添加这一行防止向上传播
    logger.propagate = False

class CurrentDateFiles:
    """当前日期文件容器类，用于存储指定日期的所有图幅的集合"""
    
    _instances = {}  # 基于日期的实例缓存
    _lock = threading.Lock()
    maps_info: Dict[int, Dict[str, Any]] = {}

    def __new__(cls, currentdate: 'DateType', *args, **kwargs):
        """改进的单例模式，基于日期创建不同实例"""
        date_key = str(currentdate)
        
        if date_key not in cls._instances:
            with cls._lock:
                if date_key not in cls._instances:
                    # 使用新的图幅管理器
                    from .mapsheet_manager import mapsheet_manager
                    cls.maps_info = mapsheet_manager.maps_info
                    cls._instances[date_key] = super(CurrentDateFiles, cls).__new__(cls)
                    cls._instances[date_key]._initialized = False
        
        return cls._instances[date_key]

    def __init__(self, currentdate: 'DateType', enable_predict: bool = False):
        """
        初始化当前日期文件集合
        
        Args:
            currentdate: 日期对象
            enable_predict: 是否启用进度预测显示
        """
        if self._initialized:
            return
            
        self.currentDate = currentdate
        self.currentDateFiles: List[MapsheetDailyFile] = []
        self.enable_predict = enable_predict  # 新增：保存predict状态
        
        # 清理所有缓存属性
        self._clear_cache()
        
        # 获取当天的文件
        self.__datacollect()
        self._initialized = True

    def _clear_cache(self):
        """清理所有缓存属性"""
        self._cached_sorted_mapsheets = None
        self._error_msg_cache = None

    @functools.cached_property
    def sorted_mapsheets(self) -> List[MapsheetDailyFile]:
        """缓存的排序图幅列表"""
        return sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)

    @functools.cached_property
    def mapsheet_targets(self) -> Dict[str, int]:
        """获取各图幅的目标点数"""
        from .mapsheet_manager import mapsheet_manager
        targets = {}
        for mapsheet in self.sorted_mapsheets:
            roman_name = mapsheet.romanName
            # 从图幅管理器获取目标点数
            target = mapsheet_manager.get_mapsheet_target(roman_name)
            targets[roman_name] = target
        return targets

    @classmethod
    def mapsInfo(cls) -> Dict[int, Dict[str, Any]]:
        """
        从100K图幅名称信息表中获取图幅的罗马名称和拉丁名称
        
        注意：此方法已弃用，请使用 mapsheet_manager.maps_info 替代
        """
        warnings.warn(
            "CurrentDateFiles.mapsInfo() 已弃用，请使用 mapsheet_manager.maps_info",
            DeprecationWarning,
            stacklevel=2
        )
        
        from .mapsheet_manager import mapsheet_manager
        return mapsheet_manager.maps_info

    def __datacollect(self) -> 'CurrentDateFiles':
        """收集当天的所有文件 - 使用统一的图幅管理器"""
        from .mapsheet_manager import mapsheet_manager
        # 使用图幅管理器创建图幅对象集合
        self.currentDateFiles = mapsheet_manager.create_mapsheet_collection(MapsheetDailyFile, self.currentDate)
        return self

    @functools.cached_property
    def totalDaiyIncreasePointNum(self) -> int:
        """本日新增点数总计"""
        return sum(
            mapsheet.dailyincreasePointNum or 0 
            for mapsheet in self.currentDateFiles
        )

    @functools.cached_property
    def dailyFinishedPoints(self) -> Dict[str, int]:
        """截止本日各图幅完成的点数"""
        finished_points = {}
        for mapsheet in self.sorted_mapsheets:
            roman_name = mapsheet.romanName
            
            # 策略1: 如果有当前文件且点数大于0，使用当前总点数
            if (mapsheet.currentTotalPointNum is not None and 
                mapsheet.currentTotalPointNum > 0):
                finished_points[roman_name] = mapsheet.currentTotalPointNum
                continue
            
            # 策略2: 如果有当前placemarks且点数大于0
            if (mapsheet.currentPlacemarks is not None and 
                mapsheet.currentPlacemarks.pointsCount > 0):
                finished_points[roman_name] = mapsheet.currentPlacemarks.pointsCount
                continue
            
            # 策略3: 关键修复 - 如果当前没有数据但有历史数据，使用历史数据
            # 这种情况表示当天没有新增，但之前有累计数据
            if (mapsheet.lastPlacemarks is not None and 
                mapsheet.lastPlacemarks.pointsCount > 0):
                finished_points[roman_name] = mapsheet.lastPlacemarks.pointsCount
                logger.info(f"图幅 {roman_name} 当天无新增，使用历史累计数据: {mapsheet.lastPlacemarks.pointsCount} 点")
                continue
            
            # 策略4: 如果当前文件存在但是空的，并且历史文件有数据
            # 这处理了当天文件存在但为空的情况
            if (mapsheet.currentPlacemarks is not None and 
                mapsheet.currentPlacemarks.pointsCount == 0 and
                mapsheet.lastPlacemarks is not None and 
                mapsheet.lastPlacemarks.pointsCount > 0):
                finished_points[roman_name] = mapsheet.lastPlacemarks.pointsCount
                logger.info(f"图幅 {roman_name} 当天文件为空，使用历史累计数据: {mapsheet.lastPlacemarks.pointsCount} 点")
                continue
            
            # 最后才设为0，并详细记录原因
            finished_points[roman_name] = 0
            has_current = mapsheet.currentPlacemarks is not None
            has_last = mapsheet.lastPlacemarks is not None
            current_count = mapsheet.currentPlacemarks.pointsCount if has_current else "N/A"
            last_count = mapsheet.lastPlacemarks.pointsCount if has_last else "N/A"
            
            logger.warning(f"图幅 {roman_name} 无法获取完成点数，设为0。详情: "
                         f"当前文件={has_current}(点数={current_count}), "
                         f"历史文件={has_last}(点数={last_count})")
        
        return finished_points

    def estimate_progress(self, confidence_level: float = 0.8, 
                          include_charts: bool = False) -> Dict[str, Any]:
        """
        估算整体项目和各图幅的完成进度
        
        Args:
            confidence_level: 置信水平，默认0.8
            include_charts: 是否包含图表，默认False
            
        Returns:
            Dict: 包含整体和各图幅预测结果的字典
        """
        logger.info(f"开始估算项目进度，置信水平: {confidence_level}")
        
        # 导入进度预测模块
        try:
            from ..progress_estimation import EstimationFacade
            from config.config_manager import ConfigManager
            
            # 创建配置管理器和估算外观
            config = ConfigManager()
            estimation_facade = EstimationFacade(config)
        except ImportError as e:
            logger.error(f"无法导入进度预测模块: {e}")
            return {"error": "进度预测模块不可用"}
        
        # 1. 估算整体项目进度
        overall_target = sum(self.mapsheet_targets.values())
        overall_current = self.totalPointNum
        
        overall_result = estimation_facade.advanced_estimate(
            target_points=overall_target,
            current_points=overall_current,
            confidence_level=confidence_level
        )
        
        # 2. 估算各图幅进度
        mapsheet_results = {}
        for mapsheet in self.sorted_mapsheets:
            roman_name = mapsheet.romanName
            target = self.mapsheet_targets.get(roman_name, 0)
            
            # 如果没有目标点数，跳过
            if target <= 0:
                logger.warning(f"图幅 {roman_name} 没有设置目标点数，跳过进度估算")
                continue
            
            # 获取当前点数
            current = self.dailyFinishedPoints.get(roman_name, 0)
            
            # 进行图幅特定的进度估算
            try:
                result = self._estimate_mapsheet_specific(roman_name, target, current)
                mapsheet_results[roman_name] = result
                logger.info(f"图幅 {roman_name} 进度: {result.get('completion_percentage', 0):.1f}%")
            except Exception as e:
                logger.error(f"估算图幅 {roman_name} 进度失败: {e}")
                mapsheet_results[roman_name] = {"error": str(e)}
        
        # 3. 汇总结果
        completion_statuses = self._analyze_completion_statuses(mapsheet_results)
        
        return {
            "overall": overall_result,
            "mapsheets": mapsheet_results,
            "summary": {
                "total_mapsheets": len(self.sorted_mapsheets),
                "estimated_mapsheets": len(mapsheet_results),
                "completion_statuses": completion_statuses,
                "avg_completion": self._calculate_avg_completion(mapsheet_results)
            },
            "timestamp": datetime.now().isoformat()
        }
    
    def _analyze_completion_statuses(self, mapsheet_results: Dict[str, Dict]) -> Dict[str, int]:
        """分析图幅完成状态分布"""
        statuses = {
            "not_started": 0,  # 0%
            "early_stage": 0,  # 0-25%
            "in_progress": 0,  # 25-75%
            "advanced": 0,     # 75-95%
            "near_complete": 0,# 95-99%
            "completed": 0     # 100%
        }
        
        for roman_name, result in mapsheet_results.items():
            if "error" in result:
                continue
            
            # 从我们的数据计算真实完成百分比
            target = self.mapsheet_targets.get(roman_name, 1)
            current = self.dailyFinishedPoints.get(roman_name, 0)
            completion = (current / target) * 100 if target > 0 else 0
            
            if completion == 0:
                statuses["not_started"] += 1
            elif completion < 25:
                statuses["early_stage"] += 1
            elif completion < 75:
                statuses["in_progress"] += 1
            elif completion < 95:
                statuses["advanced"] += 1
            elif completion < 100:
                statuses["near_complete"] += 1
            else:
                statuses["completed"] += 1
                
        return statuses
    
    def _calculate_avg_completion(self, mapsheet_results: Dict[str, Dict]) -> float:
        """计算平均完成度"""
        completions = []
        for roman_name, result in mapsheet_results.items():
            if "error" not in result:
                # 从我们的数据计算真实完成百分比
                target = self.mapsheet_targets.get(roman_name, 1)
                current = self.dailyFinishedPoints.get(roman_name, 0)
                completion = (current / target) * 100 if target > 0 else 0
                completions.append(completion)
                
        if not completions:
            return 0.0
            
        return sum(completions) / len(completions)

    def _estimate_mapsheet_specific(self, mapsheet_name: str, target_points: int, current_points: int) -> Dict[str, Any]:
        """
        针对特定图幅进行个性化进度估算
        
        Args:
            mapsheet_name: 图幅名称
            target_points: 目标点数
            current_points: 当前点数
            
        Returns:
            Dict[str, Any]: 估算结果
        """
        from datetime import datetime, timedelta
        from collections import defaultdict
        
        try:
            # 导入数据分析器
            from ..progress_estimation._internal.data_analyzer import DataAnalyzer
            from ..data_models.date_types import DateType
            
            # 初始化数据分析器
            analyzer = DataAnalyzer()
            start_date = DateType(datetime.now() - timedelta(days=30))
            success = analyzer.load_historical_data(start_date)
            
            if not success or analyzer.historical_data.empty:
                logger.warning(f"图幅 {mapsheet_name}: 无法加载历史数据，使用默认估算")
                return self._fallback_estimation(target_points, current_points)
            
            # 统计该图幅的历史完成速度
            mapsheet_daily_points = []
            total_points = 0
            active_days = 0
            
            for _, row in analyzer.historical_data.iterrows():
                mapsheet_details = row.get('mapsheet_details', {})
                if isinstance(mapsheet_details, dict):
                    points = mapsheet_details.get(mapsheet_name, 0)
                    mapsheet_daily_points.append(points)
                    total_points += points
                    if points > 0:
                        active_days += 1
            
            # 计算该图幅的日均完成速度
            if active_days == 0:
                logger.warning(f"图幅 {mapsheet_name}: 历史无活跃记录，使用全局平均")
                return self._fallback_estimation(target_points, current_points)
            
            # 使用活跃天数计算平均值（更准确）
            daily_avg = total_points / active_days
            
            # 计算加权平均（近期数据权重更高）
            if len(mapsheet_daily_points) >= 7:
                recent_points = mapsheet_daily_points[-7:]  # 最近7天
                recent_avg = sum(recent_points) / len([p for p in recent_points if p > 0]) if any(recent_points) else 0
                # 如果近期有活动，给60%权重，否则用历史平均
                weighted_avg = (recent_avg * 0.6 + daily_avg * 0.4) if recent_avg > 0 else daily_avg
            else:
                weighted_avg = daily_avg
            
            # 确保最小速度（避免无限期预测）
            min_daily_speed = max(1.0, weighted_avg)  # 至少1点/天
            
            # 计算剩余点数和预测天数
            remaining_points = max(0, target_points - current_points)
            if remaining_points == 0:
                return {
                    'completion_percentage': 100,
                    'estimated_finish_date': datetime.now().date(),
                    'days_remaining': 0,
                    'confidence': 100,
                    'daily_target': 0,
                    'current_velocity': weighted_avg,
                    'mapsheet_specific': True,
                    'recommendations': ['已完成！']
                }
            
            estimated_days = remaining_points / min_daily_speed
            estimated_finish_date = datetime.now() + timedelta(days=estimated_days)
            
            # 计算完成百分比
            completion_percentage = (current_points / target_points) * 100 if target_points > 0 else 0
            
            # 计算置信度（基于数据质量）
            confidence = min(100, (active_days / 30) * 100)  # 活跃天数越多置信度越高
            
            # 生成建议
            recommendations = []
            if weighted_avg < daily_avg * 0.8:
                recommendations.append("近期进度放缓，建议加强资源投入")
            elif weighted_avg > daily_avg * 1.2:
                recommendations.append("近期进度良好，保持当前节奏")
            
            return {
                'completion_percentage': completion_percentage,
                'estimated_finish_date': estimated_finish_date.date(),
                'days_remaining': int(estimated_days),
                'confidence': confidence,
                'daily_target': min_daily_speed,
                'current_velocity': weighted_avg,
                'mapsheet_specific': True,
                'historical_avg': daily_avg,
                'active_days': active_days,
                'recommendations': recommendations
            }
            
        except Exception as e:
            logger.error(f"图幅 {mapsheet_name} 个性化估算失败: {e}")
            return self._fallback_estimation(target_points, current_points)
    
    def _fallback_estimation(self, target_points: int, current_points: int) -> Dict[str, Any]:
        """降级估算（当个性化估算失败时使用）"""
        from datetime import datetime, timedelta
        
        try:
            from ..progress_estimation import EstimationFacade
            from config.config_manager import ConfigManager
            
            config = ConfigManager()
            estimation_facade = EstimationFacade(config)
            return estimation_facade.quick_estimate(target_points=target_points, current_points=current_points)
        except Exception:
            # 最基本的估算
            remaining = max(0, target_points - current_points)
            # 假设每天5点的最低速度
            days = remaining / 5 if remaining > 0 else 0
            
            return {
                'completion_percentage': (current_points / target_points) * 100 if target_points > 0 else 0,
                'estimated_finish_date': (datetime.now() + timedelta(days=days)).date(),
                'days_remaining': int(days),
                'confidence': 20,  # 低置信度
                'daily_target': 5,
                'current_velocity': 5,
                'recommendations': ['数据不足，建议人工核实']
            }

    def display_progress_estimation(self) -> None:
        """在屏幕上显示进度估算结果"""
        try:
            # 获取进度估算结果
            results = self.estimate_progress()
            
            if "error" in results:
                print(f"❌ 进度估算失败: {results['error']}")
                return
            
            # 显示整体进度
            overall = results.get("overall", {}).get("basic_estimation", {})
            
            print("\n" + "="*60)
            print(f"📊 项目进度估算 - {self.currentDate}")
            print("="*60)
            
            # 计算完成百分比
            target_points = overall.get('target_points', 1)
            current_points = overall.get('current_points', 0)
            completion_percentage = (current_points / target_points) * 100 if target_points > 0 else 0
            
            print(f"\n🔍 整体进度: {completion_percentage:.1f}%")
            
            finish_date = overall.get('estimated_finish_date')
            if finish_date:
                print(f"📅 预计完成日期: {finish_date.strftime('%Y年%m月%d日')}")
            
            days_remaining = overall.get('estimated_days_remaining', 0)
            print(f"⏱️ 剩余天数: {days_remaining} 天")
            
            # 计算每日目标
            daily_average = overall.get('weighted_daily_average', 0)
            print(f"🎯 每日目标: {daily_average:.1f} 点/天")
            
            # 显示图幅状态统计
            statuses = results.get("summary", {}).get("completion_statuses", {})
            print("\n🗂️ 图幅状态分布:")
            print(f"  ⚪ 未开始 (0%): {statuses.get('not_started', 0)} 个")
            print(f"  🔵 初期阶段 (0-25%): {statuses.get('early_stage', 0)} 个")
            print(f"  🟡 进行中 (25-75%): {statuses.get('in_progress', 0)} 个")
            print(f"  🟠 后期阶段 (75-95%): {statuses.get('advanced', 0)} 个")
            print(f"  🟢 接近完成 (95-99%): {statuses.get('near_complete', 0)} 个")
            print(f"  ✅ 已完成 (100%): {statuses.get('completed', 0)} 个")
            
            # 显示近期将完成的图幅
            near_complete = []
            for roman_name, result in results.get("mapsheets", {}).items():
                if "error" not in result:
                    completion = result.get("completion_percentage", 0)
                    days = result.get("days_remaining", 0)
                    if completion >= 75 and days <= 14:
                        near_complete.append((roman_name, completion, days))
            
            if near_complete:
                print("\n🚀 近期预计完成的图幅:")
                for roman_name, completion, days in sorted(near_complete, key=lambda x: x[2])[:5]:
                    print(f"  📌 {roman_name}: {completion:.1f}% 完成，预计 {days} 天内完成")
                    
            print("\n" + "="*60)
            
        except Exception as e:
            logger.error(f"显示进度估算结果失败: {e}")
            print(f"❌ 显示进度估算结果失败: {e}")

    def add_progress_estimation_to_excel(self, excel_path: Optional[str] = None) -> bool:
        """
        将进度估算结果添加到Excel报告
        
        Args:
            excel_path: Excel文件路径，默认使用每日统计报告路径
            
        Returns:
            bool: 成功返回True，失败返回False
        """
        try:
            # 使用默认路径或指定路径
            if excel_path is None:
                excel_path = self._get_excel_output_path()
                
            # 检查文件是否存在
            if not os.path.exists(excel_path):
                logger.error(f"Excel文件不存在: {excel_path}")
                return False
                
            # 获取进度估算结果
            estimation_results = self.estimate_progress()
            
            if "error" in estimation_results:
                logger.error(f"进度估算失败: {estimation_results['error']}")
                return False
            
            # 加载Excel文件
            wb = load_workbook(excel_path)
            
            # 添加或获取进度估算工作表
            if "Progress Estimation" in wb.sheetnames:
                sheet = wb["Progress Estimation"]
            else:
                sheet = wb.create_sheet("Progress Estimation")
                
            # 设置进度估算工作表
            self._setup_progress_estimation_sheet(sheet, estimation_results)
            
            # 保存Excel文件
            wb.save(excel_path)
            logger.info(f"成功将进度估算结果添加到Excel报告: {excel_path}")
            return True
            
        except Exception as e:
            logger.error(f"添加进度估算结果到Excel报告失败: {e}")
            return False
            
    def _setup_progress_estimation_sheet(self, sheet, estimation_results: Dict) -> None:
        """设置进度估算工作表"""
        # 清空工作表
        for row in sheet.rows:
            for cell in row:
                cell.value = None
                
        # 设置标题
        sheet.cell(row=1, column=1, value=f"进度估算报告 - {self.currentDate}")
        sheet.merge_cells('A1:D1')
        
        # 设置整体进度
        overall = estimation_results.get("overall", {}).get("basic_estimation", {})
        sheet.cell(row=3, column=1, value="整体项目进度")
        sheet.cell(row=3, column=2, value=f"{overall.get('completion_percentage', 0):.1f}%")
        
        sheet.cell(row=4, column=1, value="预计完成日期")
        finish_date = overall.get('estimated_finish_date')
        if finish_date:
            sheet.cell(row=4, column=2, value=finish_date.strftime('%Y年%m月%d日'))
        
        sheet.cell(row=5, column=1, value="剩余天数")
        sheet.cell(row=5, column=2, value=overall.get('days_remaining', 0))
        
        sheet.cell(row=6, column=1, value="每日目标")
        sheet.cell(row=6, column=2, value=overall.get('daily_target', 0))
        
        # 设置图幅进度表格
        sheet.cell(row=8, column=1, value="图幅名称")
        sheet.cell(row=8, column=2, value="完成百分比")
        sheet.cell(row=8, column=3, value="预计完成日期")
        sheet.cell(row=8, column=4, value="剩余天数")
        
        # 填充图幅进度数据
        row_num = 9
        mapsheet_results = estimation_results.get("mapsheets", {})
        for roman_name, result in mapsheet_results.items():
            if "error" in result:
                continue
                
            sheet.cell(row=row_num, column=1, value=roman_name)
            sheet.cell(row=row_num, column=2, value=f"{result.get('completion_percentage', 0):.1f}%")
            
            finish_date = result.get('estimated_finish_date')
            if finish_date:
                sheet.cell(row=row_num, column=3, value=finish_date.strftime('%Y年%m月%d日'))
                
            sheet.cell(row=row_num, column=4, value=result.get('days_remaining', 0))
            
            row_num += 1
            
        # 设置样式
        font_header = Font(name='Calibri', size=12, bold=True)
        font = Font(name='Calibri', size=11)
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        # 应用样式
        for row in range(1, row_num):
            for col in range(1, 5):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                
                if row in [1, 3, 8]:
                    cell.font = font_header
                else:
                    cell.font = font
                    
        # 调整列宽
        for col in sheet.columns:
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max(max_length + 2, 15)
            sheet.column_dimensions[col[0].column_letter].width = adjusted_width

    @functools.cached_property
    def dailyIncreasedPoints(self) -> Dict[str, int]:
        """本日各图幅新增的点数"""
        return {
            mapsheet.romanName: mapsheet.dailyincreasePointNum or 0
            for mapsheet in self.sorted_mapsheets
        }

    @functools.cached_property
    def totalDaiyIncreaseRouteNum(self) -> int:
        """本日新增线路数总计"""
        return sum(
            mapsheet.dailyincreaseRouteNum or 0 
            for mapsheet in self.currentDateFiles
        )

    @functools.cached_property
    def DailyPlans(self) -> Dict[str, str]:
        """本日各图幅的计划"""
        return {
            mapsheet.romanName: '#' if (hasattr(mapsheet, 'nextfilename') and mapsheet.nextfilename) else ''
            for mapsheet in self.sorted_mapsheets
        }

    @functools.cached_property
    def totalDailyPlanNum(self) -> int:
        """本日计划总数"""
        return sum(
            1 for mapsheet in self.currentDateFiles
            if hasattr(mapsheet, 'nextfilename') and mapsheet.nextfilename
        )

    @functools.cached_property
    def totalPointNum(self) -> int:
        """截止当天所有文件的点要素总数"""
        total = 0
        for mapsheet in self.currentDateFiles:
            if mapsheet.currentPlacemarks is not None:
                total += mapsheet.currentPlacemarks.pointsCount
            elif mapsheet.lastPlacemarks is not None:
                total += mapsheet.lastPlacemarks.pointsCount
        return total

    @functools.cached_property
    def allPoints(self) -> Dict:
        """截止当天所有文件的点要素"""
        all_points = {}
        for mapsheet in self.currentDateFiles:
            if mapsheet.currentPlacemarks is not None:
                all_points.update(mapsheet.currentPlacemarks.points)
            elif mapsheet.lastPlacemarks is not None:
                all_points.update(mapsheet.lastPlacemarks.points)
        return all_points

    @functools.cached_property
    def totalRoutesNum(self) -> int:
        """截止当天所有文件的线要素总数"""
        total = 0
        for mapsheet in self.currentDateFiles:
            if mapsheet.currentPlacemarks is not None:
                total += mapsheet.currentPlacemarks.routesCount
            elif mapsheet.lastPlacemarks is not None:
                total += mapsheet.lastPlacemarks.routesCount
        return total

    @functools.cached_property
    def allRoutes(self) -> List:
        """截止当天所有文件的线要素"""
        all_routes = []
        for mapsheet in self.currentDateFiles:
            if mapsheet.currentPlacemarks is not None:
                all_routes.extend(mapsheet.currentPlacemarks.routes)
            elif mapsheet.lastPlacemarks is not None:
                all_routes.extend(mapsheet.lastPlacemarks.routes)
        return all_routes

    @property
    def errorMsg(self) -> List:
        """获取错误消息"""
        if self._error_msg_cache is None:
            self._error_msg_cache = [
                mapsheet.errorMsg for mapsheet in self.currentDateFiles 
                if mapsheet.errorMsg
            ]
        return self._error_msg_cache

    def __contains__(self, key) -> bool:
        """重写__contains__方法, 用于判断图幅文件是否存在"""
        return key in self.currentDateFiles

    def dailyKMZReport(self) -> bool:
        """生成每日KMZ报告"""
        try:
            dailykmz = KMZFile(
                placemarks=ObservationData(
                    points=self.allPoints, 
                    pointsCount=len(self.allPoints), 
                    routes=self.allRoutes, 
                    routesCount=len(self.allRoutes)
                )
            )
            output_path = os.path.join(
                ConfigManager().get('system.workspace'), 
                self.currentDate.yyyymm_str, 
                self.currentDate.yyyymmdd_str, 
                f"GMAS_Points_and_tracks_until_{self.currentDate.yyyymmdd_str}.kmz"
            )
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            success = dailykmz.write_as(newpath=output_path)
            if success:
                logger.info(f"成功生成每日KMZ报告: {output_path}")
            return success
            
        except Exception as e:
            logger.error(f"生成每日KMZ报告失败: {e}")
            return False

    def dailyExcelReport(self) -> bool:
        """生成每日Excel报告"""
        try:
            output_path = self._get_excel_output_path()
            
            # 确保输出目录存在
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # 删除已存在的文件
            if os.path.exists(output_path):
                os.remove(output_path)
                logger.info(f"删除已存在的Excel文件: {output_path}")
            
            # 创建Excel报告
            self._create_excel_workbook(output_path)
            
            # # 添加进度估算结果 - 新增部分
            # self.add_progress_estimation_to_excel(output_path)
            
            logger.info(f"成功创建每日统计报告: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"生成每日Excel报告失败: {e}")
            return False

    def _get_excel_output_path(self) -> str:
        """获取Excel输出路径"""
        return os.path.join(
            ConfigManager().get('system.workspace'), 
            self.currentDate.yyyymm_str, 
            self.currentDate.yyyymmdd_str, 
            f"{self.currentDate.yyyymmdd_str}_Daily_Statistics.xlsx"
        )

    def _get_roman_names_list(self) -> List[str]:
        """获取罗马名称列表"""
        config = ConfigManager()
        sequence_min = config.get('mapsheet.sequence_min')
        sequence_max = config.get('mapsheet.sequence_max')
        return [
            self.__class__.maps_info[sequence]['Roman Name'] 
            for sequence in range(sequence_min, sequence_max + 1)
        ]

    def _create_excel_workbook(self, output_path: str) -> None:
        """创建Excel工作簿"""
        roman_names_list = self._get_roman_names_list()
        max_table_rows = len(roman_names_list) + 5
        max_table_columns = 4
        
        # 创建新的Excel文件
        book = Workbook()
        sheet = book.active
        sheet.title = "Daily Statistics"
        
        # 设置表头、样式和数据
        self._setup_excel_headers(sheet, max_table_rows, max_table_columns, roman_names_list)
        self._setup_excel_styles(sheet, max_table_rows, max_table_columns)
        self._setup_excel_data(sheet, max_table_rows)
        
        # 保存工作簿
        book.save(output_path)

    def _setup_excel_headers(self, sheet, maxTableRows: int, maxTableColumns: int, romanNames_list: List[str]):
        """设置Excel表头"""
        # 每日统计点文件的表头（前三行）
        daily_stat_header1 = ['Date', self.currentDate.yyyy_str + "/" + self.currentDate.mm_str + "/" + self.currentDate.dd_str]
        daily_stat_header2 = [
            'Map sheet name',
            'Regular observation points finished',
            'Field points on revised route'
        ]
        daily_stat_header3 = [
            '', '', 'Added observation points',
            'Added Structure points, photo points, mineralization points'
        ]
        
        # 写入表头
        for col_num, value in enumerate(daily_stat_header1, start=1):
            sheet.cell(row=1, column=col_num, value=value)
        for col_num, value in enumerate(daily_stat_header2, start=1):
            sheet.cell(row=2, column=col_num, value=value)
        for col_num, value in enumerate(daily_stat_header3, start=1):
            sheet.cell(row=3, column=col_num, value=value)
        
        # 写入图幅名称
        for i, value in enumerate(romanNames_list, start=4):
            sheet.cell(row=i, column=1, value=value)
        
        # 写入表尾
        daily_stat_footer = ['Today', '', '', '']
        total_Point_Num_footer = ['TOTAL (Group 3)', '', '', '']
        
        for col_num, value in enumerate(daily_stat_footer, start=1):
            sheet.cell(row=maxTableRows-1, column=col_num, value=value)
        for col_num, value in enumerate(total_Point_Num_footer, start=1):
            sheet.cell(row=maxTableRows, column=col_num, value=value)

    def _setup_excel_styles(self, sheet, maxTableRows: int, maxTableColumns: int):
        """设置Excel样式"""
        # 设置字体
        font_header = Font(name='Calibri', size=12, bold=True)
        font = Font(name='Calibri', size=11)
        
        # 设置边框
        border = Border(
            left=Side(border_style='thin'),
            right=Side(border_style='thin'),
            top=Side(border_style='thin'),
            bottom=Side(border_style='thin')
        )
        
        # 应用样式
        for row in range(1, maxTableRows + 1):
            for col in range(1, maxTableColumns + 1):
                cell = sheet.cell(row=row, column=col)
                cell.border = border
                
                # 设置字体
                if row in [1, 2, 3, maxTableRows-1, maxTableRows]:
                    cell.font = font_header
                else:
                    cell.font = font
        
        # 设置对齐
        center_aligned = Alignment(horizontal='center', vertical='center')
        for row in range(1, maxTableRows + 1):
            for col in range(1, maxTableColumns + 1):
                sheet.cell(row=row, column=col).alignment = center_aligned
        
        # 调整列宽
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    def _setup_excel_data(self, sheet, maxTableRows: int):
        """设置Excel数据和公式"""
        # 填充实际数据到Excel表格
        self._fill_excel_data(sheet, maxTableRows)
        
        # 设置合计行的公式
        sheet.cell(row=maxTableRows-1, column=2).value = f"=SUM(B4:B{maxTableRows-2})"
        sheet.cell(row=maxTableRows-1, column=3).value = f"=SUM(C4:C{maxTableRows-2})"
        sheet.cell(row=maxTableRows-1, column=4).value = f"=SUM(D4:D{maxTableRows-2})"

        # 在倒数第二行（合计行）写入当日新增数量
        sheet.cell(row=maxTableRows-1, column=2, value=self.totalDaiyIncreasePointNum) 

        # 在最后一行（TOTAL行）写入总数
        sheet.cell(row=maxTableRows, column=2, value=self.totalPointNum)             # 累计总数

        # 合并单元格
        sheet.merge_cells('B1:D1')
        sheet.merge_cells('C2:D2')
        sheet.merge_cells('A2:A3')
        sheet.merge_cells('B2:B3')

    def _fill_excel_data(self, sheet, maxTableRows: int):
        """填充实际数据到Excel表格"""
        try:
            # 获取数据字典
            daily_increased = self.dailyIncreasedPoints
            daily_finished = self.dailyFinishedPoints
            daily_plans = self.DailyPlans
            
            # 按序号排序的图幅列表
            sorted_mapsheets = sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)
            
            # 从第4行开始填充数据 (前3行是表头)
            current_row = 4
            
            for mapsheet in sorted_mapsheets:
                roman_name = mapsheet.romanName
                
                # 第1列：图幅名称
                sheet.cell(row=current_row, column=1, value=roman_name)
                
                # 第2列：当日新增点数 (如果为0显示空值)
                increased_points = daily_increased.get(roman_name, 0)
                sheet.cell(row=current_row, column=2, 
                          value=increased_points if increased_points > 0 else None)
                
                # 第3列：当日新增线路/结构点等 (暂时留空，可以后续扩展)
                sheet.cell(row=current_row, column=3, value=None)
                
                # 第4列：累计完成点数 (如果为0显示空值)
                # finished_points = daily_finished.get(roman_name, 0)
                # sheet.cell(row=current_row, column=4, 
                #           value=finished_points if finished_points > 0 else None)
                
                current_row += 1
                
                # 防止超出表格范围
                if current_row >= maxTableRows - 1:
                    break
                    
            logger.info(f"成功填充 {current_row-4} 行数据到Excel表格")
            
        except Exception as e:
            logger.error(f"填充Excel数据失败: {e}")
            raise

    def write_completed_data_to_statistics_excel(self, target_excel_path: str) -> bool:
        """
        将当日新增的数据列写入指定的统计Excel文件
        
        Args:
            target_excel_path: 目标Excel文件路径，例如 "D:\\RouteDesign\\Daily_statistics_details_for_Group_3.2.xlsx"
            
        Returns:
            bool: 写入成功返回True，失败返回False
        """
        try:
            # 检查目标文件是否存在
            if not os.path.exists(target_excel_path):
                logger.error(f"目标Excel文件不存在: {target_excel_path}")
                return False
            
            # 加载现有工作簿
            wb = load_workbook(target_excel_path)
            
            # 使用"总表"工作表
            if "总表" in wb.sheetnames:
                ws = wb["总表"]
            else:
                ws = wb.active
                logger.warning("未找到'总表'工作表，使用默认工作表")
            
            # 获取当前日期的新增数据（而不是累计完成数据）
            daily_increased = self.dailyIncreasedPoints
            
            # 查找日期所在的列
            target_col = self._find_date_row_in_excel(ws)
            if target_col is None:
                logger.error(f"在Excel中未找到日期 {self.currentDate} 对应的列")
                return False
            
            self._fill_increased_data_to_col(ws, target_col)
            
            # 写入新增数据到相应的行
            # self._write_increased_data_to_col(ws, target_col, daily_increased)
            
            # 保存文件
            wb.save(target_excel_path)
            logger.info(f"成功将当日新增数据写入Excel文件: {target_excel_path}")
            return True
            
        except Exception as e:
            logger.error(f"写入当日新增数据到Excel文件失败: {e}")
            return False

    def _fill_increased_data_to_col(self, ws, target_col):
        """
        Purpose: Fill in the increased data for the specified column
        """
        """填充实际数据到Excel表格"""
        try:
            # 获取数据字典
            daily_increased = self.dailyIncreasedPoints
            
            # 按序号排序的图幅列表
            sorted_mapsheets = sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)
            
            # 从第3行开始填充数据 (前2行是表头)
            current_row = 3

            for mapsheet in sorted_mapsheets:
                roman_name = mapsheet.romanName
            
                # 当日新增点数 (如果为0显示空值)
                increased_points = daily_increased.get(roman_name, 0)
                ws.cell(row=current_row, column=target_col, 
                          value=increased_points if increased_points > 0 else None)
                
                current_row += 1
                    
            logger.info(f"成功填充 {current_row-3} 行数据到Daily statics Excel表格")
        
        except Exception as e:
            logger.error(f"写入当日新增数据到Daily statics Excel表格失败: {e}")
            return False
    
    def _find_date_row_in_excel(self, worksheet) -> Optional[int]:
        """
        在Excel工作表中查找当前日期对应的行
        
        Args:
            worksheet: openpyxl工作表对象
            
        Returns:
            int or None: 找到的行号，未找到返回None
        """
        try:
            # 在第一行查找日期列
            target_date = self.currentDate.date_datetime  # datetime对象
            
            # 检查第一行的日期列（从第9列开始，基于Excel结构分析）
            for col in range(9, min(worksheet.max_column + 1, 110)):  # 扩展搜索范围
                cell_value = worksheet.cell(row=1, column=col).value
                if cell_value:
                    # 如果是datetime对象，直接比较日期
                    if hasattr(cell_value, 'date'):
                        if cell_value.date() == target_date.date():
                            logger.info(f"在Excel第1行第{col}列找到匹配日期: {cell_value}")
                            return col  # 返回列号而不是行号
                    # 如果是字符串，尝试解析
                    elif isinstance(cell_value, str):
                        cell_str = cell_value.strip()
                        # 可能的日期格式
                        possible_date_formats = [
                            self.currentDate.yyyymmdd_str,  # "20250831"
                            f"{self.currentDate.yyyy_str}-{self.currentDate.mm_str}-{self.currentDate.dd_str}",  # "2025-08-31"
                            f"{self.currentDate.yyyy_str}/{self.currentDate.mm_str}/{self.currentDate.dd_str}",  # "2025/08/31"
                        ]
                        
                        if any(date_format in cell_str for date_format in possible_date_formats):
                            logger.info(f"在Excel第1行第{col}列找到日期字符串: {cell_str}")
                            return col  # 返回列号
            
            logger.warning(f"在Excel第1行中未找到日期 {target_date.date()} 对应的列")
            return None
            
        except Exception as e:
            logger.error(f"查找日期列失败: {e}")
            return None

    def onScreenDisplay(self) -> None:
        """在屏幕上显示统计信息 - 使用统一显示管理器"""
        from display import CollectionDisplay
        
        # 委托给CollectionDisplay处理，传递predict参数
        CollectionDisplay.show_statistics(self, enable_predict=self.enable_predict)
        
        # 额外显示进度估算结果
        self.display_progress_estimation()

    def __str__(self) -> str:
        """字符串表示"""
        try:
            return (
                f"当前日期文件集合\n"
                f"{'='*40}\n"
                f"日期: {self.currentDate}\n"
                f"总文件数: {len(self.currentDateFiles)}\n"
                f"总点数: {self.totalPointNum:,}\n"
                f"日增点数: {self.totalDaiyIncreasePointNum:,}\n"
                f"总路线数: {self.totalRoutesNum:,}\n"
                f"日增路线数: {self.totalDaiyIncreaseRouteNum:,}\n"
                f"计划数: {self.totalDailyPlanNum}\n"
                f"错误数: {len(self.errorMsg)}"
            )
        except Exception as e:
            return f"CurrentDateFiles(日期={self.currentDate}, 错误={e})"

    def __repr__(self) -> str:
        """开发者友好的字符串表示"""
        return (
            f"CurrentDateFiles(currentDate={self.currentDate!r}, "
            f"files_count={len(self.currentDateFiles)}, "
            f"points={self.totalPointNum}, "
            f"daily_increase={self.totalDaiyIncreasePointNum})"
        )
