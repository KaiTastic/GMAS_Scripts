"""
Excel数据连接器模块

从GMAS统计Excel文件读取历史数据，作为核心数据访问层
"""

import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from openpyxl import load_workbook

from ..data_models.date_types import DateType
from config.config_manager import ConfigManager

logger = logging.getLogger(__name__)


class ExcelDataConnector:
    """Excel数据连接器类，从GMAS统计Excel文件读取历史数据"""

    def __init__(self, workspace_path: str = None):
        """
        初始化Excel数据连接器
        
        Args:
            workspace_path: 工作空间路径
        """
        self.workspace_path = workspace_path or os.getcwd()
        self.config_manager = ConfigManager()
        self.excel_file_path: Optional[str] = None
        self.excel_data: Optional[pd.DataFrame] = None
        
        # 从配置获取Excel文件路径
        self._initialize_excel_path()
        
    def _initialize_excel_path(self) -> None:
        """初始化Excel文件路径"""
        try:
            # 直接从主配置文件获取Excel文件路径
            self.excel_file_path = self.config_manager.get_daily_details_file()
            logger.info(f"从主配置获取Excel文件路径: {self.excel_file_path}")
            
            # 验证文件是否存在
            if not os.path.exists(self.excel_file_path):
                logger.warning(f"Excel文件不存在: {self.excel_file_path}")
                self._try_backup_path()
                
        except Exception as e:
            logger.error(f"初始化Excel路径时发生错误: {e}")
            # 使用备用路径
            self._try_backup_path()
    
    def _try_backup_path(self):
        """尝试使用备用路径"""
        try:
            workspace = self.config_manager.get_workspace()
            filename = self.config_manager.get_main_setting('reports.statistics.daily_details_file_name', 
                                                           'Daily_statistics_details_for_Group_3.2.xlsx')
            
            # 尝试在工作空间根目录
            primary_path = os.path.join(workspace, filename)
            if os.path.exists(primary_path):
                self.excel_file_path = primary_path
                logger.info(f"使用工作空间路径: {self.excel_file_path}")
                return
            
            # 尝试备用目录
            backup_dir = os.path.join(workspace, "backup")
            backup_path = os.path.join(backup_dir, filename)
            if os.path.exists(backup_path):
                self.excel_file_path = backup_path
                logger.info(f"使用备用路径: {self.excel_file_path}")
                return
            
            # 最后使用主路径（即使文件不存在）
            self.excel_file_path = primary_path
            logger.warning(f"文件不存在，但设置路径为: {self.excel_file_path}")
            
        except Exception as e:
            logger.error(f"设置备用路径失败: {e}")
            self.excel_file_path = None
    
    def load_excel_data(self) -> bool:
        """
        加载Excel数据
        
        Returns:
            bool: 是否成功加载
        """
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            logger.error(f"Excel文件不存在: {self.excel_file_path}")
            return False
        
        try:
            # 使用openpyxl读取Excel文件
            wb = load_workbook(self.excel_file_path, data_only=True)
            
            # 尝试使用"总表"工作表
            if "总表" in wb.sheetnames:
                ws = wb["总表"]
            else:
                ws = wb.active
                logger.warning("未找到'总表'工作表，使用默认工作表")
            
            # 将工作表数据转换为DataFrame
            data = []
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 读取所有数据
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(row)
            
            # 创建DataFrame
            if data:
                # 使用第一行作为列名，其余行作为数据
                if len(data) > 1:
                    columns = [str(col) if col is not None else f"Col_{i}" for i, col in enumerate(data[0])]
                    self.excel_data = pd.DataFrame(data[1:], columns=columns)
                    logger.info(f"成功加载Excel数据: {len(data)-1}行 x {len(columns)}列")
                    logger.info(f"列名: {columns}")
                else:
                    # 只有一行数据的情况
                    self.excel_data = pd.DataFrame(data)
                    logger.info(f"成功加载Excel数据: {len(data)}行")
                return True
            else:
                logger.error("Excel文件为空")
                return False
                
        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
            return False
    
    def extract_mapsheet_metadata(self) -> Dict[str, Dict[str, Any]]:
        """
        提取图幅元数据信息，包括目标点数、当前总点数等
        
        Returns:
            Dict[str, Dict[str, Any]]: 图幅元数据字典
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                logger.error("无法加载Excel数据")
                return {}
        
        try:
            mapsheet_metadata = {}
            
            # 查找关键列的索引
            columns = self.excel_data.columns
            column_indices = {}
            
            # 寻找关键列
            for i, col in enumerate(columns):
                col_str = str(col).strip()
                if 'sheet name' in col_str.lower() or '图幅名' in col_str:
                    column_indices['sheet_name'] = i
                elif 'team' in col_str.lower() or '团队' in col_str:
                    column_indices['team'] = i
                elif 'total' in col_str.lower() or '总计' in col_str:
                    column_indices['total'] = i
                elif 'adjusted' in col_str.lower() and 'num' in col_str.lower():
                    column_indices['adjusted_num'] = i
                elif 'percentage' in col_str.lower() or '百分比' in col_str:
                    column_indices['percentage'] = i
                elif 'person' in col_str.lower() and 'charge' in col_str.lower():
                    column_indices['person_in_charge'] = i
            
            logger.info(f"找到的关键列索引: {column_indices}")
            
            # 遍历每一行提取图幅信息
            for row_idx in range(len(self.excel_data)):
                try:
                    # 获取图幅标识
                    mapsheet_id = None
                    
                    # 优先使用图幅名称
                    if 'sheet_name' in column_indices:
                        sheet_name = self.excel_data.iloc[row_idx, column_indices['sheet_name']]
                        if pd.notna(sheet_name) and str(sheet_name).strip():
                            mapsheet_id = str(sheet_name).strip()
                    
                    # 备用方案：使用团队信息
                    if not mapsheet_id and 'team' in column_indices:
                        team_name = self.excel_data.iloc[row_idx, column_indices['team']]
                        if pd.notna(team_name) and str(team_name).strip():
                            mapsheet_id = str(team_name).strip()
                    
                    # 如果都没有，跳过这行
                    if not mapsheet_id:
                        continue
                    
                    # 提取元数据
                    metadata = {
                        'mapsheet_id': mapsheet_id,
                        'row_index': row_idx
                    }
                    
                    # 提取总点数
                    if 'total' in column_indices:
                        total_value = self.excel_data.iloc[row_idx, column_indices['total']]
                        try:
                            metadata['total_points'] = float(total_value) if pd.notna(total_value) else 0
                        except (ValueError, TypeError):
                            metadata['total_points'] = 0
                    else:
                        metadata['total_points'] = 0
                    
                    # 提取目标点数 (Adjusted Num)
                    if 'adjusted_num' in column_indices:
                        adjusted_value = self.excel_data.iloc[row_idx, column_indices['adjusted_num']]
                        try:
                            metadata['target_points'] = float(adjusted_value) if pd.notna(adjusted_value) else 1000
                        except (ValueError, TypeError):
                            metadata['target_points'] = 1000
                    else:
                        metadata['target_points'] = 1000  # 默认值
                    
                    # 提取完成百分比
                    if 'percentage' in column_indices:
                        percentage_value = self.excel_data.iloc[row_idx, column_indices['percentage']]
                        try:
                            metadata['completion_percentage'] = float(percentage_value) if pd.notna(percentage_value) else 0
                        except (ValueError, TypeError):
                            metadata['completion_percentage'] = 0
                    else:
                        # 计算完成百分比
                        if metadata['target_points'] > 0:
                            metadata['completion_percentage'] = (metadata['total_points'] / metadata['target_points']) * 100
                        else:
                            metadata['completion_percentage'] = 0
                    
                    # 提取负责人信息
                    if 'person_in_charge' in column_indices:
                        person_value = self.excel_data.iloc[row_idx, column_indices['person_in_charge']]
                        metadata['person_in_charge'] = str(person_value).strip() if pd.notna(person_value) else ""
                    else:
                        metadata['person_in_charge'] = ""
                    
                    # 计算剩余点数
                    metadata['remaining_points'] = max(0, metadata['target_points'] - metadata['total_points'])
                    
                    # 判断完成状态
                    metadata['status'] = 'completed' if metadata['total_points'] >= metadata['target_points'] else 'in_progress'
                    
                    mapsheet_metadata[mapsheet_id] = metadata
                    
                except Exception as e:
                    logger.debug(f"处理第{row_idx}行时出错: {e}")
                    continue
            
            logger.info(f"成功提取 {len(mapsheet_metadata)} 个图幅的元数据")
            
            # 打印元数据摘要
            if mapsheet_metadata:
                logger.info("图幅元数据摘要:")
                for mapsheet_id, meta in list(mapsheet_metadata.items())[:5]:  # 只打印前5个
                    logger.info(f"  {mapsheet_id}: 总计{meta['total_points']:.0f}, 目标{meta['target_points']:.0f}, "
                              f"完成{meta['completion_percentage']:.1f}%")
            
            return mapsheet_metadata
            
        except Exception as e:
            logger.error(f"提取图幅元数据失败: {e}")
            return {}
    
    def extract_mapsheet_historical_data(self, start_date: DateType, end_date: DateType = None) -> Dict[str, List[Dict]]:
        """
        提取每个图幅的历史数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期，默认为今天
            
        Returns:
            Dict[str, List[Dict]]: 每个图幅的历史数据，格式为 {mapsheet_name: [daily_records]}
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                return {}
        
        if end_date is None:
            end_date = DateType(datetime.now())
        
        try:
            mapsheet_data = {}
            
            # 找到图幅名称列
            if 'Sheet Name' not in self.excel_data.columns:
                logger.warning("未找到'Sheet Name'列")
                return {}
            
            # 找到日期列（从第8列开始是日期列）
            date_columns = []
            for col_name in self.excel_data.columns[8:]:  # 日期从第8列开始
                if isinstance(col_name, str) and '2025-' in col_name:  # 日期列
                    try:
                        # 验证是否可以解析为日期
                        datetime.strptime(col_name, '%Y-%m-%d %H:%M:%S')
                        date_columns.append(col_name)
                    except:
                        continue
            
            if not date_columns:
                logger.warning("未找到有效的日期列")
                return {}
            
            logger.info(f"开始提取图幅历史数据 - 找到 {len(date_columns)} 个日期列")
            
            # 遍历每一行（每行代表一个图幅的数据）
            for row_idx in range(len(self.excel_data)):
                try:
                    # 获取图幅名称
                    mapsheet_name = self.excel_data.iloc[row_idx, self.excel_data.columns.get_loc('Sheet Name')]
                    
                    if pd.isna(mapsheet_name) or str(mapsheet_name).strip() == '':
                        continue
                    
                    mapsheet_name = str(mapsheet_name).strip()
                    
                    if mapsheet_name not in mapsheet_data:
                        mapsheet_data[mapsheet_name] = []
                    
                    # 遍历每个日期列，提取该图幅在该日期的数据
                    for date_col in date_columns:
                        try:
                            # 解析日期
                            date_obj = datetime.strptime(date_col, '%Y-%m-%d %H:%M:%S')
                            date_type = DateType(date_obj)
                            
                            # 检查是否在指定范围内
                            if not (start_date.date_datetime <= date_type.date_datetime <= end_date.date_datetime):
                                continue
                            
                            # 获取该日期的点数
                            col_idx = self.excel_data.columns.get_loc(date_col)
                            points_value = self.excel_data.iloc[row_idx, col_idx]
                            
                            points = 0
                            if not pd.isna(points_value):
                                try:
                                    points = int(float(points_value))
                                    if points < 0:
                                        points = 0
                                except (ValueError, TypeError):
                                    points = 0
                            
                            # 创建该图幅在该日期的记录
                            record = {
                                'date': date_type.yyyymmdd_str,
                                'date_obj': date_type.date_datetime,
                                'daily_points': points,
                                'workday': date_type.date_datetime.weekday() < 5
                            }
                            
                            mapsheet_data[mapsheet_name].append(record)
                            
                        except Exception as e:
                            logger.debug(f"解析图幅 {mapsheet_name} 在 {date_col} 的数据失败: {e}")
                            continue
                        
                except Exception as e:
                    logger.debug(f"解析第{row_idx}行数据失败: {e}")
                    continue
            
            # 为每个图幅计算累计完成量
            for mapsheet_name in mapsheet_data:
                records = mapsheet_data[mapsheet_name]
                if records:
                    # 按日期排序
                    records.sort(key=lambda x: x['date'])
                    
                    # 计算累计完成量
                    cumulative = 0
                    for record in records:
                        cumulative += record['daily_points']
                        record['cumulative_points'] = cumulative
            
            logger.info(f"成功提取 {len(mapsheet_data)} 个图幅的历史数据")
            return mapsheet_data
            
        except Exception as e:
            logger.error(f"提取图幅历史数据失败: {e}")
            return {}
    
    def get_project_timeline(self) -> Dict[str, Any]:
        """
        获取项目时间线信息
        
        Returns:
            Dict: 包含时间线信息的字典
        """
        try:
            if not self.load_excel_data():
                return {}
            
            # 从日期列获取时间线信息
            date_columns = []
            for col_name in self.excel_data.columns[8:]:  # 日期从第8列开始
                if isinstance(col_name, str) and '2025-' in col_name:  # 日期列
                    try:
                        date_obj = datetime.strptime(col_name, '%Y-%m-%d %H:%M:%S')
                        date_columns.append(date_obj)
                    except:
                        continue
            
            if not date_columns:
                return {}
            
            date_columns.sort()
            start_date = date_columns[0]
            latest_date = date_columns[-1]
            
            return {
                'start_date': start_date,
                'latest_date': latest_date,
                'total_days': (latest_date - start_date).days + 1,
                'data_points': len(date_columns)
            }
            
        except Exception as e:
            logger.error(f"获取项目时间线失败: {e}")
            return {}
    
    def get_current_project_status(self) -> Dict[str, Any]:
        """
        获取当前项目状态
        
        Returns:
            Dict: 包含当前状态的字典
        """
        try:
            metadata = self.extract_mapsheet_metadata()
            
            if not metadata:
                return {}
            
            total_points = sum(meta['total_points'] for meta in metadata.values())
            total_target = sum(meta['target_points'] for meta in metadata.values())
            
            return {
                'current_points': total_points,
                'target_points': total_target,
                'total_mapsheets': len(metadata),
                'completion_percentage': (total_points / total_target * 100) if total_target > 0 else 0,
                'completed_mapsheets': sum(1 for meta in metadata.values() if meta['status'] == 'completed'),
                'latest_date': datetime.now().strftime('%Y-%m-%d')
            }
            
        except Exception as e:
            logger.error(f"获取当前项目状态失败: {e}")
            return {}
    
    def get_connection_status(self) -> Dict[str, Any]:
        """获取连接状态摘要"""
        file_info = self._get_file_info()
        validation = self._validate_data_structure()
        
        return {
            'excel_file': file_info,
            'data_loaded': self.excel_data is not None,
            'data_validation': validation,
            'config_source': {
                'workspace': self.config_manager.get_workspace(),
                'mapsheet_range': self.config_manager.get_mapsheet_range()
            }
        }
    
    def _get_file_info(self) -> Dict[str, Any]:
        """获取Excel文件信息"""
        if not self.excel_file_path:
            return {'exists': False, 'path': None}
        
        try:
            if os.path.exists(self.excel_file_path):
                stat = os.stat(self.excel_file_path)
                return {
                    'exists': True,
                    'path': self.excel_file_path,
                    'size': stat.st_size,
                    'modified': datetime.fromtimestamp(stat.st_mtime),
                    'size_mb': round(stat.st_size / (1024 * 1024), 2)
                }
            else:
                return {
                    'exists': False,
                    'path': self.excel_file_path
                }
        except Exception as e:
            logger.error(f"获取文件信息失败: {e}")
            return {'exists': False, 'path': self.excel_file_path, 'error': str(e)}
    
    def _validate_data_structure(self) -> Dict[str, Any]:
        """验证数据结构是否符合预期"""
        if self.excel_data is None:
            return {
                'valid': False,
                'issues': ['Excel数据未加载']
            }
        
        issues = []
        
        # 检查数据是否为空
        if self.excel_data.empty:
            issues.append('Excel数据为空')
        
        # 检查是否有足够的列
        if len(self.excel_data.columns) < 8:
            issues.append('Excel数据列数不足（至少需要8列）')
        
        return {
            'valid': len(issues) == 0,
            'issues': issues,
            'data_shape': self.excel_data.shape if self.excel_data is not None else None,
            'columns': list(self.excel_data.columns) if self.excel_data is not None else None
        }
