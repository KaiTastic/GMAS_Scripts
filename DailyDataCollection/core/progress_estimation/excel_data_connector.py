"""
真实数据连接器模块

连接进度估算模块与实际的GMAS Excel数据文件
"""

import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from openpyxl import load_workbook

from ..data_models.date_types import DateType
from config.config_manager import ConfigManager

logger = logging.getLogger(__name__)


class ExcelDataConnector:
    """Excel数据连接器类，从GMAS统计Excel文件读取历史数据"""

    def __init__(self, workspace_path: str = None):
        """
        初始化Excel数据连接器
        
        Args:
            workspace_path: 工作空间路径
        """
        self.workspace_path = workspace_path or os.getcwd()
        self.config_manager = ConfigManager()
        self.excel_file_path: Optional[str] = None
        self.excel_data: Optional[pd.DataFrame] = None
        
        # 从配置获取Excel文件路径
        self._initialize_excel_path()
        
    def _initialize_excel_path(self) -> None:
        """初始化Excel文件路径"""
        try:
            # 从配置文件获取统计详情文件路径
            stats_config = self.config_manager.get('reports.statistics')
            if stats_config and 'daily_details_file' in stats_config:
                file_path_template = stats_config['daily_details_file']
                
                # 构建解析上下文
                context = {
                    'stats_config': stats_config
                }
                
                # 使用配置管理器的模板解析功能
                self.excel_file_path = self.config_manager.resolve_path_template(file_path_template, context)
                logger.info(f"Excel文件路径: {self.excel_file_path}")
            else:
                # 使用默认路径
                workspace = self.config_manager.get('system.workspace', 'D:\\RouteDesign')
                filename = "Daily_statistics_details_for_Group_3.2.xlsx"
                self.excel_file_path = os.path.join(workspace, filename)
                logger.warning(f"配置中未找到Excel路径，使用默认路径: {self.excel_file_path}")
                
        except Exception as e:
            # 异常情况使用默认路径
            workspace = self.config_manager.get('system.workspace', 'D:\\RouteDesign')
            filename = "Daily_statistics_details_for_Group_3.2.xlsx"
            self.excel_file_path = os.path.join(workspace, filename)
            logger.error(f"初始化Excel路径时发生错误: {e}，使用默认路径: {self.excel_file_path}")
    
    def load_excel_data(self) -> bool:
        """
        加载Excel数据
        
        Returns:
            bool: 是否成功加载
        """
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            logger.error(f"Excel文件不存在: {self.excel_file_path}")
            return False
        
        try:
            # 使用openpyxl读取Excel文件
            wb = load_workbook(self.excel_file_path, data_only=True)
            
            # 尝试使用"总表"工作表
            if "总表" in wb.sheetnames:
                ws = wb["总表"]
            else:
                ws = wb.active
                logger.warning("未找到'总表'工作表，使用默认工作表")
            
            # 将工作表数据转换为DataFrame
            data = []
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 读取所有数据
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(row)
            
            # 创建DataFrame
            if data:
                # 使用第一行作为列名，其余行作为数据
                if len(data) > 1:
                    columns = [str(col) if col is not None else f"Col_{i}" for i, col in enumerate(data[0])]
                    self.excel_data = pd.DataFrame(data[1:], columns=columns)
                    logger.info(f"成功加载Excel数据: {len(data)-1}行 x {len(columns)}列")
                    logger.info(f"列名: {columns}")
                else:
                    # 只有一行数据的情况
                    self.excel_data = pd.DataFrame(data)
                    logger.info(f"成功加载Excel数据: {len(data)}行")
                return True
            else:
                logger.error("Excel文件为空")
                return False
                
        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
            return False
    
    def extract_historical_data(self, start_date: DateType, end_date: DateType = None) -> List[Dict[str, Any]]:
        """
        从Excel数据中提取历史数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期，默认为今天
            
        Returns:
            List[Dict]: 历史数据列表
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                return []
        
        if end_date is None:
            end_date = DateType(datetime.now())
        
        try:
            historical_records = []
            
            # 分析Excel结构：第一列是日期，其他列是图幅
            if len(self.excel_data.columns) < 2:
                logger.error("Excel数据列数不足")
                return []
            
            # 获取图幅列名（跳过第一列日期和最后一列总计）
            mapsheet_columns = []
            for col_name in self.excel_data.columns[1:-1]:  # 跳过第一列和最后一列
                if 'H50E' in str(col_name):  # 图幅名称格式
                    mapsheet_columns.append(col_name)
            
            if not mapsheet_columns:
                logger.warning("未找到有效的图幅列")
                return []
            
            logger.info(f"找到 {len(mapsheet_columns)} 个图幅列: {mapsheet_columns[:3]}...")
            
            # 遍历每一行（每行代表一个日期的数据）
            for row_idx in range(len(self.excel_data)):
                try:
                    # 获取日期（第一列）
                    date_value = self.excel_data.iloc[row_idx, 0]
                    
                    if pd.isna(date_value):
                        continue
                    
                    # 解析日期
                    date_str = str(int(date_value)) if isinstance(date_value, (int, float)) else str(date_value).strip()
                    
                    # 转换为DateType对象
                    date_obj = self._parse_date_string(date_str)
                    if not date_obj:
                        continue
                    
                    date_type = DateType(date_obj)
                    
                    # 检查是否在指定范围内
                    if not (start_date.date_datetime <= date_type.date_datetime <= end_date.date_datetime):
                        continue
                    
                    # 提取该日期的点数数据
                    daily_points = {}
                    total_points = 0
                    
                    # 遍历图幅列
                    for mapsheet_name in mapsheet_columns:
                        try:
                            col_idx = self.excel_data.columns.get_loc(mapsheet_name)
                            points_value = self.excel_data.iloc[row_idx, col_idx]
                            
                            points = 0
                            if not pd.isna(points_value):
                                try:
                                    points = int(float(points_value))
                                    if points < 0:
                                        points = 0
                                except (ValueError, TypeError):
                                    points = 0
                            
                            daily_points[mapsheet_name] = points
                            total_points += points
                            
                        except Exception as e:
                            logger.debug(f"解析图幅 {mapsheet_name} 数据失败: {e}")
                            daily_points[mapsheet_name] = 0
                    
                    if total_points > 0:  # 只记录有数据的日期
                        # 估算活跃团队数量（基于有数据的图幅数量）
                        active_mapsheets = len([p for p in daily_points.values() if p > 0])
                        teams_active = max(1, active_mapsheets // 3)  # 假设每3个图幅一个团队
                        
                        record = {
                            'date': date_type.yyyymmdd_str,
                            'completed_points': total_points,
                            'teams_active': teams_active,
                            'workday': date_type.date_datetime.weekday() < 5,
                            'mapsheet_details': daily_points
                        }
                        
                        historical_records.append(record)
                        
                except Exception as e:
                    logger.debug(f"解析第{row_idx}行数据失败: {e}")
                    continue
            
            # 按日期排序
            historical_records.sort(key=lambda x: x['date'])
            
            # 计算累计完成量
            cumulative = 0
            for record in historical_records:
                cumulative += record['completed_points']
                record['cumulative_points'] = cumulative
            
            logger.info(f"成功提取 {len(historical_records)} 天的历史数据，总点数: {cumulative}")
            return historical_records
            
        except Exception as e:
            logger.error(f"提取历史数据失败: {e}")
            return []
    
    def _find_date_columns(self) -> Dict[str, int]:
        """
        查找日期列
        
        Returns:
            Dict[str, int]: 日期字符串到列索引的映射
        """
        date_columns = {}
        
        if self.excel_data is None or len(self.excel_data) < 1:
            return date_columns
        
        try:
            # 检查第一行的日期列（从第9列开始）
            first_row = self.excel_data.iloc[0]
            
            for col_idx in range(8, len(first_row)):  # 从第9列开始 (索引8)
                cell_value = first_row.iloc[col_idx]
                
                if pd.isna(cell_value):
                    continue
                
                # 尝试解析日期
                date_str = str(cell_value).strip()
                if self._is_date_string(date_str):
                    date_columns[date_str] = col_idx
            
            logger.info(f"找到 {len(date_columns)} 个日期列")
            return date_columns
            
        except Exception as e:
            logger.error(f"查找日期列失败: {e}")
            return {}
    
    def _find_mapsheet_rows(self) -> Dict[str, int]:
        """
        查找图幅行
        
        Returns:
            Dict[str, int]: 图幅名称到行索引的映射
        """
        mapsheet_rows = {}
        
        if self.excel_data is None or len(self.excel_data) < 3:
            return mapsheet_rows
        
        try:
            # 从第3行开始查找图幅名称（前两行是表头）
            for row_idx in range(2, len(self.excel_data)):
                first_cell = self.excel_data.iloc[row_idx, 0]  # 第一列
                
                if pd.isna(first_cell):
                    continue
                
                mapsheet_name = str(first_cell).strip()
                if mapsheet_name and not mapsheet_name.lower().startswith(('today', 'total', '合计')):
                    mapsheet_rows[mapsheet_name] = row_idx
            
            logger.info(f"找到 {len(mapsheet_rows)} 个图幅行")
            return mapsheet_rows
            
        except Exception as e:
            logger.error(f"查找图幅行失败: {e}")
            return {}
    
    def _extract_daily_points(self, col_idx: int, mapsheet_rows: Dict[str, int]) -> Dict[str, int]:
        """
        提取指定列的每日点数数据
        
        Args:
            col_idx: 列索引
            mapsheet_rows: 图幅行映射
            
        Returns:
            Dict[str, int]: 图幅名称到点数的映射
        """
        daily_points = {}
        
        try:
            for mapsheet_name, row_idx in mapsheet_rows.items():
                if row_idx < len(self.excel_data) and col_idx < len(self.excel_data.columns):
                    cell_value = self.excel_data.iloc[row_idx, col_idx]
                    
                    # 转换为数值
                    points = 0
                    if not pd.isna(cell_value):
                        try:
                            points = int(float(cell_value))
                        except (ValueError, TypeError):
                            points = 0
                    
                    daily_points[mapsheet_name] = points
            
            return daily_points
            
        except Exception as e:
            logger.error(f"提取日点数数据失败: {e}")
            return {}
    
    def _parse_date_string(self, date_str: str) -> Optional[datetime]:
        """
        解析日期字符串
        
        Args:
            date_str: 日期字符串
            
        Returns:
            datetime对象或None
        """
        if not date_str:
            return None
        
        date_str = str(date_str).strip()
        
        try:
            # 尝试不同的日期格式
            formats = [
                '%Y%m%d',      # 20250810
                '%Y-%m-%d',    # 2025-08-10
                '%Y/%m/%d',    # 2025/08/10
                '%d/%m/%Y',    # 10/08/2025
                '%d-%m-%Y',    # 10-08-2025
            ]
            
            for fmt in formats:
                try:
                    return datetime.strptime(date_str, fmt)
                except ValueError:
                    continue
            
            # 如果都不匹配，尝试pandas解析
            return pd.to_datetime(date_str).to_pydatetime()
            
        except Exception as e:
            logger.debug(f"解析日期字符串失败: {date_str}, 错误: {e}")
            return None

    def _is_date_string(self, date_str: str) -> bool:
        """
        检查字符串是否是日期格式
        
        Args:
            date_str: 待检查的字符串
            
        Returns:
            bool: 是否是日期格式
        """
        # 常见的日期格式模式
        date_patterns = [
            r'^\d{4}-\d{1,2}-\d{1,2}$',      # YYYY-MM-DD
            r'^\d{4}/\d{1,2}/\d{1,2}$',      # YYYY/MM/DD
            r'^\d{8}$',                       # YYYYMMDD
            r'^\d{4}\.\d{1,2}\.\d{1,2}$',    # YYYY.MM.DD
        ]
        
        import re
        for pattern in date_patterns:
            if re.match(pattern, date_str):
                return True
        
        # 尝试解析为日期
        try:
            self._parse_date_string(date_str)
            return True
        except:
            return False
    
    def _parse_date_string(self, date_str: str) -> Optional[datetime]:
        """
        解析日期字符串
        
        Args:
            date_str: 日期字符串
            
        Returns:
            datetime: 解析后的日期对象，失败返回None
        """
        # 常见的日期格式
        date_formats = [
            '%Y-%m-%d',
            '%Y/%m/%d',
            '%Y%m%d',
            '%Y.%m.%d',
            '%m/%d/%Y',
            '%d/%m/%Y',
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt)
            except ValueError:
                continue
        
        # 尝试pandas的日期解析
        try:
            return pd.to_datetime(date_str).to_pydatetime()
        except:
            return None
    
    def get_current_project_status(self) -> Dict[str, Any]:
        """
        获取当前项目状态
        
        Returns:
            Dict: 包含当前状态的字典
        """
        try:
            if not self.load_excel_data():
                return {}
            
            # 获取最新数据
            date_columns = self._find_date_columns()
            mapsheet_rows = self._find_mapsheet_rows()
            
            if not date_columns or not mapsheet_rows:
                return {}
            
            # 找到最新的日期列
            latest_date_str = max(date_columns.keys(), key=lambda x: self._parse_date_string(x) or datetime.min)
            latest_col_idx = date_columns[latest_date_str]
            
            # 计算总完成点数
            daily_points = self._extract_daily_points(latest_col_idx, mapsheet_rows)
            current_points = sum(daily_points.values())
            
            # 估算目标点数（基于图幅数量）
            total_mapsheets = len(mapsheet_rows)
            estimated_target = total_mapsheets * 200  # 假设每个图幅平均200个点
            
            return {
                'current_points': current_points,
                'estimated_target': estimated_target,
                'total_mapsheets': total_mapsheets,
                'latest_date': latest_date_str,
                'mapsheet_details': daily_points
            }
            
        except Exception as e:
            logger.error(f"获取当前项目状态失败: {e}")
            return {}
    
    def get_project_timeline(self) -> Dict[str, Any]:
        """
        获取项目时间线信息
        
        Returns:
            Dict: 包含时间线信息的字典
        """
        try:
            if not self.load_excel_data():
                return {}
            
            date_columns = self._find_date_columns()
            if not date_columns:
                return {}
            
            # 解析所有日期
            dates = []
            for date_str in date_columns.keys():
                date_obj = self._parse_date_string(date_str)
                if date_obj:
                    dates.append(date_obj)
            
            if not dates:
                return {}
            
            dates.sort()
            start_date = dates[0]
            latest_date = dates[-1]
            
            return {
                'start_date': DateType(start_date),
                'latest_date': DateType(latest_date),
                'total_days': (latest_date - start_date).days + 1,
                'data_points': len(dates)
            }
            
        except Exception as e:
            logger.error(f"获取项目时间线失败: {e}")
            return {}
    
    def validate_data_availability(self, start_date: DateType, end_date: DateType) -> Dict[str, Any]:
        """
        验证指定日期范围内的数据可用性
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            Dict: 数据可用性验证结果
        """
        validation_result = {
            'total_days': 0,
            'available_days': 0,
            'missing_days': 0,
            'data_quality_score': 0.0,
            'issues': [],
            'recommendations': []
        }
        
        try:
            if not self.load_excel_data():
                validation_result['issues'].append("Excel文件无法加载")
                return validation_result
            
            # 计算总天数
            total_days = (end_date.date_datetime - start_date.date_datetime).days + 1
            validation_result['total_days'] = total_days
            
            # 提取实际可用的数据
            historical_data = self.extract_historical_data(start_date, end_date)
            available_days = len(historical_data)
            validation_result['available_days'] = available_days
            validation_result['missing_days'] = total_days - available_days
            
            # 计算数据质量分数
            if total_days > 0:
                availability_score = available_days / total_days
                
                # 检查数据一致性（每日点数是否合理）
                if historical_data:
                    daily_points = [record['completed_points'] for record in historical_data]
                    avg_points = sum(daily_points) / len(daily_points)
                    
                    # 检查是否有异常值（过高或过低的点数）
                    outliers = 0
                    for points in daily_points:
                        if points > avg_points * 3 or points < 0:
                            outliers += 1
                    
                    consistency_score = 1.0 - (outliers / len(daily_points)) if daily_points else 0.0
                    
                    # 综合分数
                    validation_result['data_quality_score'] = (availability_score * 0.7 + consistency_score * 0.3)
                else:
                    validation_result['data_quality_score'] = 0.0
            
            # 生成问题和建议
            if available_days > 0:
                availability_score = available_days / total_days
                if availability_score < 0.8:
                    validation_result['issues'].append(f"数据缺失较多 ({validation_result['missing_days']}天)")
                    validation_result['recommendations'].append("检查数据收集流程，确保数据完整性")
            
                if validation_result['data_quality_score'] < 0.7:
                    validation_result['issues'].append("数据质量偏低")
                    validation_result['recommendations'].append("检查数据输入的准确性和一致性")
            
            logger.info(f"数据验证完成: {available_days}/{total_days} 天可用，质量分数: {validation_result['data_quality_score']:.2f}")
            
        except Exception as e:
            logger.error(f"数据验证失败: {e}")
            validation_result['issues'].append(f"验证过程出错: {str(e)}")
        
        return validation_result
    
    def extract_mapsheet_historical_data(self, start_date: DateType, end_date: DateType = None) -> Dict[str, List[Dict]]:
        """
        提取每个图幅的历史数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期，默认为今天
            
        Returns:
            Dict[str, List[Dict]]: 每个图幅的历史数据，格式为 {mapsheet_name: [daily_records]}
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                return {}
        
        if end_date is None:
            end_date = DateType(datetime.now())
        
        try:
            mapsheet_data = {}
            
            # 获取图幅列名（跳过第一列日期和最后一列总计）
            mapsheet_columns = []
            for col_name in self.excel_data.columns[1:-1]:
                if 'H50E' in str(col_name):  # 图幅名称格式
                    mapsheet_columns.append(col_name)
                    mapsheet_data[col_name] = []
            
            if not mapsheet_columns:
                logger.warning("未找到有效的图幅列")
                return {}
            
            logger.info(f"开始提取 {len(mapsheet_columns)} 个图幅的历史数据")
            
            # 遍历每一行（每行代表一个日期的数据）
            for row_idx in range(len(self.excel_data)):
                try:
                    # 获取日期（第一列）
                    date_value = self.excel_data.iloc[row_idx, 0]
                    
                    if pd.isna(date_value):
                        continue
                    
                    # 解析日期
                    date_str = str(int(date_value)) if isinstance(date_value, (int, float)) else str(date_value).strip()
                    
                    # 转换为DateType对象
                    date_obj = self._parse_date_string(date_str)
                    if not date_obj:
                        continue
                    
                    date_type = DateType(date_obj)
                    
                    # 检查是否在指定范围内
                    if not (start_date.date_datetime <= date_type.date_datetime <= end_date.date_datetime):
                        continue
                    
                    # 为每个图幅提取数据
                    for mapsheet_name in mapsheet_columns:
                        try:
                            col_idx = self.excel_data.columns.get_loc(mapsheet_name)
                            points_value = self.excel_data.iloc[row_idx, col_idx]
                            
                            points = 0
                            if not pd.isna(points_value):
                                try:
                                    points = int(float(points_value))
                                    if points < 0:
                                        points = 0
                                except (ValueError, TypeError):
                                    points = 0
                            
                            # 创建该图幅在该日期的记录
                            record = {
                                'date': date_type.yyyymmdd_str,
                                'date_obj': date_type.date_datetime,
                                'daily_points': points,
                                'workday': date_type.date_datetime.weekday() < 5
                            }
                            
                            mapsheet_data[mapsheet_name].append(record)
                            
                        except Exception as e:
                            logger.debug(f"解析图幅 {mapsheet_name} 在 {date_str} 的数据失败: {e}")
                            continue
                        
                except Exception as e:
                    logger.debug(f"解析第{row_idx}行数据失败: {e}")
                    continue
            
            # 为每个图幅计算累计完成量
            for mapsheet_name in mapsheet_data:
                records = mapsheet_data[mapsheet_name]
                if records:
                    # 按日期排序
                    records.sort(key=lambda x: x['date'])
                    
                    # 计算累计完成量
                    cumulative = 0
                    for record in records:
                        cumulative += record['daily_points']
                        record['cumulative_points'] = cumulative
            
            logger.info(f"成功提取 {len(mapsheet_data)} 个图幅的历史数据")
            return mapsheet_data
            
        except Exception as e:
            logger.error(f"提取图幅历史数据失败: {e}")
            return {}
    
    def get_mapsheet_current_status(self) -> Dict[str, Dict]:
        """
        获取每个图幅的当前状态
        
        Returns:
            Dict[str, Dict]: 每个图幅的当前状态信息
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                return {}
        
        try:
            mapsheet_status = {}
            
            # 获取图幅列名（跳过第一列日期和最后一列总计）
            mapsheet_columns = []
            for col_name in self.excel_data.columns[1:-1]:
                if 'H50E' in str(col_name):
                    mapsheet_columns.append(col_name)
            
            # 计算每个图幅的总完成量和最新数据
            for mapsheet_name in mapsheet_columns:
                try:
                    col_idx = self.excel_data.columns.get_loc(mapsheet_name)
                    mapsheet_column = self.excel_data.iloc[:, col_idx]
                    
                    # 计算总点数
                    total_points = 0
                    latest_points = 0
                    latest_date = None
                    active_days = 0
                    
                    for row_idx in range(len(self.excel_data)):
                        points_value = mapsheet_column.iloc[row_idx]
                        date_value = self.excel_data.iloc[row_idx, 0]
                        
                        if not pd.isna(points_value) and not pd.isna(date_value):
                            try:
                                points = int(float(points_value))
                                if points < 0:
                                    points = 0
                                
                                total_points += points
                                
                                if points > 0:
                                    active_days += 1
                                    latest_points = points
                                    
                                    # 解析日期
                                    date_str = str(int(date_value)) if isinstance(date_value, (int, float)) else str(date_value).strip()
                                    date_obj = self._parse_date_string(date_str)
                                    if date_obj:
                                        latest_date = date_obj
                                        
                            except (ValueError, TypeError):
                                continue
                    
                    # 估算目标点数（基于当前完成情况和标准目标）
                    # 假设每个图幅的目标点数相似，可以基于总体目标平均分配
                    estimated_target = max(total_points * 2, 500)  # 最少500点
                    
                    mapsheet_status[mapsheet_name] = {
                        'current_points': total_points,
                        'estimated_target': estimated_target,
                        'completion_rate': (total_points / estimated_target * 100) if estimated_target > 0 else 0,
                        'active_days': active_days,
                        'latest_daily_points': latest_points,
                        'latest_date': latest_date.strftime('%Y-%m-%d') if latest_date else None,
                        'avg_daily_points': (total_points / active_days) if active_days > 0 else 0
                    }
                    
                except Exception as e:
                    logger.debug(f"处理图幅 {mapsheet_name} 状态失败: {e}")
                    continue
            
            logger.info(f"获取 {len(mapsheet_status)} 个图幅的当前状态")
            return mapsheet_status
            
        except Exception as e:
            logger.error(f"获取图幅状态失败: {e}")
            return {}
