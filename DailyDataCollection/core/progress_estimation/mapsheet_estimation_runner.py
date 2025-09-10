#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
图幅完成日期估算主程序

使用Excel中的真实数据，对不同图幅应用多种方法预估完成日期
支持的估算方法：
1. 简单平均法 (simple_average)
2. 加权平均法 (weighted_average) 
3. 线性回归法 (linear_regression)
4. 蒙特卡洛模拟法 (monte_carlo)
"""

import os
import sys
import logging
import pandas as pd
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Any
from pathlib import Path

# 添加项目根目录到路径
sys.path.append(os.path.dirname(os.path.dirname(os.path.dirname(__file__))))

from core.data_models.date_types import DateType
from core.progress_estimation.data_analyzer import DataAnalyzer
from core.progress_estimation.finish_date_estimator import FinishDateEstimator
from core.progress_estimation.progress_charts import ProgressCharts
from core.progress_estimation.real_excel_data_connector import RealExcelDataConnector
from core.progress_estimation.method_integrator import MethodIntegrator
from core.progress_estimation.completed_project_handler import CompletedProjectHandler
from config.logger_manager import get_logger

# 设置日志
logger = get_logger(__name__)


class MapsheetEstimationRunner:
    """图幅完成日期估算运行器"""
    
    def __init__(self, output_dir: str = None, skip_completed_estimation: bool = False):
        """
        初始化估算运行器
        
        Args:
            output_dir: 输出目录路径
            skip_completed_estimation: 是否跳过已完成项目的复杂估算
        """
        self.output_dir = output_dir or os.path.join(os.getcwd(), 'mapsheet_estimation_results')
        os.makedirs(self.output_dir, exist_ok=True)
        
        # 初始化数据连接器
        self.data_connector = RealExcelDataConnector()
        
        # 初始化方法集成器
        self.method_integrator = MethodIntegrator()
        
        # 初始化已完成项目处理器
        self.completed_handler = CompletedProjectHandler(skip_estimation=skip_completed_estimation)
        self.skip_completed_estimation = skip_completed_estimation
        
        # 估算方法列表
        self.estimation_methods = [
            'simple_average',
            'weighted_average', 
            'linear_regression',
            'monte_carlo'
        ]
        
        # 方法中文名称映射
        self.method_names = {
            'simple_average': '简单平均法',
            'weighted_average': '加权平均法',
            'linear_regression': '线性回归法',
            'monte_carlo': '蒙特卡洛模拟法'
        }
        
        logger.info(f"图幅估算运行器已初始化，输出目录: {self.output_dir}")
    
    def _format_completion_date(self, estimation: Dict[str, Any], is_completed: bool = False) -> str:
        """
        格式化完成日期显示
        
        Args:
            estimation: 估算结果
            is_completed: 是否已完成
            
        Returns:
            str: 格式化的日期字符串，已完成项目返回"-"
        """
        # 检查是否已完成
        status = estimation.get('status', '')
        if is_completed or status in ['completed', 'completed_skipped']:
            return '-'
        
        # 处理正常的估算日期
        est_date = estimation.get('estimated_date')
        if isinstance(est_date, str):
            return est_date
        elif est_date:
            return est_date.strftime('%Y-%m-%d')
        else:
            return '无法计算'
    
    def run_mapsheet_estimations(self, 
                                start_date: DateType = None,
                                days_back: int = 30,
                                confidence_level: float = 0.8) -> Dict[str, Any]:
        """
        运行图幅完成日期估算
        
        Args:
            start_date: 数据开始日期，默认为30天前
            days_back: 回溯天数
            confidence_level: 置信度
            
        Returns:
            Dict: 估算结果汇总
        """
        logger.info("开始运行图幅完成日期估算...")
        
        # 设置日期范围
        if start_date is None:
            start_date = DateType(datetime.now() - timedelta(days=days_back))
        end_date = DateType(datetime.now())
        
        # 加载Excel数据
        if not self.data_connector.load_excel_data():
            logger.error("无法加载Excel数据")
            return {}
        
        # 获取图幅历史数据
        mapsheet_data = self.data_connector.extract_mapsheet_historical_data(start_date, end_date)
        if not mapsheet_data:
            logger.error("未找到图幅历史数据")
            return {}
        
        logger.info(f"找到 {len(mapsheet_data)} 个图幅的数据")
        
        # 获取图幅当前状态
        team_status = self.data_connector.get_team_current_status()
        mapsheet_status = self._aggregate_team_status_by_mapsheet(team_status)
        
        # 为每个图幅进行估算
        estimation_results = {}
        
        for sheet_no, historical_records in mapsheet_data.items():
            logger.info(f"开始处理图幅 {sheet_no}...")
            
            try:
                # 获取该图幅的状态
                sheet_status = mapsheet_status.get(sheet_no, {})
                if not sheet_status:
                    logger.warning(f"图幅 {sheet_no} 缺少状态信息，跳过")
                    continue
                
                # 为该图幅创建数据分析器
                sheet_analyzer = self._create_mapsheet_analyzer(sheet_no, historical_records)
                
                if not sheet_analyzer:
                    logger.warning(f"图幅 {sheet_no} 无法创建分析器，跳过")
                    continue
                
                # 创建估算器
                estimator = FinishDateEstimator(sheet_analyzer, self.skip_completed_estimation)
                
                # 检查是否为已完成项目
                current_points = sheet_status.get('current_points', 0)
                target_points = sheet_status.get('estimated_target', 0)
                is_completed = self.completed_handler.is_project_completed(current_points, target_points)
                
                if is_completed:
                    logger.info(f"图幅 {sheet_no} 已完成 ({current_points}/{target_points})")
                
                # 应用多种方法进行估算（包括已完成项目的特殊处理）
                sheet_results = self._estimate_with_multiple_methods(
                    estimator, 
                    sheet_status,
                    confidence_level,
                    is_completed
                )
                
                # 智能集成多种估算结果
                data_quality_info = self._assess_data_quality(historical_records)
                integration_result = self.method_integrator.integrate_estimation_results(
                    sheet_results, data_quality_info
                )
                
                # 生成图表
                self._generate_mapsheet_charts(sheet_no, sheet_analyzer, sheet_status, estimator)
                
                estimation_results[sheet_no] = {
                    'sheet_info': {
                        'sheet_no': sheet_no,
                        'current_points': sheet_status.get('current_points', 0),
                        'target_points': sheet_status.get('estimated_target', 0),
                        'completion_rate': sheet_status.get('completion_rate', 0),
                        'active_days': sheet_status.get('active_days', 0),
                        'avg_daily_points': sheet_status.get('avg_daily_points', 0)
                    },
                    'estimations': sheet_results,
                    'integration': integration_result,  # 新增：智能集成结果
                    'data_quality': data_quality_info,
                    'recommendations': self._generate_recommendations(sheet_status, sheet_results, integration_result)
                }
                
                logger.info(f"图幅 {sheet_no} 处理完成")
                
            except Exception as e:
                logger.error(f"处理图幅 {sheet_no} 时发生错误: {e}")
                continue
        
        # 生成汇总报告
        summary_report = self._generate_summary_report(estimation_results)
        
        # 保存结果
        self._save_results(estimation_results, summary_report)
        
        logger.info(f"图幅估算完成，处理了 {len(estimation_results)} 个图幅")
        return {
            'mapsheet_results': estimation_results,
            'summary_report': summary_report,
            'output_dir': self.output_dir
        }
    
    def _aggregate_team_status_by_mapsheet(self, team_status: Dict[str, Dict]) -> Dict[str, Dict]:
        """将团队状态按图幅聚合"""
        mapsheet_status = {}
        
        for team_identifier, status in team_status.items():
            team_info = status['team_info']
            sheet_no = team_info.get('sheet_no', 'Unknown')
            
            if sheet_no not in mapsheet_status:
                mapsheet_status[sheet_no] = {
                    'current_points': 0,
                    'estimated_target': 0,
                    'active_days': 0,
                    'teams': [],
                    'latest_date': None
                }
            
            # 聚合数据
            mapsheet_status[sheet_no]['current_points'] += status.get('current_points', 0)
            mapsheet_status[sheet_no]['estimated_target'] += status.get('estimated_target', 0)
            mapsheet_status[sheet_no]['active_days'] = max(
                mapsheet_status[sheet_no]['active_days'], 
                status.get('active_days', 0)
            )
            mapsheet_status[sheet_no]['teams'].append(team_identifier)
            
            # 更新最新日期
            latest_date = status.get('latest_date')
            if latest_date:
                if (mapsheet_status[sheet_no]['latest_date'] is None or 
                    latest_date > mapsheet_status[sheet_no]['latest_date']):
                    mapsheet_status[sheet_no]['latest_date'] = latest_date
        
        # 计算完成率和平均日完成量
        for sheet_no, status in mapsheet_status.items():
            current = status['current_points']
            target = status['estimated_target']
            active_days = status['active_days']
            
            status['completion_rate'] = (current / target * 100) if target > 0 else 0
            status['avg_daily_points'] = (current / active_days) if active_days > 0 else 0
            status['team_count'] = len(status['teams'])
        
        return mapsheet_status
    
    def _create_mapsheet_analyzer(self, sheet_no: str, historical_records: List[Dict]) -> Optional[DataAnalyzer]:
        """为图幅创建数据分析器"""
        try:
            # 创建分析器（使用真实数据模式，但提供预处理的数据）
            analyzer = DataAnalyzer(use_real_data=False)  # 使用模拟模式以便手动注入数据
            
            # 将历史记录转换为DataFrame
            if not historical_records:
                return None
            
            df = pd.DataFrame(historical_records)
            df['date'] = pd.to_datetime(df['date_obj'])
            df = df.sort_values('date')
            
            # 手动设置历史数据
            analyzer.historical_data = df
            
            # 计算日统计
            analyzer._calculate_daily_statistics()
            
            logger.debug(f"为图幅 {sheet_no} 创建分析器成功，包含 {len(df)} 天的数据")
            return analyzer
            
        except Exception as e:
            logger.error(f"为图幅 {sheet_no} 创建分析器失败: {e}")
            return None
    
    def _estimate_with_multiple_methods(self, 
                                      estimator: FinishDateEstimator,
                                      sheet_status: Dict,
                                      confidence_level: float,
                                      is_completed: bool = False) -> Dict[str, Dict]:
        """使用多种方法进行估算"""
        results = {}
        
        current_points = sheet_status.get('current_points', 0)
        target_points = sheet_status.get('estimated_target', 0)
        
        if is_completed and self.skip_completed_estimation:
            # 策略A：完全跳过复杂估算，快速返回已完成状态
            logger.debug(f"跳过已完成项目的复杂估算 (当前: {current_points}, 目标: {target_points})")
            for method in self.estimation_methods:
                results[method] = self.completed_handler.create_completed_estimation_result(
                    current_points=current_points,
                    target_points=target_points,
                    method=method,
                    historical_data=estimator.data_analyzer.historical_data
                )
            return results
        
        if target_points <= current_points:
            # 策略B：使用增强的已完成分析
            for method in self.estimation_methods:
                results[method] = estimator.completed_handler.create_completed_estimation_result(
                    current_points=current_points,
                    target_points=target_points,
                    method=method,
                    historical_data=estimator.data_analyzer.historical_data
                )
                results[method]['method_name'] = self.method_names[method]
            return results
        
        # 对每种方法进行估算
        for method in self.estimation_methods:
            try:
                result = estimator.estimate_finish_date(
                    target_points=target_points,
                    current_points=current_points,
                    method=method,
                    confidence_level=confidence_level
                )
                
                result['method_name'] = self.method_names[method]
                results[method] = result
                
                logger.debug(f"方法 {method} 估算完成")
                
            except Exception as e:
                logger.warning(f"方法 {method} 估算失败: {e}")
                # 提供回退估算
                results[method] = {
                    'estimated_date': datetime.now() + timedelta(days=30),
                    'days_remaining': 30,
                    'confidence': 0.1,
                    'status': 'fallback',
                    'error': str(e),
                    'method_name': self.method_names[method]
                }
        
        return results
    
    def _generate_mapsheet_charts(self, 
                                sheet_no: str,
                                analyzer: DataAnalyzer,
                                sheet_status: Dict,
                                estimator: FinishDateEstimator):
        """为图幅生成图表"""
        try:
            sheet_output_dir = os.path.join(self.output_dir, f'mapsheet_{sheet_no}')
            os.makedirs(sheet_output_dir, exist_ok=True)
            
            # 创建图表生成器
            charts = ProgressCharts(analyzer, sheet_output_dir)
            
            # 生成综合仪表板
            target_points = sheet_status.get('estimated_target', 0)
            current_points = sheet_status.get('current_points', 0)
            
            dashboard_path = charts.generate_progress_dashboard(
                target_points=target_points,
                current_points=current_points,
                finish_estimator=estimator
            )
            
            if dashboard_path:
                logger.info(f"图幅 {sheet_no} 图表已生成: {dashboard_path}")
                
        except Exception as e:
            logger.error(f"为图幅 {sheet_no} 生成图表失败: {e}")
    
    def _assess_data_quality(self, historical_records: List[Dict]) -> Dict[str, Any]:
        """评估数据质量"""
        if not historical_records:
            return {'quality': 'poor', 'reason': '无历史数据'}
        
        total_days = len(historical_records)
        active_days = sum(1 for r in historical_records if r.get('daily_points', 0) > 0)
        
        if total_days < 7:
            quality = 'poor'
            reason = f'数据不足（仅{total_days}天）'
        elif active_days / total_days < 0.3:
            quality = 'medium'
            reason = f'活跃度较低（{active_days}/{total_days}天有数据）'
        elif active_days / total_days < 0.6:
            quality = 'good'
            reason = f'数据较好（{active_days}/{total_days}天有数据）'
        else:
            quality = 'excellent'
            reason = f'数据优秀（{active_days}/{total_days}天有数据）'
        
        return {
            'quality': quality,
            'reason': reason,
            'total_days': total_days,
            'active_days': active_days,
            'activity_rate': active_days / total_days if total_days > 0 else 0
        }
    
    def _generate_recommendations(self, sheet_status: Dict, estimations: Dict[str, Dict], 
                                integration_result: Dict[str, Any] = None) -> List[str]:
        """生成建议"""
        recommendations = []
        
        completion_rate = sheet_status.get('completion_rate', 0)
        avg_daily = sheet_status.get('avg_daily_points', 0)
        
        # 基于完成率的建议
        if completion_rate < 10:
            recommendations.append("项目刚开始，建议建立稳定的工作节奏")
        elif completion_rate < 30:
            recommendations.append("项目进度较慢，建议分析进度瓶颈")
        elif completion_rate < 70:
            recommendations.append("项目进展正常，继续保持当前工作节奏")
        elif completion_rate < 90:
            recommendations.append("项目即将完成，注意质量控制")
        else:
            recommendations.append("项目接近完成，准备验收工作")
        
        # 基于速度的建议  
        if avg_daily < 5:
            recommendations.append("日均完成量较低，建议增加工作强度或人员")
        elif avg_daily > 20:
            recommendations.append("日均完成量较高，注意工作质量和团队状态")
        
        # 基于估算结果的建议
        methods_count = len([r for r in estimations.values() if r.get('status') != 'fallback'])
        if methods_count < 3:
            recommendations.append("历史数据不足，估算准确性有限，建议增加数据采集")
        
        # 基于智能集成结果的建议
        if integration_result:
            # 添加方法选择建议
            method_recommendations = integration_result.get('recommendations', [])
            recommendations.extend(method_recommendations)
            
            # 添加最佳方法建议
            best_method = integration_result.get('best_method')
            if best_method:
                method_name = best_method.get('method', '')
                reliability = best_method.get('reliability_score', 0)
                if reliability > 0.8:
                    recommendations.append(f"推荐使用{method_name}方法，可靠性评分: {reliability:.1%}")
                elif reliability > 0.6:
                    recommendations.append(f"建议参考{method_name}方法，但需注意不确定性")
            
            # 添加组合估算建议
            ensemble = integration_result.get('ensemble_estimation')
            if ensemble and ensemble.get('confidence', 0) > 0.7:
                recommendations.append("多方法组合估算置信度较高，可作为主要参考")
            
            # 添加一致性建议
            consistency = integration_result.get('consistency_analysis', {})
            if consistency.get('score', 0) < 0.4:
                recommendations.append("估算方法间差异较大，建议谨慎解读结果")
        
        # 检查估算一致性（保留原有逻辑作为备份）
        completion_dates = [r.get('estimated_date') for r in estimations.values() 
                          if r.get('estimated_date') and r.get('status') != 'fallback']
        if len(completion_dates) >= 2 and not integration_result:
            date_range = max(completion_dates) - min(completion_dates)
            if date_range.days > 30:
                recommendations.append("不同方法估算差异较大，项目存在不确定性")
        
        return recommendations
    
    def _generate_summary_report(self, estimation_results: Dict[str, Any]) -> Dict[str, Any]:
        """生成汇总报告"""
        total_mapsheets = len(estimation_results)
        if total_mapsheets == 0:
            return {'error': '无有效估算结果'}
        
        # 统计信息
        total_current = sum(r['sheet_info']['current_points'] for r in estimation_results.values())
        total_target = sum(r['sheet_info']['target_points'] for r in estimation_results.values())
        overall_completion = (total_current / total_target * 100) if total_target > 0 else 0
        
        # 按完成率分类
        completion_categories = {
            'completed': [],      # >= 90%
            'near_completion': [], # 70-89%
            'in_progress': [],     # 30-69%
            'starting': [],        # 10-29%
            'not_started': []      # < 10%
        }
        
        for sheet_no, result in estimation_results.items():
            completion_rate = result['sheet_info']['completion_rate']
            if completion_rate >= 90:
                completion_categories['completed'].append(sheet_no)
            elif completion_rate >= 70:
                completion_categories['near_completion'].append(sheet_no)
            elif completion_rate >= 30:
                completion_categories['in_progress'].append(sheet_no)
            elif completion_rate >= 10:
                completion_categories['starting'].append(sheet_no)
            else:
                completion_categories['not_started'].append(sheet_no)
        
        # 数据质量统计
        quality_stats = {
            'excellent': [],
            'good': [],
            'medium': [],
            'poor': []
        }
        
        for sheet_no, result in estimation_results.items():
            quality = result['data_quality']['quality']
            quality_stats[quality].append(sheet_no)
        
        # 估算方法可靠性统计
        method_reliability = {}
        integration_analysis = {
            'best_methods_distribution': {},
            'ensemble_confidence_avg': 0,
            'consistency_distribution': {},
            'method_agreement_rate': 0
        }
        
        ensemble_confidences = []
        consistency_scores = []
        best_method_counts = {}
        
        for method in ['simple_average', 'weighted_average', 'linear_regression', 'monte_carlo']:
            successful_estimates = 0
            total_estimates = 0
            avg_confidence = 0
            
            for result in estimation_results.values():
                if method in result['estimations']:
                    total_estimates += 1
                    estimation = result['estimations'][method]
                    if estimation.get('status') != 'fallback':
                        successful_estimates += 1
                        avg_confidence += estimation.get('confidence', 0)
                
                # 收集集成分析数据
                integration = result.get('integration', {})
                if integration:
                    # 统计最佳方法分布
                    best_method = integration.get('best_method', {})
                    if best_method and best_method.get('method'):
                        best_method_name = best_method['method']
                        best_method_counts[best_method_name] = best_method_counts.get(best_method_name, 0) + 1
                    
                    # 收集组合估算置信度
                    ensemble = integration.get('ensemble_estimation', {})
                    if ensemble and 'confidence' in ensemble:
                        ensemble_confidences.append(ensemble['confidence'])
                    
                    # 收集一致性评分
                    consistency = integration.get('consistency_analysis', {})
                    if consistency and 'score' in consistency:
                        consistency_scores.append(consistency['score'])
            
            if total_estimates > 0:
                method_reliability[method] = {
                    'success_rate': successful_estimates / total_estimates,
                    'avg_confidence': avg_confidence / successful_estimates if successful_estimates > 0 else 0,
                    'successful_count': successful_estimates,
                    'total_count': total_estimates
                }
        
        # 计算集成分析统计
        if ensemble_confidences:
            integration_analysis['ensemble_confidence_avg'] = sum(ensemble_confidences) / len(ensemble_confidences)
        
        if consistency_scores:
            # 一致性分布统计
            high_consistency = len([s for s in consistency_scores if s > 0.7])
            medium_consistency = len([s for s in consistency_scores if 0.4 <= s <= 0.7])
            low_consistency = len([s for s in consistency_scores if s < 0.4])
            
            integration_analysis['consistency_distribution'] = {
                'high': high_consistency,
                'medium': medium_consistency, 
                'low': low_consistency
            }
        
        integration_analysis['best_methods_distribution'] = best_method_counts
        
        return {
            'timestamp': datetime.now().isoformat(),
            'total_mapsheets': total_mapsheets,
            'overall_statistics': {
                'total_current_points': total_current,
                'total_target_points': total_target,
                'overall_completion_rate': overall_completion,
                'average_completion_rate': sum(r['sheet_info']['completion_rate'] 
                                             for r in estimation_results.values()) / total_mapsheets
            },
            'completion_categories': completion_categories,
            'data_quality_distribution': quality_stats,
            'method_reliability': method_reliability,
            'integration_analysis': integration_analysis,  # 新增：集成分析统计
            'output_directory': self.output_dir
        }
    
    def _save_results(self, estimation_results: Dict[str, Any], summary_report: Dict[str, Any]):
        """保存结果到文件"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        saved_files = []
        
        try:
            # 保存详细结果（JSON）
            import json
            
            results_file = os.path.join(self.output_dir, f'mapsheet_estimations_{timestamp}.json')
            try:
                with open(results_file, 'w', encoding='utf-8') as f:
                    # 处理datetime对象以便JSON序列化
                    serializable_results = self._make_json_serializable(estimation_results)
                    json.dump(serializable_results, f, ensure_ascii=False, indent=2)
                saved_files.append(f"详细结果: {results_file}")
            except Exception as e:
                logger.warning(f"保存详细结果JSON失败: {e}")
            
            # 保存汇总报告（JSON）
            summary_file = os.path.join(self.output_dir, f'summary_report_{timestamp}.json')
            try:
                with open(summary_file, 'w', encoding='utf-8') as f:
                    serializable_summary = self._make_json_serializable(summary_report)
                    json.dump(serializable_summary, f, ensure_ascii=False, indent=2)
                saved_files.append(f"汇总报告: {summary_file}")
            except Exception as e:
                logger.warning(f"保存汇总报告JSON失败: {e}")
            
            # 生成人类可读的报告（TXT）
            readable_file = os.path.join(self.output_dir, f'readable_report_{timestamp}.txt')
            try:
                self._generate_readable_report(estimation_results, summary_report, readable_file)
                saved_files.append(f"可读报告: {readable_file}")
            except Exception as e:
                logger.warning(f"生成可读报告失败: {e}")
            
            # 生成Excel报告
            excel_file = os.path.join(self.output_dir, f'estimation_report_{timestamp}.xlsx')
            try:
                self._generate_excel_report(estimation_results, summary_report, excel_file)
                saved_files.append(f"Excel报告: {excel_file}")
            except Exception as e:
                logger.warning(f"生成Excel报告失败: {e}")
                # 如果Excel生成失败，尝试生成简化版本
                try:
                    self._generate_simple_excel_report(estimation_results, summary_report, excel_file)
                    saved_files.append(f"简化Excel报告: {excel_file}")
                except Exception as e2:
                    logger.error(f"生成简化Excel报告也失败: {e2}")
            
            if saved_files:
                logger.info("结果已保存:")
                for file_info in saved_files:
                    logger.info(f"  {file_info}")
            else:
                logger.error("所有文件保存都失败了")
            
        except Exception as e:
            logger.error(f"保存结果时发生严重错误: {e}")
    
    def _make_json_serializable(self, obj):
        """将对象转换为JSON可序列化格式"""
        import numpy as np
        
        if isinstance(obj, datetime):
            return obj.isoformat()
        elif isinstance(obj, dict):
            return {k: self._make_json_serializable(v) for k, v in obj.items()}
        elif isinstance(obj, list):
            return [self._make_json_serializable(item) for item in obj]
        elif isinstance(obj, np.integer):
            return int(obj)
        elif isinstance(obj, np.floating):
            return float(obj)
        elif isinstance(obj, np.ndarray):
            return obj.tolist()
        elif pd.notna(obj) and hasattr(obj, 'item'):  # pandas scalar types
            return obj.item()
        else:
            return obj
    
    def _generate_readable_report(self, estimation_results: Dict[str, Any], 
                                summary_report: Dict[str, Any], output_file: str):
        """生成人类可读的报告"""
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write("GMAS 图幅完成日期估算报告\n")
            f.write("=" * 50 + "\n\n")
            
            # 汇总信息
            f.write("汇总信息:\n")
            f.write("-" * 20 + "\n")
            overall = summary_report['overall_statistics']
            f.write(f"处理图幅数量: {summary_report['total_mapsheets']}\n")
            f.write(f"总目标点数: {overall['total_target_points']:,}\n")
            f.write(f"总完成点数: {overall['total_current_points']:,}\n")
            f.write(f"整体完成率: {overall['overall_completion_rate']:.1f}%\n")
            f.write(f"平均完成率: {overall['average_completion_rate']:.1f}%\n\n")
            
            # 分类统计
            f.write("完成状态分布:\n")
            f.write("-" * 20 + "\n")
            categories = summary_report['completion_categories']
            f.write(f"已完成 (≥90%): {len(categories['completed'])} 个\n")
            f.write(f"接近完成 (70-89%): {len(categories['near_completion'])} 个\n")
            f.write(f"进行中 (30-69%): {len(categories['in_progress'])} 个\n")
            f.write(f"刚开始 (10-29%): {len(categories['starting'])} 个\n")
            f.write(f"未开始 (<10%): {len(categories['not_started'])} 个\n\n")
            
            # 详细图幅信息
            f.write("详细图幅信息:\n")
            f.write("-" * 20 + "\n")
            
            for sheet_no, result in estimation_results.items():
                sheet_info = result['sheet_info']
                f.write(f"\n图幅 {sheet_no}:\n")
                f.write(f"  当前进度: {sheet_info['current_points']:,} / {sheet_info['target_points']:,} 点 ({sheet_info['completion_rate']:.1f}%)\n")
                f.write(f"  活跃天数: {sheet_info['active_days']} 天\n")
                f.write(f"  日均完成: {sheet_info['avg_daily_points']:.1f} 点\n")
                f.write(f"  数据质量: {result['data_quality']['quality']} - {result['data_quality']['reason']}\n")
                
                # 估算结果
                f.write("  估算结果:\n")
                sheet_info = result['sheet_info']
                is_completed = sheet_info['completion_rate'] >= 100
                
                for method, estimation in result['estimations'].items():
                    if estimation.get('status') != 'fallback':
                        method_name = estimation.get('method_name', method)
                        est_date_str = self._format_completion_date(estimation, is_completed)
                        days_remaining = estimation.get('days_remaining', 0)
                        confidence = estimation.get('confidence', 0)
                        
                        if is_completed:
                            f.write(f"    {method_name}: {est_date_str} (已完成, 置信度{confidence:.1%})\n")
                        else:
                            f.write(f"    {method_name}: {est_date_str} (剩余{days_remaining:.1f}天, 置信度{confidence:.1%})\n")
                
                # 建议
                if result['recommendations']:
                    f.write("  建议:\n")
                    for rec in result['recommendations']:
                        f.write(f"    - {rec}\n")
    
    def _generate_simple_excel_report(self, estimation_results: Dict[str, Any], summary_report: Dict[str, Any], output_file: str):
        """生成简化的Excel报告（备选方案）"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            
            # 1. 汇总工作表
            ws_summary = wb.active
            ws_summary.title = "汇总信息"
            
            # 基本标题
            ws_summary['A1'] = "图幅估算汇总报告"
            ws_summary['A1'].font = Font(bold=True, size=14)
            
            row = 3
            
            # 总体统计
            ws_summary[f'A{row}'] = "总体统计"
            ws_summary[f'A{row}'].font = Font(bold=True)
            row += 1
            
            if 'overall_statistics' in summary_report:
                stats = summary_report['overall_statistics']
                for key, value in stats.items():
                    ws_summary[f'A{row}'] = str(key)
                    ws_summary[f'B{row}'] = str(value)
                    row += 1
            
            row += 1
            
            # 2. 图幅结果工作表
            ws_results = wb.create_sheet("图幅结果")
            
            # 标题行
            headers = ["图幅名称", "简单平均", "加权平均", "线性回归", "蒙特卡洛"]
            for col, header in enumerate(headers, 1):
                cell = ws_results.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # 数据行
            row = 2
            for mapsheet_name, result in estimation_results.items():
                if isinstance(result, dict) and 'estimations' in result:
                    ws_results.cell(row=row, column=1, value=mapsheet_name)
                    
                    # 检查是否已完成
                    sheet_info = result.get('sheet_info', {})
                    is_completed = sheet_info.get('completion_rate', 0) >= 100
                    
                    estimations = result['estimations']
                    for col, method in enumerate(['simple_average', 'weighted_average', 'linear_regression', 'monte_carlo'], 2):
                        if method in estimations:
                            est_data = estimations[method]
                            if isinstance(est_data, dict):
                                if is_completed or est_data.get('status') in ['completed', 'completed_skipped']:
                                    # 已完成项目显示"-"
                                    ws_results.cell(row=row, column=col, value="-")
                                elif 'estimated_finish_date' in est_data:
                                    date_val = est_data['estimated_finish_date']
                                    if isinstance(date_val, str):
                                        ws_results.cell(row=row, column=col, value=date_val)
                                    else:
                                        ws_results.cell(row=row, column=col, value=str(date_val))
                                else:
                                    ws_results.cell(row=row, column=col, value="无法计算")
                            else:
                                ws_results.cell(row=row, column=col, value="无法计算")
                    row += 1
            
            # 调整列宽
            for ws in [ws_summary, ws_results]:
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 30)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(output_file)
            
        except Exception as e:
            logger.error(f"生成简化Excel报告失败: {e}")
            raise

    def _generate_excel_report(self, estimation_results: Dict[str, Any], 
                             summary_report: Dict[str, Any], output_file: str):
        """生成Excel报告"""
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                # 汇总工作表
                summary_data = []
                for sheet_no, result in estimation_results.items():
                    sheet_info = result['sheet_info']
                    data_quality = result['data_quality']
                    is_completed = sheet_info['completion_rate'] >= 100
                    
                    # 获取最佳估算（选择置信度最高的）
                    best_estimation = None
                    best_confidence = 0
                    for estimation in result['estimations'].values():
                        if (estimation.get('status') != 'fallback' and 
                            estimation.get('confidence', 0) > best_confidence):
                            best_estimation = estimation
                            best_confidence = estimation.get('confidence', 0)
                    
                    if best_estimation:
                        est_date_str = self._format_completion_date(best_estimation, is_completed)
                        days_remaining = best_estimation.get('days_remaining', 0)
                        method_name = best_estimation.get('method_name', '')
                    else:
                        est_date_str = '无法估算'
                        days_remaining = 0
                        method_name = '无'
                    
                    # 对已完成项目，剩余天数也显示为"-"
                    days_remaining_str = '-' if is_completed else str(round(days_remaining, 1))
                    
                    summary_data.append({
                        '图幅编号': sheet_no,
                        '当前点数': sheet_info['current_points'],
                        '目标点数': sheet_info['target_points'],
                        '完成率(%)': round(sheet_info['completion_rate'], 1),
                        '活跃天数': sheet_info['active_days'],
                        '日均完成': round(sheet_info['avg_daily_points'], 1),
                        '预计完成日期': est_date_str,
                        '剩余天数': days_remaining_str,
                        '估算方法': method_name,
                        '置信度(%)': round(best_confidence * 100, 1) if best_estimation else 0,
                        '数据质量': data_quality['quality'],
                        '数据质量说明': data_quality['reason']
                    })
                
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='汇总', index=False)
                
                # 详细估算工作表
                detail_data = []
                for sheet_no, result in estimation_results.items():
                    sheet_info = result['sheet_info']
                    is_completed = sheet_info['completion_rate'] >= 100
                    
                    for method, estimation in result['estimations'].items():
                        if estimation.get('status') != 'fallback':
                            est_date_str = self._format_completion_date(estimation, is_completed)
                            days_remaining = estimation.get('days_remaining', 0)
                            days_remaining_str = '-' if is_completed else str(round(days_remaining, 1))
                            
                            detail_data.append({
                                '图幅编号': sheet_no,
                                '估算方法': estimation.get('method_name', method),
                                '预计完成日期': est_date_str,
                                '剩余天数': days_remaining_str,
                                '置信度(%)': round(estimation.get('confidence', 0) * 100, 1),
                                '日均速度': round(estimation.get('daily_velocity', 0), 1),
                                '状态': estimation.get('status', '未知')
                            })
                
                detail_df = pd.DataFrame(detail_data)
                detail_df.to_excel(writer, sheet_name='详细估算', index=False)
                
            logger.info(f"Excel报告已生成: {output_file}")
            
        except Exception as e:
            logger.error(f"生成Excel报告失败: {e}")


def main():
    """主函数"""
    # 获取日志记录器
    logger = get_logger('mapsheet_estimation')
    
    # 创建运行器
    runner = MapsheetEstimationRunner()
    
    # 运行估算
    results = runner.run_mapsheet_estimations(
        days_back=30,  # 使用最近30天的数据
        confidence_level=0.8  # 80%置信度
    )
    
    if results:
        print("\n" + "="*60)
        print("GMAS 图幅完成日期估算完成!")
        print("="*60)
        
        summary = results.get('summary_report', {})
        if 'error' not in summary:
            overall = summary.get('overall_statistics', {})
            print(f"处理图幅数量: {summary.get('total_mapsheets', 0)}")
            print(f"整体完成率: {overall.get('overall_completion_rate', 0):.1f}%")
            print(f"输出目录: {results.get('output_dir', '')}")
            
            # 显示各类别图幅数量
            categories = summary.get('completion_categories', {})
            print(f"\n状态分布:")
            print(f"  已完成: {len(categories.get('completed', []))} 个")
            print(f"  接近完成: {len(categories.get('near_completion', []))} 个") 
            print(f"  进行中: {len(categories.get('in_progress', []))} 个")
            print(f"  刚开始: {len(categories.get('starting', []))} 个")
            print(f"  未开始: {len(categories.get('not_started', []))} 个")
        
        print("\n详细结果请查看输出目录中的报告文件。")
    else:
        print("估算运行失败，请检查日志信息。")


if __name__ == "__main__":
    main()
