"""
真实数据连接器模块

连接进度估算模块与实际的GMAS Excel数据文件
"""

import os
import logging
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
from openpyxl import load_workbook

from ..data_models.date_types import DateType
from config.config_manager import ConfigManager

logger = logging.getLogger(__name__)


class RealDataConnector:
    """真实数据连接器类，从GMAS Excel文件读取历史数据"""

    def __init__(self, workspace_path: str = None):
        """
        初始化真实数据连接器
        
        Args:
            workspace_path: 工作空间路径
        """
        self.workspace_path = workspace_path or os.getcwd()
        self.config_manager = ConfigManager()
        self.excel_file_path: Optional[str] = None
        self.excel_data: Optional[pd.DataFrame] = None
        
        # 从配置获取Excel文件路径
        self._initialize_excel_path()
        
    def _initialize_excel_path(self) -> None:
        """初始化Excel文件路径"""
        try:
            # 从配置文件获取统计详情文件路径
            stats_config = self.config_manager.get('reports.statistics')
            if stats_config and 'daily_details_file' in stats_config:
                file_path_template = stats_config['daily_details_file']
                # 使用配置管理器的模板解析功能
                self.excel_file_path = self.config_manager.resolve_path_template(file_path_template)
                logger.info(f"Excel文件路径: {self.excel_file_path}")
            else:
                # 回退到默认路径
                workspace = self.config_manager.get('system.workspace')
                filename = self.config_manager.get('reports.statistics.daily_details_file_name', 
                                                 'Daily_statistics_details_for_Group_3.2.xlsx')
                self.excel_file_path = os.path.join(workspace, filename)
                logger.warning(f"使用默认Excel文件路径: {self.excel_file_path}")
                
        except Exception as e:
            logger.error(f"初始化Excel文件路径失败: {e}")
            self.excel_file_path = None
    
    def load_excel_data(self) -> bool:
        """
        加载Excel数据
        
        Returns:
            bool: 是否成功加载
        """
        if not self.excel_file_path or not os.path.exists(self.excel_file_path):
            logger.error(f"Excel文件不存在: {self.excel_file_path}")
            return False
        
        try:
            # 使用openpyxl读取Excel文件
            wb = load_workbook(self.excel_file_path, data_only=True)
            
            # 尝试使用"总表"工作表
            if "总表" in wb.sheetnames:
                ws = wb["总表"]
            else:
                ws = wb.active
                logger.warning("未找到'总表'工作表，使用默认工作表")
            
            # 将工作表数据转换为DataFrame
            data = []
            max_row = ws.max_row
            max_col = ws.max_column
            
            # 读取所有数据
            for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
                data.append(row)
            
            # 创建DataFrame
            if data:
                self.excel_data = pd.DataFrame(data)
                logger.info(f"成功加载Excel数据: {len(data)}行 x {len(data[0])}列")
                return True
            else:
                logger.error("Excel文件为空")
                return False
                
        except Exception as e:
            logger.error(f"加载Excel数据失败: {e}")
            return False
    
    def extract_historical_data(self, start_date: DateType, end_date: DateType = None) -> List[Dict[str, Any]]:
        """
        从Excel数据中提取历史数据
        
        Args:
            start_date: 开始日期
            end_date: 结束日期，默认为今天
            
        Returns:
            List[Dict]: 历史数据列表
        """
        if self.excel_data is None:
            if not self.load_excel_data():
                return []
        
        if end_date is None:
            end_date = DateType(datetime.now())
        
        try:
            historical_records = []
            
            # 分析Excel结构
            date_columns = self._find_date_columns()
            mapsheet_rows = self._find_mapsheet_rows()
            
            if not date_columns or not mapsheet_rows:
                logger.error("无法解析Excel文件结构")
                return []
            
            # 提取指定日期范围的数据
            for date_str, col_idx in date_columns.items():
                try:
                    # 解析日期
                    date_obj = self._parse_date_string(date_str)
                    if not date_obj:
                        continue
                    
                    date_type = DateType(date_obj)
                    
                    # 检查是否在指定范围内
                    if not (start_date.date_datetime <= date_type.date_datetime <= end_date.date_datetime):
                        continue
                    
                    # 提取该日期的点数数据
                    daily_points = self._extract_daily_points(col_idx, mapsheet_rows)
                    total_points = sum(daily_points.values()) if daily_points else 0
                    
                    if total_points > 0:  # 只记录有数据的日期
                        # 估算活跃团队数量（基于有数据的图幅数量）
                        active_mapsheets = len([p for p in daily_points.values() if p > 0])
                        teams_active = max(1, active_mapsheets // 3)  # 假设每3个图幅一个团队
                        
                        record = {
                            'date': date_type.yyyymmdd_str,
                            'completed_points': total_points,
                            'teams_active': teams_active,
                            'workday': date_type.date_datetime.weekday() < 5,
                            'mapsheet_details': daily_points
                        }
                        
                        historical_records.append(record)
                        
                except Exception as e:
                    logger.debug(f"解析日期 {date_str} 数据失败: {e}")
                    continue
            
            # 按日期排序
            historical_records.sort(key=lambda x: x['date'])
            
            # 计算累计完成量
            cumulative = 0
            for record in historical_records:
                cumulative += record['completed_points']
                record['cumulative_points'] = cumulative
            
            logger.info(f"成功提取 {len(historical_records)} 天的历史数据")
            return historical_records
            
        except Exception as e:
            logger.error(f"提取历史数据失败: {e}")
            return []
        """
        try:
            # 使用GMAS系统的CurrentDateFiles获取当日数据
            current_files = CurrentDateFiles(date_obj)
            
            # 获取当日的统计数据
            daily_data = {
                'date': date_obj.yyyymmdd_str,
                'completed_points': current_files.totalDaiyIncreasePointNum,
                'cumulative_points': 0,  # 将在后续计算中更新
                'teams_active': self._calculate_active_teams(current_files),
                'workday': date_obj.date_datetime.weekday() < 5,
                'mapsheets_updated': len([m for m in current_files.currentDateFiles if m.dailyincreasePointNum and m.dailyincreasePointNum > 0]),
                'total_routes': current_files.totalDaiyIncreaseRouteNum,
                'mapsheet_details': self._get_mapsheet_details(current_files)
            }
            
            # 如果当天没有任何数据，返回None
            if daily_data['completed_points'] == 0 and daily_data['total_routes'] == 0:
                logger.debug(f"日期 {date_obj.yyyymmdd_str} 无进度数据")
                return None
                
            logger.debug(f"获取到日期 {date_obj.yyyymmdd_str} 的进度数据: {daily_data['completed_points']} 点")
            return daily_data
            
        except Exception as e:
            logger.warning(f"获取日期 {date_obj.yyyymmdd_str} 的数据失败: {e}")
            return None
    
    def _calculate_active_teams(self, current_files: CurrentDateFiles) -> int:
        """
        计算活跃团队数量
        
        基于当日有数据更新的图幅数量来估算活跃团队
        """
        try:
            # 统计有更新的图幅数量
            updated_mapsheets = [
                mapsheet for mapsheet in current_files.currentDateFiles 
                if mapsheet.dailyincreasePointNum and mapsheet.dailyincreasePointNum > 0
            ]
            
            # 简化的团队数量估算：假设每个团队平均负责1-2个图幅
            active_teams = max(1, len(updated_mapsheets) // 2)
            
            # 考虑工作模式调整
            date_obj = current_files.currentDate
            if date_obj.date_datetime.weekday() >= 5:  # 周末
                active_teams = max(1, active_teams // 2)
            
            return min(active_teams, 6)  # 最多6个团队
            
        except Exception as e:
            logger.warning(f"计算活跃团队数量失败: {e}")
            return 1  # 默认值
    
    def _get_mapsheet_details(self, current_files: CurrentDateFiles) -> Dict[str, Any]:
        """获取图幅详细信息"""
        try:
            return {
                'daily_finished_points': current_files.dailyFinishedPoints,
                'daily_increased_points': current_files.dailyIncreasedPoints,
                'total_mapsheets': len(current_files.currentDateFiles),
                'active_mapsheets': len([m for m in current_files.currentDateFiles if m.dailyincreasePointNum and m.dailyincreasePointNum > 0])
            }
        except Exception as e:
            logger.warning(f"获取图幅详细信息失败: {e}")
            return {}
    
    def get_cumulative_progress(self, end_date: DateType) -> int:
        """
        获取截止到指定日期的累计进度
        
        Args:
            end_date: 结束日期
            
        Returns:
            int: 累计完成的点数
        """
        try:
            current_files = CurrentDateFiles(end_date)
            
            # 获取所有图幅的累计完成点数
            total_cumulative = sum(current_files.dailyFinishedPoints.values())
            
            logger.info(f"截止到 {end_date.yyyymmdd_str} 的累计进度: {total_cumulative} 点")
            return total_cumulative
            
        except Exception as e:
            logger.error(f"获取累计进度失败: {e}")
            return 0
    
    def get_project_target(self) -> Optional[int]:
        """
        获取项目目标点数
        
        从配置或历史数据中推断项目目标
        """
        try:
            # 首先尝试从配置中获取
            target_points = self.config_manager.get('progress.target_points', None)
            if target_points:
                return int(target_points)
            
            # 如果配置中没有，可以基于图幅数量估算
            # 这里使用一个简化的估算方法
            from ..mapsheet.mapsheet_manager import mapsheet_manager
            total_mapsheets = len(mapsheet_manager.maps_info)
            
            # 假设每个图幅平均100个观测点
            estimated_target = total_mapsheets * 100
            
            logger.info(f"基于 {total_mapsheets} 个图幅估算项目目标: {estimated_target} 点")
            return estimated_target
            
        except Exception as e:
            logger.warning(f"获取项目目标失败: {e}")
            return None
    
    def validate_data_availability(self, start_date: DateType, end_date: DateType) -> Dict[str, Any]:
        """
        验证指定日期范围内的数据可用性
        
        Args:
            start_date: 开始日期
            end_date: 结束日期
            
        Returns:
            Dict: 数据可用性报告
        """
        validation_report = {
            'total_days': 0,
            'available_days': 0,
            'missing_days': [],
            'data_coverage': 0.0,
            'recommendation': ''
        }
        
        try:
            current = start_date.date_datetime
            end = end_date.date_datetime
            
            while current <= end:
                validation_report['total_days'] += 1
                date_obj = DateType(current)
                
                daily_data = self.get_daily_progress_data(date_obj)
                if daily_data:
                    validation_report['available_days'] += 1
                else:
                    validation_report['missing_days'].append(date_obj.yyyymmdd_str)
                
                current += timedelta(days=1)
            
            # 计算数据覆盖率
            if validation_report['total_days'] > 0:
                validation_report['data_coverage'] = validation_report['available_days'] / validation_report['total_days']
            
            # 生成建议
            coverage = validation_report['data_coverage']
            if coverage >= 0.8:
                validation_report['recommendation'] = '数据覆盖率良好，可以进行可靠的进度分析'
            elif coverage >= 0.5:
                validation_report['recommendation'] = '数据覆盖率一般，分析结果可能存在偏差'
            else:
                validation_report['recommendation'] = '数据覆盖率较低，建议扩大日期范围或检查数据源'
            
            logger.info(f"数据可用性验证完成，覆盖率: {coverage:.1%}")
            
        except Exception as e:
            logger.error(f"数据可用性验证失败: {e}")
            validation_report['recommendation'] = '数据验证过程中发生错误'
        
        return validation_report
    
    def get_historical_data_summary(self, days_back: int = 30) -> Dict[str, Any]:
        """
        获取历史数据摘要
        
        Args:
            days_back: 往前追溯的天数
            
        Returns:
            Dict: 历史数据摘要
        """
        end_date = DateType(datetime.now())
        start_date = DateType(datetime.now() - timedelta(days=days_back))
        
        summary = {
            'period': {
                'start_date': start_date.yyyymmdd_str,
                'end_date': end_date.yyyymmdd_str,
                'days': days_back
            },
            'total_progress': 0,
            'daily_average': 0,
            'peak_day': {'date': '', 'points': 0},
            'active_days': 0,
            'data_validation': self.validate_data_availability(start_date, end_date)
        }
        
        try:
            daily_points = []
            current = start_date.date_datetime
            
            while current <= end_date.date_datetime:
                date_obj = DateType(current)
                daily_data = self.get_daily_progress_data(date_obj)
                
                if daily_data:
                    points = daily_data['completed_points']
                    daily_points.append(points)
                    summary['total_progress'] += points
                    summary['active_days'] += 1
                    
                    # 记录峰值日
                    if points > summary['peak_day']['points']:
                        summary['peak_day'] = {
                            'date': date_obj.yyyymmdd_str,
                            'points': points
                        }
                
                current += timedelta(days=1)
            
            # 计算平均值
            if summary['active_days'] > 0:
                summary['daily_average'] = summary['total_progress'] / summary['active_days']
            
            logger.info(f"历史数据摘要: {summary['active_days']} 个活跃日，"
                       f"总进度 {summary['total_progress']} 点，"
                       f"日均 {summary['daily_average']:.1f} 点")
            
        except Exception as e:
            logger.error(f"获取历史数据摘要失败: {e}")
        
        return summary
