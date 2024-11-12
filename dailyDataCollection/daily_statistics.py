import argparse
import os
import zipfile
import rarfile
import shutil
import fnmatch
import re
import stat
from osgeo import ogr
from datetime import datetime, timedelta
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import *


## 解压当前目录下所有的压缩文件至特定目录
def extract_all_files_in_directory(directory: str, output_directory: str) -> None:
    for filename in os.listdir(directory):
        if filename.endswith(".zip"):
            with zipfile.ZipFile(os.path.join(directory, filename), 'r') as zip_ref:
                zip_ref.extractall(output_directory)
        elif filename.endswith(".rar"):
            with rarfile.RarFile(os.path.join(directory, filename), 'r') as rar_ref:
                rar_ref.extractall(output_directory)
    return None


## 返回当前目录下所有的特定类型（后缀名）文件名称
def list_all_files_with_ext(directory: str, ext: str) -> list:
    
    while is_valid_path(directory):
        names = os.listdir(directory)
        files = [name for name in names if name.endswith(ext)]
        return files
    return None

## 返回当前目录下所有含特定字符串的文件全路径
def list_fullpath_of_files_with_keyword(directory: str, keyword: str) -> list:
    matches = []
    for root, _, filenames in os.walk(directory):
        for filename in fnmatch.filter(filenames, '*' + keyword + '*'):
            matches.append(os.path.join(root, filename))
    return matches

#返回当前目录下所有含多个特定字符串(不区分大小写)列表的文件全路径
def list_fullpath_of_files_with_keywords(directory: str, keywords: list) -> list:
    matches = []
    for root, _, files in os.walk(directory):
        for file in files:
            if all(keyword.lower() in file.lower() for keyword in keywords):
                matches.append(os.path.join(root, file))
    return matches

##返回当前目录及子目录下所有的特定类型文件的全路径
def list_fullpath_of_all_files_with_ext(directory, ext):
    file_paths = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(ext):
                file_paths.append(os.path.join(root, file))
    return file_paths


#移动包含特征字符串的文件至指定文件夹
def move_files_with_keyword(files, keyword, target_directory):
    for file in files:
        if keyword.lower() in os.path.basename(file).lower():
            shutil.move(file, os.path.join(target_directory, os.path.basename(file)))


#移动特定后缀名的文件至指定文件夹
def move_files_with_ext(files, ext, target_directory):
    for file in files:
        if file.endswith(ext):
            shutil.move(file, os.path.join(target_directory, os.path.basename(file)))

# 从文件名中提取日期字符串，并将其转换为 `datetime` 对象。
def get_date_from_filename(filename):
    match = re.search(r'\d{8}', filename)
    if match:
        return datetime.strptime(match.group(), "%Y%m%d")
    return None

def get_number_in_parentheses(filename):
    # 提取括号中的数字
    match = re.search(r'\((\d+)\)', filename)
    if match:
        return int(match.group(1))
    return -1

def find_files_with_max_number(directory) -> dict:
    files_dict = {}
    # 遍历目录中的所有文件
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            base_name = re.sub(r'\(\d+\)', '', filename)
            number = get_number_in_parentheses(filename)
            file_path = os.path.join(root, filename)
            if base_name not in files_dict:
                files_dict[base_name] = (file_path, number)
            else:
                existing_file, existing_number = files_dict[base_name]
                if number > existing_number:
                    files_dict[base_name] = (file_path, number)
    # 删除括号中数字为-1的文件
    files_dict = {base_name: (file_path, number) for base_name, (file_path, number) in files_dict.items() if number != -1}
    # 输出包含相同基础文件名中括号内数字最大的文件
    return files_dict


# 将文件属性改为可读写
def make_files_read_write(directory):
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            file_path = os.path.join(root, filename)
            os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
            # print(f"已将文件属性改为可读写: {file_path}")

#创建文件夹
def create_directory(directory):
    if not os.path.exists(directory):
        try:
            os.makedirs(directory)
            return f"目录 {directory} 创建成功"
        except PermissionError:
            return f"无权限创建目录 {directory}."
        except NotADirectoryError:
            return f"The path {directory} is a file, not a directory."
        except OSError as e:
            return f"An error occurred while creating the directory {directory}: {e}"
    else:
        return f"目录 {directory} 已存在"

#删除文件目录
def delete_directory(directory):
    if os.path.exists(directory):
        shutil.rmtree(directory)

#检查路径是否有效
def is_valid_path(path):
    # 检查路径是否为空
    if not path:
        print("路径为空")
        return False
    # 检查路径是否包含无效字符
    invalid_chars = set('<>"/|?*')
    if any(char in invalid_chars for char in path):
        print(f"路径 {path} 包含无效字符")
        return False
    # 检查路径是否存在
    if not os.path.exists(path):
        # print(f"路径 {path} 不存在")
        return False
    return True

#检查文件名是否有效
def is_valid_filename(filename):
    # 检查路径是否为空
    if not filename:
        return False
    # 检查路径是否包含无效字符
    invalid_chars = set('<>:"/\\|?*')
    if any(char in invalid_chars for char in filename):
        print(f"文件名 {filename} 包含无效字符")
        return False
    return True

#获取前一天的日期
def get_previous_day(date_str):
    # 将日期字符串转换为datetime对象
    date = datetime.strptime(date_str, "%Y%m%d")
    # 计算前一天的日期
    previous_day = date - timedelta(days=1)
    # 将前一天的日期转换回字符串
    previous_day_str = previous_day.strftime("%Y%m%d")
    return previous_day_str

#获取后一天的日期
def get_next_day(date_str):
    # 将日期字符串转换为datetime对象
    date = datetime.strptime(date_str, "%Y%m%d")
    # 计算后一天的日期
    next_day = date + timedelta(days=1)
    # 将后一天的日期转换回字符串
    next_day_str = next_day.strftime("%Y%m%d")
    return next_day_str


#拷贝文件
def copy_file(src, dst):
    try:
        shutil.copy(src, dst)
        return "File copied successfully."
    except Exception as e:
        return f"An error occurred while copying the file: {e}"

def rename_file(src, new_name):
    dir_name = os.path.dirname(src)
    new_path = os.path.join(dir_name, new_name)
    os.rename(src, new_path)




# 列表截取特定字符串之前部分，不区分大小写
def truncate_list(list, str):
    return [item.split(str.lower())[0] for item in list]

def is_subset(list1, list2):
    return set(list1).issubset(set(list2))

# 为字典中的列表填充空值，使其长度相等
def pad_dict(dict, pad_value=None):
    max_len = max(len(lst) for lst in dict.values())
    for key in dict:
        dict[key] += [pad_value] * (max_len - len(dict[key]))
    return dict

## 计算KMZ文件中所有layer的点要素总数量
## 计算原理：
## 1. 使用GDAL打开KMZ文件；2. 遍历KMZ文件中的每个layer；3. 遍历每个layer中的每个feature；4. 如果feature的description字段中包含点号，则计数加1。 判断方法为：使用正则表达式匹配成对的 <td> 和 </td>，且期间的字符串长度等于 9，前5位和后三位为数字，第四位为字母中的一个
def count_num_points_in_kmz(kmz_file):
    # 用于存储所有图层点的数量
    count_point = 0
    all_matches = []
    count_dict = {}
    ds = ogr.Open(kmz_file)
    if ds is None:
        print(f"Failed to open {kmz_file}")
        return None
    for i in range(ds.GetLayerCount()):
        layer = ds.GetLayerByIndex(i)
        if layer is None:
            continue

        count_dict[layer.GetName()] = layer.GetFeatureCount()
        for j in range(layer.GetFeatureCount()):
            feat = layer.GetNextFeature()
            for key in feat.keys():
                if feat[key] is None:
                    continue
                # 解析description里面包含的obsptid信息
                if key == "description" and len(feat[key]) >= 50:
                    if "obsptid" in feat[key].lower():
                        matches = re.findall(r'<td>(\d{5}[A-Za-z]\d{3})</td>', feat[key], re.IGNORECASE)
                        if matches:
                            if len(matches) > 1:
                                if all(x == matches[0] for x in matches):
                                    matches = [matches[0]]
                                else:
                                    print(f"不相符和列表: {matches}")
                            else:
                                None
                                # print(matches)
                            all_matches.extend(matches)
                            count_point += 1
    # print(f"{kmz_file}文件包括的点数: ", count_point)
    ds = None
    # 计算 all_matches 是否有重复元素，并输出重复的元素为新列表
    unique_matches = set()
    duplicate_matches = set()
    for match in all_matches:
        if match in unique_matches:
            duplicate_matches.add(match)
        else:
            unique_matches.add(match)

    is_duplicate = duplicate_matches == True
    
    # 通过第六位解析组号并存储在字典中
    obsptidStatistics_dict = {}
    for match in all_matches:
        key = match[5]
        if key not in obsptidStatistics_dict:
            obsptidStatistics_dict[key] = []
        obsptidStatistics_dict[key].append(match)

    # 分别计算第六位相同的个数，以及第六位相同的all_matches列表元素后三位是否连续
    # 建立与statistics_dict相同的字典，键值为原键值+"组"
    statistics_dict = {}
    for key, values in obsptidStatistics_dict.items():
        count = len(values)
        values.sort(key=lambda x: int(x[-3:]))  # 按后三位数字排序
        is_consecutive = all(int(values[i][-3:]) + 1 == int(values[i + 1][-3:]) for i in range(len(values) - 1))
        max_value = max(values, key=lambda x: int(x[-3:]))
        new_key = key + "组"
        statistics_dict[new_key] = {
            "完成点数": count,
            "点号是否连续": is_consecutive,
            "最大列表元素": max_value
        }
    statistics_dict["总数"] = {
        "总完成点数": count_point,
        "点号是否重复": is_duplicate,
        "最大列表元素": None
    }

    all_consecutive = all(value["点号是否连续"] for key, value in statistics_dict.items() if key != "总数")

    # 检查所有“点号是否重复”是否为 False
    all_not_duplicate = not statistics_dict["总数"]["点号是否重复"]

    # 累加所有“完成点数”
    total_completed_points = sum(value["完成点数"] for key, value in statistics_dict.items() if key != "总数")

    # 检查总完成点数是否等于所有“完成点数相加”
    total_points_match = total_completed_points == statistics_dict["总数"]["总完成点数"]

    # 如果以上条件均满足，则输出“验证通过”
    if all_consecutive and all_not_duplicate and total_points_match:
        None
        # print(f"{kmz_file}数据验证通过")
    else:
        print(f"{kmz_file}数据验证失败")
        print(statistics_dict)

    return count_point

# 检查文件是否被占用
def is_file_locked(filepath):
    if os.path.exists(filepath):
        try:
            with open(filepath, 'a'):
                pass
        except IOError:
            return True
    return False


# 根据列“Sequence”的值大小，输出指定值区间的相应“Roman Name”
def get_columns_in_range(df, minSequenceValue, maxSequenceValue):
    filtered_df = df[
        (df['Sequence'] >= minSequenceValue) & (df['Sequence'] <= maxSequenceValue)
    ]

    # 按照 'Sequence' 列值从小到大排序
    sorted_df = filtered_df.sort_values(by='Sequence')

    return (
        sorted_df['Roman Name'].tolist(),
        sorted_df['File Name'].tolist()
    )

def create_daily_statistics_excel(
        date_str: str,
        romanNames: list,
        excel_filename: str,
        ):
    """_summary_

    Args:
        date_str (str): 日期字符串，格式为“YYYY/MM/DD”
        excel_filename (str): Excel文件路径
        minSequenceValue (int): 最小填图序列号
        maxSequenceValue (int): 最大填图序列号

    Returns:
        _type_: 无
    """
    # 计算输出表格的行数和列数
    maxTableRows, maxTableColumns = len(romanNames) + 4, 4
    # 每日统计点文件的表头（前三行）
    daily_stat_header1 = ['Date', date_str]
    daily_stat_header2 = ['Map sheet name',
                          'Regular observation points finished',
                          'Field points on revised route'
                          ]
    daily_stat_header3 = ['', '', 'Added observation points',
                          'Added Structure points, photo points, mineralization points'
                          ]
    # 每日统计点文件的合计行（最后一行）
    daily_stat_footer = ['Today', '', '', '']

    # 创建一个新的 Excel 文件
    try:
        book = load_workbook(excel_filename)
    except FileNotFoundError:
        book = Workbook()

    sheet = book.active
    sheet.title = "Daily Statistics"
    # 写入表头到前三行
    for col_num, value in enumerate(daily_stat_header1, start=1):
        sheet.cell(row=1, column=col_num, value=value)
    for col_num, value in enumerate(daily_stat_header2, start=1):
        sheet.cell(row=2, column=col_num, value=value)
    for col_num, value in enumerate(daily_stat_header3, start=1):
        sheet.cell(row=3, column=col_num, value=value)
    # 写入表尾到最后一行
    for col_num, value in enumerate(daily_stat_footer, start=1):
        sheet.cell(row=maxTableRows, column=col_num, value=value)
    # 从第四行起，写入图幅罗马名称
    for i, value in enumerate(romanNames, start=4):
        sheet.cell(row=i, column=1, value=value)

    # 设置字体样式
    # 设置表头字体样式
    font_header = Font(
        name='Calibri',
        size=12,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )
    # 设置正文字体样式
    font = Font(
        name='Calibri',
        size=11,
        bold=False,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='00000000'
    )

    border = Border(
        left=Side(border_style='thin', color='FF000000'),
        right=Side(border_style='thin', color='FF000000'),
        top=Side(border_style='thin', color='FF000000'),
        bottom=Side(border_style='thin', color='FF000000'),
        diagonal=Side(border_style='thin', color='FF000000'),
        diagonal_direction=0,
        outline=Side(border_style='thin', color='FF000000'),
        vertical=Side(border_style='thin', color='FF000000'),
        horizontal=Side(border_style='thin', color='FF000000')
    )

    # 将字体样式应用到指定单元格
    # 前三行字体为表头字体
    for row in range(1, 4):
        for col in range(1, maxTableColumns + 1):
            cell = sheet.cell(row, column=col)
            cell.font = font_header
    # 最后一行字体为表头字体
    for col in range(1, maxTableColumns + 1):
        cell = sheet.cell(maxTableRows, column=col)
        cell.font = font_header
    # 其他字体为正文字体
    for row in range(4, maxTableRows):
        for col in range(1, maxTableColumns + 1):
            cell = sheet.cell(maxTableRows-1, column=col)
            cell.font = font

    # 将单元格格式应用到指定单元格
    for row in range(maxTableRows):
        for col in range(maxTableColumns):
            cell = sheet.cell(row=row+1, column=col+1)
            cell.border = border

    # 设置多列的宽度，adjusted_width是单元格宽度，比最大长度多2
    for column in sheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_length + 2)  # 设置单元格宽度，比最大长度多2
        sheet.column_dimensions[column[0].column_letter].width = adjusted_width

    # 创建一个居中对齐的对象
    center_aligned = Alignment(horizontal='center', vertical='center')
    # 将居中对齐应用到指定单元格
    for row in sheet['A1:D3']:  # 修改这个范围以适应你的需要
        for cell in row:
            cell.alignment = center_aligned
    for row in range(4, maxTableRows):
        for col in range(2, maxTableColumns):
            cell = sheet.cell(row, column=col)
            cell.alignment = center_aligned
    for row in range(maxTableRows + 1):
        for col in range(1, maxTableColumns + 1):
            cell = sheet.cell(row+1, column=col)
            cell.alignment = center_aligned

    # 设置合计行的公式
    sheet.cell(row=maxTableRows, column=2).value = f"=SUM(B4:B{maxTableRows-1})"
    sheet.cell(row=maxTableRows, column=3).value = f"=SUM(C4:C{maxTableRows-1})"
    sheet.cell(row=maxTableRows, column=4).value = f"=SUM(D4:D{maxTableRows-1})"

    sheet.merge_cells('B1:D1')
    sheet.merge_cells('C2:D2')
    sheet.merge_cells('A2:A3')
    sheet.merge_cells('B2:B3')

    # 保存工作簿
    book.save(excel_filename)
    return print(f"创建每日统计点 {excel_filename} 空表成功。")

def daily_statistics(date_str: str) -> None:
    """_summary_

    Args:
        date_str (str): 8位字符串类型的年月日YYMMDD

    Returns:
        _type_: 每日统计点的完成函数
    """


    workspace = r"D:\RouteDesigen"
    # 设置需要统计的图幅的最小和最大序号
    minSequenceValue = 1
    maxSequenceValue = 22
    # 设置每日统计点的Excel文件名，例如：20211010点统计.xlsx
    dailyReportExcelFilename = date_str + "点统计.xlsx"
    _100kSheetNames = (r"D:\RouteDesigen\PythonRun\100K_sheet_names_271_name_V3_after_GEOSA_edit.xlsx")
    df = pd.read_excel(_100kSheetNames, sheet_name="Sheet1", header=0, index_col=0)


    # 将日期字符串转换为datetime对象
    date = datetime.strptime(date_str, "%Y%m%d")
    # 格式化日期字符串，例如：2021/1/1
    formatted_date_str = date.strftime("%Y/%#m/%#d")
    # 格式化日期字符串，例如：年月日2021010
    yearAndmonth_str = date.strftime("%Y%m")
    # 格式化日期字符串，例如：年，2021
    year_str = date.strftime("%Y")
    # 格式化日期字符串，例如：月，10
    month_str = date.strftime("%m")
    # 当前的微信聊天记录文件夹
    currentWechatDic_str = os.path.join(r"D:\Users\lenovo\Documents\WeChat Files\WeChat Files\bringsmile\FileStorage\File", year_str + "-" + month_str)

    # 拉丁文地名（文件名）
    romanNames, quadangleFilenames = get_columns_in_range(df, minSequenceValue, maxSequenceValue)

    # 创建年月、年月日和kmz文件存储目录并输出返回值
    print(create_directory(os.path.join(workspace, yearAndmonth_str, date_str)))
    print(create_directory(os.path.join(workspace, yearAndmonth_str, date_str, "Planned routes")))
    if os.path.exists(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points")):
        shutil.rmtree(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"))
        os.makedirs(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"))  # 重新创建空目录
    print(create_directory(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points")))

    # 写入空统计表格
    if os.path.exists(os.path.join(workspace, yearAndmonth_str, date_str, dailyReportExcelFilename)):
        os.remove(os.path.join(workspace, yearAndmonth_str, date_str, dailyReportExcelFilename))
    create_daily_statistics_excel(formatted_date_str, romanNames, os.path.join(workspace, yearAndmonth_str, date_str, dailyReportExcelFilename))

    # ------------------- 拷贝微信聊天记录文件夹中的kmz文件相应的"Finished points"文件夹和"Planned routes"文件夹------------------------------
    # ------------------- 拷贝微信聊天记录文件夹中当天的的kmz文件至当天的Finished points文件夹------------------------------------------------
    files = list_fullpath_of_files_with_keywords(currentWechatDic_str, [date_str, "finished_points_and_tracks"])
    # print(f"微信聊天记录文件夹中的文件{len(files)}个""：", files)
    for file in files:
        copy_file(file, os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"))
        # print(file, "已拷贝至", date_str, "Finished points")
    # 将Finished points文件夹下的所有文件属性改为可读写
    make_files_read_write(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"))
    # 列出并删除有重复的文件，只保留最新的文件
    duplicateFiles_dict = find_files_with_max_number(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"))
    # print(duplicateFiles_dict)
    for base_name, (max_file_path, max_number) in duplicateFiles_dict.items():
        for root, _, filenames in os.walk(workspace, yearAndmonth_str, date_str, "Finished points"):
            for filename in filenames:
                if re.sub(r'\(\d+\)', '', filename) == base_name:
                    number = get_number_in_parentheses(filename)
                    file_path = os.path.join(root, filename)
                    if number < max_number:
                        os.remove(file_path)
                        # print(f"已删除文件: {file_path}")
    for base_name, (file_path, number) in duplicateFiles_dict.items():
        new_file_path = os.path.join(os.path.dirname(file_path), base_name)
        os.rename(file_path, new_file_path)
        # print(f"已重命名文件: {file_path} 为 {new_file_path}")
    print(f"{date_str}的完成的共有", len(list_fullpath_of_files_with_keywords(os.path.join(workspace, yearAndmonth_str, date_str, "Finished points"), [date_str, "finished_points_and_tracks"])), "组")

    # 拷贝下一个工作日的plan文件至下一个工作日的Planned routes文件夹中
    # 思路为：从当前日期开始，向后查找文件名字后8位为下一个工作日的文件，拷贝至下一个工作日的Planned routes文件夹中


    currentDay_str = date_str
    currentDate = datetime.strptime(currentDay_str, "%Y%m%d")
    formatted_currentDay_str = currentDate.strftime("%Y/%#m/%#d")
    currentYearAndMonth_str = currentDate.strftime("%Y%m")
    currentDayDir = os.path.join(workspace, yearAndmonth_str, currentDay_str, "Finished points")
    files = list_fullpath_of_files_with_keywords(currentWechatDic_str, ["plan"])
    # 循环查找下一个工作日的文件
    found_files = []
    # 下一天的日期字符串
    next_day_str = get_next_day(currentDay_str)
    nextDay_yearAndmonth_str = datetime.strptime(next_day_str, "%Y%m%d").strftime("%Y%m")
    while not found_files:
        # 查找包含下一天日期字符串的文件
        found_files = list_fullpath_of_files_with_keywords(currentWechatDic_str, [next_day_str, "plan"])
        days_differece = datetime.strptime(next_day_str, "%Y%m%d") - datetime.strptime(currentDay_str, "%Y%m%d")
        if found_files:
            print(create_directory(os.path.join(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes")))
            for file in found_files:
                copy_file(file, os.path.join(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes"))
                # print(file, "已拷贝至", next_day_str, "Planned routes")
            make_files_read_write(os.path.join(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes"))
            # 列出并删除有重复的文件，只保留最新的文件
            duplicateFiles_dict = find_files_with_max_number(os.path.join(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes"))
            # print(duplicateFiles_dict)
            for base_name, (max_file_path, max_number) in duplicateFiles_dict.items():
                for root, _, filenames in os.walk(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes"):
                    for filename in filenames:
                        if re.sub(r'\(\d+\)', '', filename) == base_name:
                            number = get_number_in_parentheses(filename)
                            file_path = os.path.join(root, filename)
                            if number < max_number:
                                os.remove(file_path)
                                # print(f"已删除文件: {file_path}")
            for base_name, (file_path, number) in duplicateFiles_dict.items():
                new_file_path = os.path.join(os.path.dirname(file_path), base_name)
                os.rename(file_path, new_file_path)
                # print(f"已重命名文件: {file_path} 为 {new_file_path}")
            print(f"{next_day_str}的计划共有", len(list_fullpath_of_files_with_keywords(os.path.join(workspace, nextDay_yearAndmonth_str, next_day_str, "Planned routes"), [next_day_str, "plan"])), "组")
        if not found_files and days_differece.days < 5:
            next_day_str = get_next_day(next_day_str)
            continue
        elif days_differece.days >= 5:
            print("五天内无下一天工作计划")
            break
    # 结束------------------- 拷贝微信聊天记录文件夹中的kmz文件相应的"Finished points"文件夹和"Planned routes"文件夹-------------------


    # 目前还缺少的上一天数据补齐拷贝至当天
    # 循环直到找到所有缺失的文件
    currentDay_str = date_str
    currentDate = datetime.strptime(currentDay_str, "%Y%m%d")
    formatted_currentDay_str = currentDate.strftime("%Y/%#m/%#d")
    currentYearAndMonth_str = currentDate.strftime("%Y%m")
    currentDayDir = os.path.join(workspace, yearAndmonth_str, currentDay_str, "Finished points")

    files = list_fullpath_of_files_with_keywords(currentDayDir, [currentDay_str, "kmz"])
    # print(f"{currentDay_str}共完成", len(files), "组:", files)
    str = "_finished_points_and_tracks"
    Quadangles_truncate_list = truncate_list(files, str)
    # print("Quadangles_truncate_list:", Quadangles_truncate_list)

    while date > datetime.strptime("20240901", "%Y%m%d"):
        files = list_all_files_with_ext(currentDayDir, ".kmz")
        Quadangles_truncate_list = truncate_list(files, str)
        date -= timedelta(days=1)
        date_str = date.strftime("%Y%m%d")
        yearAndmonth_str = date.strftime("%Y%m")
        previousDayDir = os.path.join(workspace, yearAndmonth_str, date_str, "Finished points")
        # print("被比较的文件目录:", previousDayDir)
        files = list_all_files_with_ext(previousDayDir, ".kmz")
        # print("previousDayDir文件", len(files), "：", files)
        set1 = set(Quadangles_truncate_list)
        # print("set1:", set1)
        set2 = set(quadangleFilenames)
        diffs = set2 - set1
        # print("不同的的文件数量", len(diffs), "个：", diffs)
        for diff in diffs:
            # print("缺失的文件：", diff)
            missing_files = list_fullpath_of_files_with_keywords(previousDayDir, [diff, "finished_points_and_tracks"])
            # print("缺失的文件全路径为：", list_fullpath_of_files_with_keywords(previousDayDir, [diff]))
            for missing_file in missing_files:
                copy_file(missing_file, currentDayDir)
                # print(missing_file, "已拷贝至", currentDay_str, "Finished points")
        # 更新当前文件夹的文件列表
        files = list_all_files_with_ext(currentDayDir, ".kmz")
        # print(f"经过与{date_str}比较，{currentDayDir}更新后的文件数量", len(files), "个：", files)
        Quadangles_truncate_list = truncate_list(files, str)
        # print("更新后的Quadangles_truncate_list:", Quadangles_truncate_list)
        # if len(Quadangles_truncate_list) == len(quadangleFilenames):
        #     break

    # print("所有当天无路线的组，已补全最近的文件")

    # ---------------------------------使用GDAL统计KMZ文件中的点要素数量---------------------------------
    dict = {}
    for quadangleFilename in quadangleFilenames:
        kmz_file = list_fullpath_of_files_with_keywords(os.path.join(workspace, currentYearAndMonth_str, currentDay_str, "Finished points"), [quadangleFilename, currentDay_str, 'kmz'])
        if len(kmz_file) == 0:
            # print(f"{currentDay_str} 无 {quadangleFilename} KMZ")
            dict[quadangleFilename] = {"Today": None, "Previous Day": None}
        else:
            kmz_file = kmz_file[0]
            point_feature_counts = count_num_points_in_kmz(kmz_file)
            dict[quadangleFilename] = {"Today": point_feature_counts, "Previous Day": None}

            # 回溯查找上一天的文件
            previous_date = currentDate - timedelta(days=1)
            found_previous = False
            while not found_previous and previous_date >= datetime.strptime("20240901", "%Y%m%d"):
                previousDate_str = previous_date.strftime("%Y%m%d")
                previousYearAndmonth_str = previous_date.strftime("%Y%m")
                previous_kmz_file = list_fullpath_of_files_with_keywords(os.path.join(workspace, previousYearAndmonth_str, previousDate_str, "Finished points"), [quadangleFilename, previousDate_str, 'kmz'])
                if len(previous_kmz_file) == 1:
                    previous_kmz_file = previous_kmz_file[0]
                    previous_point_feature_counts = count_num_points_in_kmz(previous_kmz_file)
                    dict[quadangleFilename]["Previous Day"] = previous_point_feature_counts
                    found_previous = True
                previous_date -= timedelta(days=1)
                    # 计算 Today 和 Previous Day 的差值
            if dict[quadangleFilename]["Today"] is not None and dict[quadangleFilename]["Previous Day"] is not None:
                if dict[quadangleFilename]["Today"] - dict[quadangleFilename]["Previous Day"] > 0:
                    dict[quadangleFilename]["Daily Increase"] = dict[quadangleFilename]["Today"] - dict[quadangleFilename]["Previous Day"]
                # else:
                #     dict[quadangleFilename]["Daily Increase"] = None
            else:
                dict[quadangleFilename]["Daily Increase"] = None
    # 补齐三列数据：确保字典中的所有键值对都包含 `Today`、`Previous Day` 和 `Daily Increase` 键
    for key in dict.keys():
        if 'Today' not in dict[key]:
            dict[key]['Today'] = None
        if 'Previous Day' not in dict[key]:
            dict[key]['Previous Day'] = None
        if 'Daily Increase' not in dict[key]:
            dict[key]['Daily Increase'] = None

    # print(dict)

    # 累加字典中所有的 `Daily Increase`
    total_daily_increase = sum(value['Daily Increase'] for value in dict.values() if value['Daily Increase'] is not None)
    print(f"总的 Daily Increase: {total_daily_increase}")

    completed_points = [value["Daily Increase"] for key, value in dict.items()]
    completed_points_series = pd.Series(completed_points)
    excel_filename = os.path.join(workspace, currentYearAndMonth_str, currentDay_str, dailyReportExcelFilename)
    book = load_workbook(excel_filename)
    sheet = book['Daily Statistics']
    # 将数据写入到指定的单元格
    for i, value in enumerate(completed_points_series, start=4):
        sheet.cell(row=i, column=2, value=value)
    # 保存工作簿
    book.save(excel_filename)
    os.startfile(os.path.join(workspace, currentYearAndMonth_str, currentDay_str, dailyReportExcelFilename))

    return None


def parse_args():
    # 无参数输入时，默认date_str为当天，有参数时，为指定参数
    parser = argparse.ArgumentParser(description="处理日期字符串")
    parser.add_argument("date_str", nargs='?', default=datetime.now().strftime("%Y%m%d"), type=str, help="日期字符串，格式为 YYYYMMDD")
    return parser.parse_args()

def main():
    args = parse_args()
    date_str = args.date_str
    daily_statistics(date_str)

if __name__ == "__main__":
    main()