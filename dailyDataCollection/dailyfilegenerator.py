import os
import re
import zipfile
import shutil
import hashlib
import xmlschema
import logging
import time
import json
import stat
import pandas as pd
from dataclasses import dataclass, field
from lxml import etree
from datetime import datetime, timedelta
from config import *
from abc import ABC, abstractmethod
from osgeo import ogr, osr
from openpyxl.styles import *
from openpyxl import Workbook, load_workbook
from tabulate import tabulate


# 创建 logger 实例
logger = logging.getLogger('Daily File Generator')
logger.setLevel(logging.INFO)
handler = logging.StreamHandler()
formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
handler.setFormatter(formatter)
logger.addHandler(handler)


#返回当前目录下所有含多个特定字符串(不区分大小写)列表的文件全路径
def list_fullpath_of_files_with_keywords(directory: str, keywords: list) -> list:
    matches = []
    for root, _, files in os.walk(directory):
        for file in files:
            if all(keyword.lower() in file.lower() for keyword in keywords):
                matches.append(os.path.join(root, file))
    return matches

def find_files_with_max_number(directory: str) -> dict:
    """
    从指定目录中查找包含相同基础文件名的文件，并返回括号中数字最大的文件
    """
    files_dict = {}
    # 遍历目录中的所有文件
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            # 去掉括号中的数字，例如：'file(1).txt' 返回 'file.txt'
            base_name = re.sub(r'\(\d+\)', '', filename)
            # 在文件名中查找括号内的数字。正则表达式 `r'\((\d+)\)'` 匹配括号内的一个或多个数字，并返回一个 `Match` 对象。如果没有找到匹配项，则返回 `None`
            match = re.search(r'\((\d+)\)', filename)
            # 从 `Match` 对象中提取括号内的数字，并将其转换为整数。如果找到了匹配项，则使用 `int(match.group(1))` 提取括号内的数字；否则，返回 -1
            # 提取括号中的数字，例如：'file(1).txt' 返回 1，'file.txt' 返回 -1
            number = int(match.group(1)) if match else -1        
            file_path = os.path.join(root, filename)
            # 如果文件名不在字典中，则将其添加到字典中
            if base_name not in files_dict:
                files_dict[base_name] = (file_path, number)
            # 如果文件名在字典中，则比较括号中的数字，保留数字最大的文件
            else:
                existing_file, existing_number = files_dict[base_name]
                if number > existing_number:
                    files_dict[base_name] = (file_path, number)
    # 清除括号中数字为-1的文件
    files_dict = {base_name: (file_path, number) for base_name, (file_path, number) in files_dict.items() if number != -1}
    # 输出包含相同基础文件名中括号内数字最大的文件
    return files_dict


class FileIO(ABC):
    def __init__(self, filepath=None):
        if filepath is not None and os.path.exists(filepath) and os.path.isfile(filepath):
            self.filepath = filepath
        else:
            logger.error(f"文件路径无效（为空/不存在/不是有效文件路径）: {filepath}")
            self.filepath = None

    @abstractmethod
    def read(self):
        pass

    @abstractmethod
    def write(self, content):
        pass

    @abstractmethod
    def delete(self):
        pass

    @abstractmethod
    def update(self, content):
        pass


class GeneralIO(FileIO):

    def __init__(self, filepath=None, mode='r', encoding = 'utf-8'):
        super().__init__(filepath)
        self.data = None
        self.mode = mode
        self.encoding = encoding

    def read(self, method='r', encoding='utf-8'):
        if os.path.exists(self.filepath) and os.path.isfile(self.filepath):
            with open(self.filepath, 'r', encoding='utf-8') as file:
                return file.read()

    def write(self, content):
        if os.path.exists(self.filepath) and os.path.isfile(self.filepath):
            with open(self.filepath, 'w', encoding='utf-8') as file:
                file.write(content)

    def write_as(self, newpath: str = None):
        """
        write_as 另存为新文件
        另存为新文件，将当前文件复制到新路径，将内存中的文件更新并更新当前文件路径
        :param newpath: 新文件路径
        :type newpath: str
        :raises FileNotFoundError: _description_
        """
        if not newpath:
            logger.error(f"新文件路径为空: {newpath}")
            return False
        if os.path.isdir(newpath):
            logger.error(f"文件路径{newpath}是一个目录，不是文件")
            return False
        #TODO: 另存为新文件，需要利用os.chmod(dest, stat.S_IWRITE | stat.S_IREAD)解决权限问题？
        # Step 1: 从当前文件路径复制文件到新路径
        # 先检查新路径是否存在，如果存在，询问用户是否覆盖
        if os.path.exists(newpath) and os.path.isfile(newpath):
            user_input = input(f"文件 '{newpath}' 已存在，是否覆盖？(YES(Y)/NO(N)): ").lower()
            if user_input not in ['yes', 'y']:
                print("操作已取消")
                return False
            shutil.copy(self.filepath, newpath)
        else:
            if not os.path.exists(os.path.dirname(newpath)):
                os.makedirs(os.path.dirname(newpath), exist_ok=True)
            shutil.copy(self.filepath, newpath)
        return True

    def delete(self):
        if os.path.exists(self.filepath) and os.path.isfile(self.filepath):
            os.remove(self.filepath)
        else:
            print(f"File delete fail: file \'{self.filepath}\' do not exist")
        
    def update(self, content):
        pass

    def __enter__(self):
        # 打开文件
        self.data = open(self.filepath, self.mode, self.encoding)
        return self.data
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        if self.data:
            self.data.close()
        return False


class FileAttributes(object):
    def __init__(self, filepath=None):
        # Basic file attributes
        if filepath and os.path.exists(filepath) and os.path.isfile(filepath):
            self._filepath = filepath
            self._filename = os.path.basename(filepath)
            self._file_dir = os.path.dirname(filepath)
            self._size = os.path.getsize(filepath)
            self._creation_time = time.ctime(os.path.getctime(filepath))
            self._modification_time = time.ctime(os.path.getmtime(filepath))
            self._access_time = time.ctime(os.path.getatime(filepath))
            self._file_type = self.__get_file_type()
        else:
            logger.error(f"文件路径无效（为空/不存在/不是有效文件路径）: {filepath}")
            self._filepath = None
            self._filename = None
            self._file_dir = None
            self._size = None
            self._creation_time = None
            self._modification_time = None
            self._access_time = None
            self._file_type = None
        # File content
        self._data = None
        # hash value
        self._hashMD5 = None
        self._hashSHA265 = None

    def __getattr__(self, name):
        if name == 'filepath':
            if self._filepath and os.path.exists(self._filepath) and os.path.isfile(self._filepath):
                return self._filepath
            return None
        if name == 'filename':
            if self._filename is None and self._filepath:
                self._filename = os.path.basename(self._filepath)
            return self._filename
        elif name == 'file_dir':
            if self._file_dir is None and self._filepath:
                self._file_dir = os.path.dirname(self._filepath)
            return self._file_dir
        elif name == 'size':
            if self._size is None and self._filepath:
                self._size = os.path.getsize(self._filepath)
            return self._size
        elif name == 'creation_time':
            if self._creation_time is None and self._filepath:
                self._creation_time = time.ctime(os.path.getctime(self._filepath))
            return self._creation_time
        elif name == 'modification_time':
            if self._modification_time is None and self._filepath:
                self._modification_time = time.ctime(os.path.getmtime(self._filepath))
            return self._modification_time
        elif name == 'access_time':
            if self._access_time is None and self._filepath:
                self._access_time = time.ctime(os.path.getatime(self._filepath))
            return self._access_time
        elif name == 'file_type':
            if self._file_type is None and self._filepath:
                self._file_type = self.__get_file_type()
            return self._file_type
        elif name == 'data':
            if self._data is None and self._filepath:
                with open(self._filepath, 'r') as file:
                    self._data = file.read()
            return self._data
        elif name == 'hashMD5':
            if self._filepath:
                if self._hashMD5 is None and self._data is None:
                    # Read file content
                    with open(self._filepath, 'rb') as file:
                        self._data = file.read()
                    self._hashMD5 = hashlib.md5(self._data).hexdigest()
            else:
                print(f"File path is None, hashMD5 is None")
                return None
            return self._hashMD5
        elif name == 'hashSHA265':
            if self._filepath:
                if self._hashSHA265 is None and self._data is None:
                    # Read file content
                    with open(self._filepath, 'rb') as file:
                        self._data = file.read()
                    self._hashSHA265 = hashlib.sha256(self._data).hexdigest()
            else:
                print(f"File path is None, hashSHA265 is None")
                return None
            return self._hashSHA265
        else:
            # print(f"AttributeError: {name} is not a valid attribute")
            return None


    def __get_file_type(self):
        if self._filepath and os.path.exists(self._filepath) and os.path.isfile(self._filepath):
            if os.path.splitext(self._filepath)[1]:
                return os.path.splitext(self._filepath)[1]
            else:
                return None
        else:
            return None
    
    def __str__(self):
        return f"File Name: \t\t{self._filename}\n" \
               f"File Path: \t\t{self._filepath}\n" \
               f"File Size: \t\t{self.size} bytes\n" \
               f"Creation Time: \t\t{self._creation_time}\n" \
               f"Modification Time: \t{self._modification_time}\n" \
               f"Access Time: \t\t{self._access_time}\n" \
               f"File type: \t\t{self._file_type}\n"\
               f"MD5: \t\t\t{self._hashMD5}\n" \
               f"SHA265: \t\t{self._hashSHA265}"


@dataclass
class ObservationData:

    OSPID_PATTERN = r'\d{5}[A-Za-z]\d{3}'
    LONGITUDE_PATTERN = r'<td>Longitude</td>\s*<td>(-?\d+\.\d+)</td>'
    LATITUDE_PATTERN = r'<td>Latitude</td>\s*<td>(-?\d+\.\d+)</td>'

    points: dict = field(default_factory=dict)
    pointsCount: int = 0
    routes: list = field(default_factory=list)
    routesCount: int = 0
    __errorMsg: list = field(default_factory=list)

    kml_content: str = None

    def __post_init__(self):
        if self.kml_content:
            self.__getPoints()
            self.__getRoutes()
            # 在解析KML内容后删除KML内容，以释放内存
            del self.kml_content
            # 检查点号
            self.__pointCheck()

    def __getPoints(self) -> None: 
        """
        从KML内容中提取点要素，包括OBSID、经度和纬度
        """
        root = etree.fromstring(self.kml_content, parser=etree.XMLParser(encoding='utf-8',recover=True))
        # Step 1: 通过Point元素提取点要素，Point元素下的name元素为OBSID，coordinates元素为经纬度，name元素可能为空或不符合命名规范，coordinates元素一定存在
        for placemark in root.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
            point = placemark.find('.//{http://www.opengis.net/kml/2.2}Point')
            if point is not None:
                name = placemark.find('.//{http://www.opengis.net/kml/2.2}name')
                if name is not None and name.text:
                    obspid_pattern = re.compile(ObservationData.OSPID_PATTERN)
                    if obspid_pattern.match(name.text):
                        obspid = name.text
                        coordinates = point.find('.//{http://www.opengis.net/kml/2.2}coordinates').text.split(',')
                        longitude, latitude = coordinates[0], coordinates[1]
                        self.points[obspid] = {'longitude': longitude, 'latitude': latitude}
                    else:
                        error = f"点要素{name.text}的标签格式不符合OBSID命名规范"
                        if type(self.__errorMsg) == list:
                            self.__errorMsg.append(error)
                        else:
                            self.__errorMsg = [error]
                        # logger.warning(error)
        # Step 2: 通过Description元素提取点要素，Description元素下的内容可能包含OBSID、经度和纬度3个要素，OBSID、经度或纬度均可能为空
        for description in root.findall('.//{http://www.opengis.net/kml/2.2}description'):
            if description is not None and description.text:
                obspidPattern = re.compile(ObservationData.OSPID_PATTERN)
                longitudePattern = re.compile(ObservationData.LONGITUDE_PATTERN)
                latitudePattern = re.compile(ObservationData.LATITUDE_PATTERN)
                obspid_match = obspidPattern.search(description.text)
                longitude_match = longitudePattern.search(description.text)
                latitude_match = latitudePattern.search(description.text)
                obspid = obspid_match.group(0) if obspid_match else None
                longitude = float(longitude_match.group(1)) if longitude_match else None
                latitude = float(latitude_match.group(1)) if latitude_match else None
            # if obspid == None or longitude == None or latitude == None:
            #     print(f"点要素没有OBSID{obspid}或经度{longitude}或纬度{latitude}")
                if obspid and longitude and latitude:
                    if not obspid in self.points:
                        self.points[obspid] = { 'longitude': longitude, 'latitude': latitude}
                # 如果OBSID、经度和纬度中有一个为空，则记录日志
                if obspid and not longitude and not latitude:
                    error = f"点要素{obspid}的属性表中缺少经度值和纬度值"
                    if type(self.__errorMsg) == list:
                        self.__errorMsg.append(error)
                    else:
                        self.__errorMsg = [error]
                    # logger.warning(error)
        # Step 3: 删除缺少经度或纬度的点要素
        # for obspid, coords in self.points.items():
        #     if coords['longitude'] == None or coords['latitude'] == None:
        #         del self.points[obspid]
        #         error = f"点要素{obspid}缺少经度或纬度"
        #         self.errorMsg.append(error)
        #         print(error)
        self.pointsCount = len(self.points)

    def __getRoutes(self) -> None:
        # 查找所有的Placemark元素下的LineString元素下的coordinates元素，提取线要素
        root = etree.fromstring(self.kml_content, parser=etree.XMLParser(encoding='utf-8',recover=True))
        for placemark in root.findall('.//{http://www.opengis.net/kml/2.2}Placemark'):
            linestring = placemark.find('.//{http://www.opengis.net/kml/2.2}LineString')
            if linestring is not None:
                coordinates = linestring.find('.//{http://www.opengis.net/kml/2.2}coordinates')
                if coordinates is not None:
                    self.routes.append(coordinates.text.strip())
        self.routesCount = len(self.routes)

    def __pointCheck(self) -> bool:
        # 通过第self.points字典的键，第6位解析组号并存储在字典中
        # 得到一个字典，键为组号，值为OBSID
        obsptidStatistics_dict = {}
        for key in self.points:
            teamkey = key[5]
            if teamkey not in obsptidStatistics_dict:
                obsptidStatistics_dict[teamkey] = []
            obsptidStatistics_dict[teamkey].append(key)
        #TODO 计算前5位，判断是否有重复值？
        #TODO 可以不做，因为OBSID是唯一的，同时该类可以是多个图幅的数据，不同图幅的图幅号不同
        # 分别计算第六位相同的个数（每组的点数），以及第六位相同的all_matches列表元素后三位是否连续（小组点数是否连续）
        # 建立与statistics_dict相同的字典，键值为原键值+"组"
        statistics_dict = {}
        for key, values in obsptidStatistics_dict.items():
            count = len(values)
            values.sort(key=lambda x: int(x[-3:]))  # 按后三位数字排序，bool
            is_duplicate = len(values) != len(set(values)) # 检查是否有重复的点号，bool
            is_consecutive = all(int(values[i][-3:]) + 1 == int(values[i + 1][-3:]) for i in range(len(values) - 1))    # 检查是否连续
            # 遍历数组，检查相邻元素之间的差是否为1
            pointValues = [int(value[-3:]) for value in values]
            gaps = []
            for i in range(1, count):
                if pointValues[i] - pointValues[i - 1] != 1:
                    # 如果差不为1，记录间断的位置
                    gaps.append((f"{key}{pointValues[i - 1]}", f"{key}{pointValues[i]}"))
            # 如果有间断，输出间断的位置
            if gaps:
                msg = f"组{key}的点号间断位置：{gaps}"
            else:
                msg = None
            max_value = max(values, key=lambda x: int(x[-3:]))
            new_key = key + "组"
            statistics_dict[new_key] = { "完成点数": count, "最大列表元素": max_value, "点号是否重复": is_duplicate, "点号是否连续": is_consecutive, "间断位置": msg}
        # 检查所有“点号是否连续”是否为 True
        all_consecutive = all(value["点号是否连续"] for key, value in statistics_dict.items())
        # 检查所有“点号是否重复”是否为 False
        all_not_duplicate = all(not value["点号是否重复"] for key, value in statistics_dict.items())
        # 累加所有“完成点数”
        total_completed_points = sum(value["完成点数"] for key, value in statistics_dict.items())
        # 检查总完成点数是否等于所有“完成点数相加”
        total_points_match = total_completed_points == len(self.points)
        # 如果以上条件均满足，则输出“验证通过”
        if all_consecutive and all_not_duplicate and total_points_match:
            return True
        else:
            print(f"\nOBSID数据验证失败")
            print(statistics_dict, "\n")
            self.__errorMsg.append(statistics_dict)
            return False

    
    def __add__(self: 'ObservationData', other: 'ObservationData') -> 'ObservationData':
        """用加法操作实现两个文件数据合并，得到一个新的字典"""
        if not isinstance(other, ObservationData):
            raise TypeError("Operands must be of type 'ObservationData'")
        # 创建新的ObservationData对象
        new_file = ObservationData()
        # 合并点要素（相同的键进行了覆盖）
        new_file.points = {**self.points, **other.points}
        new_file.pointsCount = len(new_file.points)
        # 合并线要素（列表相加并去重）
        new_file.routes = list(set(self.routes + other.routes))
        new_file.routesCount = len(new_file.routes)
        return new_file

    def __sub__(self, other: 'ObservationData') -> 'ObservationData':
        """用减法操作实现两个文件数据的差异，得到一个新的字典"""
        if not isinstance(other, ObservationData):
            logger.error("数据类型必须是'ObservationData'")
            return None
        # 验证一个字典的键是否是另一个字典键的子集
        if not (self.points.keys() <= other.points.keys() or other.points.keys() <= self.points.keys()):
            logger.error("KML文件中的点要素不是另一个的子集")
            return None
        # if not (set(self.routes) <= set(other.routes) or set(other.routes) <= set(self.routes)):
        #     raise ValueError("KML文件中的线要素不是另一个的子集")
        # 创建新的KMLFile对象
        new_file = ObservationData()
        # 计算点要素的差异
        new_file.points = {key: value for key, value in self.points.items() if key not in other.points}
        new_file.points.update({key: value for key, value in other.points.items() if key not in self.points})
        new_file.pointsCount = len(new_file.points)
        # 计算线要素的差异并去重
        # new_file.routes = list(set(self.routes).symmetric_difference(set(other.routes)))
        # new_file.routesCount = len(new_file.routes)
        return new_file
    
    @property
    def errorMsg(self):
        if type(self.__errorMsg) == list:
            return self.__errorMsg
        else:
            return None
    
    def __str__(self):
        return f"点要素数量: {self.pointsCount}\n"f"线要素数量: {self.routesCount}\n"f"错误信息: {self.__errorMsg}"


class KMZFile(FileAttributes, GeneralIO):

    SCHEMA_22 = KML_SCHEMA_22
    SCHEMA_23 = KML_SCHEMA_23

    
    def __init__(self, filepath: str = None, placemarks: ObservationData = None):
        super().__init__(filepath)
        # TODO: 如果注释掉，在main函数测试中，会在原1219行报错
        # TODO: File "D:\MacBook\MacBookDocument\SourceCode\GMAS\dailyDataCollection\DailyFileGenerator.py", line 1222, in totalPointNum
        # TODO: totalNum += mapsheet.lastPlacemarks.pointsCount
        # TODO: AttributeError: 'NoneType' object has no attribute 'pointsCount'
        self.filepath = filepath
        if self.filepath:
            if self.filepath.endswith(".kmz"):
                self.filepath = filepath
            else:
                self.filepath = None
                logger.error(f"文件后缀名无效: {filepath}")
        self._placemarks: ObservationData = placemarks

        self._kml_content = None
        self._resources: dict = {}
        self._points: dict = {}
        self._pointsCount: int = 0
        self._routes: list = []
        self._routesCount: int = 0
        # 用于存储错误信息
        self.__errorMsg: list = []

        if self.filepath:
            self.read()

        if placemarks:
            # print("通过ObservationData数据初始化")
            self._points = self._placemarks.points
            self._pointsCount = self._placemarks.pointsCount
            self._routes = self._placemarks.routes
            self._routesCount = self._placemarks.routesCount
            if type(self._placemarks.errorMsg) == list:
                self.__errorMsg.append(self._placemarks.errorMsg)
                print("KMZ初始化时发现的错误", self.errorMsg)
            
            # if self._placemarks.errorMsg:
            #     print(self._placemarks.errorMsg)
            #     self.__errorMsg.append(self._placemarks.errorMsg)

    def __validateKMZ(self, defaultSchema = "schema22") -> bool:
        """
        验证KMZ文件是否符合KML的XSD模式
        """
        if defaultSchema == "schema22":
            schema = xmlschema.XMLSchema(KMZFile.SCHEMA_22) 
        elif defaultSchema == "schema23":
            schema = xmlschema.XMLSchema(KMZFile.SCHEMA_22)
        else:
            return print(f"无效的输入参数, 将使用默认的'{defaultSchema}'")
        if self._kml_content is None:
            return print("KML内容为空")
        else:
            try:
                # print(f"开始验证XML文件与'{defaultSchema}'模式")
                schema.validate(self._kml_content)
                logger.info(f"XML文件与XSD'{defaultSchema}'验证通过")
                return True
            except xmlschema.validators.exceptions.XMLSchemaValidationError as e:
                # error = f"XML文件与XSD'{defaultSchema}'验证不符: {e}"
                warning = f"XML文件与XSD'{defaultSchema}'验证不符"
                logger.warning(warning)
                self.__errorMsg.append(warning)
                return False
    
    def read(self, filepath: str = None, validate = False, defaultSchema = "schema22"):
        """ 解压 KMZ 文件并提取 KML 内容 """
        if filepath is not None:
            if not os.path.exists(filepath) or not os.path.isfile(filepath):
                logger.error(f"文件路径不存在: {filepath}")
                return False
            if not filepath.endswith(".kmz"):
                logger.error(f"文件后缀名无效: {filepath}")
                return False
            if self.filepath is None:
                self.filepath = filepath
        elif self.filepath is None:
            logger.error("文件路径为空")
            return False
        else:
            filepath = self.filepath

        with zipfile.ZipFile(filepath, 'r') as kmz:
            # print(f"解压 KMZ 文件: {filepath}")
            # 查找 KML 文件
            kml_files = [f for f in kmz.namelist() if f.endswith('.kml')]
            # print(f"KMZ 文件中的 KML 文件: {kml_files}")
            if kml_files:
                kml_file = kml_files[0]
                # 读取 KML 文件内容
                with kmz.open(kml_file) as kml:
                    self._kml_content = kml.read()
                    # 验证KML文件是否符合XSD模式
                    if validate:
                        self.__validateKMZ(defaultSchema)
                    # 解析KML内容
                    placemarks = ObservationData(kml_content=self._kml_content)
                    self._points = placemarks.points
                    self._pointsCount = placemarks.pointsCount
                    self._routes = placemarks.routes
                    self._routesCount = placemarks.routesCount
                    if placemarks.errorMsg:
                        self.__errorMsg.append(placemarks.errorMsg)
                        # print("KMZ读取时候发现的错误", self.errorMsg)
                    self._placemarks = placemarks
                    #! 删除KML内容，释放内存
                    del self._kml_content
            else:
                logger.error("在KMZ文件中没有找到 KML文件")

    def write(self, type='kmz'):
        if self._filepath is not None:
            os.makedirs(os.path.dirname(self._filepath), exist_ok=True)
            if type == 'kmz':
                if self.__toKMZ(self._filepath):
                    print(f"KMZ文件已保存到: {self._filepath}")
                    return True
                else:
                    print(f"KMZ文件保存失败")
                    logger.error(f"KMZ文件保存失败{self._filepath}")
                    return False
            elif type == 'shp':
                if self.__toShp(self._filepath):
                    print(f"SHP文件已保存到: {self._filepath}")
                    return True
                else:
                    print(f"SHP文件保存失败")
                    logger.error(f"SHP文件保存失败{self._filepath}")
                    return False
            else:
                # print(f"无效的输出文件类型: {type}")
                logger.error(f"无效的输出文件类型: {type}")
                return False
        else:
            print("文件路径为空")
            logger.error("文件路径为空")
            return False

    def write_as(self, newpath: str = None):
        if os.path.exists(newpath) and os.path.isfile(newpath):
            logger.warning(f"文件路径{newpath}将覆盖原文件")
        filetype = os.path.splitext(newpath)[1].lower()
        if filetype == '.kmz':
            if self.__toKMZ(newpath):
                print(f"KMZ文件已保存到: {newpath}")
                return True
            else:
                print(f"KMZ文件保存失败")
                logger.error(f"KMZ文件保存失败{newpath}")
                return False
        elif filetype == '.shp':
            if self.__toShp(newpath):
                print(f"SHP文件已保存到: {newpath}")
                return True
            else:
                print(f"SHP文件保存失败")
                logger.error(f"SHP文件保存失败{newpath}")
                return False
        else:
            print(f"无效的输出文件类型: {filetype}")
            logger.error(f"无效的输出文件类型: {filetype}")
            return False


    def delete(self):
        return super().delete()
    
    def update(self, content):
        return super().update(content)
    
    def __toKMZ(self, output_path):
        """ 将 KML 数据写入 KMZ 文件 """
        if self.placemarks is None:
            logger.error("ObservationData is empty")
            return False
        # 创建输出目录
        if os.path.exists(output_path) and os.path.isfile(output_path):
            # 验证output_path是否已.kmz结尾
            if not output_path.endswith('.kmz'):
                logger.error("Output file must be a KMZ file")
                return False
        output_dir = os.path.dirname(output_path)
        # 在输出目录下创建临时目录
        temp_dir = os.path.join(output_dir, "temp_kmz")
        os.makedirs(temp_dir, exist_ok=True)
        files_dir = os.path.join(temp_dir, 'files')
        os.makedirs(files_dir, exist_ok=True)
        # 创建临时KML文件
        temp_kml_file = os.path.join(output_dir, "temp_combined.kml")
        # 创建输出KMZ文件
        # 根节点
        document = etree.Element("Document")
        # 添加线样式
        style = etree.SubElement(document, "Style", id="lineStyle")
        line_style = etree.SubElement(style, "LineStyle")
        color = etree.SubElement(line_style, "color")
        color.text = "ff000000"  # 黑色，格式为 AABBGGRR
        width = etree.SubElement(line_style, "width")
        width.text = "1"  # 线宽度
        # 添加线数据
        for i, coordinates in enumerate(self.routes):
            placemark = etree.SubElement(document, "Placemark")
            style_url = etree.SubElement(placemark, "styleUrl")
            style_url.text = "#lineStyle"
            name = etree.SubElement(placemark, "name")
            name.text = f"Route {i+1}"
            linestring = etree.SubElement(placemark, "LineString")
            coord_elem = etree.SubElement(linestring, "coordinates")
            coord_elem.text = coordinates
        # 添加点样式
        style = etree.SubElement(document, "Style", id="pointStyle")
        icon_style = etree.SubElement(style, "IconStyle")
        icon = etree.SubElement(icon_style, "Icon")
        href = etree.SubElement(icon, "href")
        href.text = "files/Layer0_Symbol_Square.png"  # 使用自定义图标的 URL
        # 添加点要素
        for obspid, coords in self.points.items():
            placemark = etree.SubElement(document, "Placemark")
            style_url = etree.SubElement(placemark, "styleUrl")
            style_url.text = "#pointStyle"
            name = etree.SubElement(placemark, "name")
            name.text = obspid
            point = etree.SubElement(placemark, "Point")
            coordinates = etree.SubElement(point, "coordinates")
            coordinates.text = f"{coords['longitude']},{coords['latitude']}"
        # 保存KML文件
        tree = etree.ElementTree(document)
        tree.write(temp_kml_file, pretty_print=True, xml_declaration=True, encoding="UTF-8")
        # 将PNG文件移动到files文件夹中
        shutil.copy(ICON_1, files_dir)
        os.chmod(files_dir, stat.S_IWRITE | stat.S_IREAD)
        # 压缩为KMZ文件
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as kmz:
        # 将doc.kml文件添加到KMZ文件中
            kmz.write(temp_kml_file, 'doc.kml')
        # 将files文件夹中的PNG文件添加到KMZ文件中
            for root, _, files in os.walk(files_dir):
                for file in files:
                    file_path = os.path.join(root, file)
                    arcname = os.path.relpath(file_path, temp_dir)
                    kmz.write(file_path, arcname)
        # 删除临时目录
        shutil.rmtree(temp_dir)
        # 删除临时KML文件
        os.remove(temp_kml_file)
        return True
    
    def __toShp(self, output_path):
        # 创建 SHP 文件驱动
        shp_driver = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(output_path):
            shp_driver.DeleteDataSource(output_path)
        shp_ds = shp_driver.CreateDataSource(output_path)
        if shp_ds is None:
            raise RuntimeError("Failed to create SHP file")
        # 创建 SHP 图层
        srs = osr.SpatialReference()
        srs.ImportFromEPSG(4326)  # WGS84
        shp_layer = shp_ds.CreateLayer('points', srs, ogr.wkbPoint)
        # 创建字段
        field_name = ogr.FieldDefn('Name', ogr.OFTString)
        field_name.SetWidth(24)
        shp_layer.CreateField(field_name)
        field_longitude = ogr.FieldDefn('Longitude', ogr.OFTReal)
        shp_layer.CreateField(field_longitude)
        field_latitude = ogr.FieldDefn('Latitude', ogr.OFTReal)
        shp_layer.CreateField(field_latitude)
        # 创建要素并添加到图层
        for obspid, coords in self.placemarks.points.items():
            feature = ogr.Feature(shp_layer.GetLayerDefn())
            feature.SetField('Name', obspid)
            feature.SetField('Longitude', float(coords['longitude']))
            feature.SetField('Latitude', float(coords['latitude']))
            # 创建点要素
            point = ogr.Geometry(ogr.wkbPoint)
            point.AddPoint(float(coords['longitude']), float(coords['latitude']))
            feature.SetGeometry(point)
            # 添加要素
            shp_layer.CreateFeature(feature)
            feature = None  # 清理要素
        # 清理
        shp_ds = None
        print(f"点要素已成功生成 SHP 文件: {output_path}")
        return True
    
    def __getattr__(self, name):
        super().__getattr__(name)
        if name == 'placemarks':
            return self._placemarks
        elif name == 'points':
            return self._points
        elif name == 'pointsCount':
            return self._pointsCount
        elif name == 'routes':
            return self._routes
        elif name == 'routesCount':
            return self._routesCount
        elif name == 'errorMsg':
            if self.__errorMsg:
                return self.__errorMsg
            else:
                return None
        else:
            return None
    
    def __str__(self):
        super().__str__()
    
    def __add__(self, other: 'KMZFile') -> 'KMZFile':
        if not isinstance(other, KMZFile):
            raise TypeError("Operands must be of type 'KMZFile'")
        new_file = KMZFile()
        new_file.placemarks = self.placemarks + other.placemarks
        return new_file
    
    def __sub__(self, other: 'KMZFile') -> 'KMZFile':
        if not isinstance(other, KMZFile):
            raise TypeError("Operands must be of type 'KMZFile'")
        new_file = KMZFile()
        new_file.placemarks = self.placemarks - other.placemarks
        return new_file


class DateIterator(object):
    def __init__(self, start_date: str, direction: str = 'forward'):
        """
        __init__
            初始化迭代器，start_date为起始日期，direction为迭代方向
        args:
            start_date: 起始日期，格式为 'YYYYMMDD'
            direction: 迭代方向，可选值为 'forward' 或 'backward'
                        'forward': 从起始日期开始向后迭代，即向前（下一天）迭代
                        'backward': 从起始日期开始向前迭代，即向后（上一天）迭代
        """
        # 将字符串日期转换为 datetime 对象
        self.date = datetime.strptime(start_date, "%Y%m%d")
        self.direction = direction  # 'forward' 为下一天，'backward' 为前一天

    def __iter__(self):
        """返回迭代器自身"""
        return self

    def __next__(self):
        """根据迭代方向返回下一个日期"""
        if self.direction == 'forward':
            self.date += timedelta(days=1)
        elif self.direction == 'backward':
            self.date -= timedelta(days=1)
        else:
            raise ValueError("Invalid direction. Use 'forward' or 'backward'.")

        # 返回当前日期并格式化为字符串
        return self.date.strftime("%Y%m%d")

    def switch_direction(self):
        """切换迭代方向"""
        if self.direction == 'forward':
            self.direction = 'backward'
        else:
            self.direction = 'forward'

    def reset(self, new_start_date: str):
        """重置为新的起始日期"""
        self.date = datetime.strptime(new_start_date, "%Y%m%d")


class MapsheetDailyFile(object):

    maps_info: dict = {}

    def __new__(cls, *args, **kwargs):
        # 如果类属性中没有实例（第一次创建）
        # 尝试获取CurrentDateFiles.maps_info属性
        # 如果不存在，则自行初始化MapsheetsDailyFile.maps_info属性
        if not hasattr(cls, 'instance'):
            if CurrentDateFiles.maps_info:
                cls.maps_info = CurrentDateFiles.maps_info
            else:
                cls.maps_info = CurrentDateFiles.mapsInfo()
        cls.instance = super(MapsheetDailyFile, cls).__new__(cls)
        # cls.instance.__init__(*args, **kwargs)
        return cls.instance

    def __init__(self, mapsheetFileName: str, date: 'DateType'):

        self.mapsheetFileName:str = mapsheetFileName
        if not self.__mapsheetInfo(mapsheetFileName):
            print(f"图幅文件名称{mapsheetFileName}未找到")
            self.sequence: int = None
            self.romanName: str = None
            self.latinName: str = None

        # 当前日期的文件属性
        self.currentDate: DateType = date
        self.currentfilename = None
        self.currentfilepath = None
        self.currentPlacemarks: ObservationData = None
        # 上一次提交的文件属性
        self.lastDate: DateType =  None
        self.lastfilename = None
        self.lastfilepath = None
        self.lastPlacemarks: ObservationData = None
        # 下一次提交的工作计划文件属性
        self.nextDate: DateType =  None
        self.nextfilename = None
        self.nextfilepath = None
        self.nextPlacemarks: ObservationData = None
        # 本日新增点数、线路数、点要素和线要素
        self.dailyincreasePointNum: int = None
        self.dailyincreaseRouteNum: int = None
        self.dailyincreasePoints: dict = {}
        self.dailyincreaseRoutes: list = []
        # 包含的错误信息
        self.__errorMsg: dict = {}
        # 截止当前日期的总点数、线路数
        self.currentTotalPointNum: int = None
        self.currentTotalRouteNum: int = None

        self.__mapsheetfiles()

    def __mapsheetInfo(self, mapsheetFileName):
        """
        根据 File Name 从 mapsheet_info 字典中获取 Sequence 并赋值给实例变量 self.sequence
        """
        if self.__class__.maps_info:
            for sequence, info in self.__class__.maps_info.items():
                if info['File Name'] == mapsheetFileName:
                    self.sequence = sequence
                    break
            else:
                print(f"文件名称未找到\n请查证是否文件名 '{mapsheetFileName}' 有误？\文件名列表如下：")
                print(json.dumps(self.__class__.maps_info, indent=4, ensure_ascii=False))
                print("程序退出")
                exit()
                # raise ValueError(f"File Name {mapsheetFileName} not found in mapsheet_info")
            self.romanName = self.__class__.maps_info[self.sequence]['Roman Name']
            self.latinName = self.__class__.maps_info[self.sequence]['Latin Name']
            return True
        # else:
        #     print(f"类属性maps_info不存在，请先通过类方法'MapsheetDailyFile.mapsInfo()'获取图幅信息")
        return False

    @classmethod
    def getCurrentDateFile(cls, instance):
        """
        获取当天的文件，文件名格式为图幅名称+finished_points_and_tracks+日期+.kmz
        一般情况下，当天的文件是最新的文件，因此优先查找当天的文件，但微信中可能会有多个文件，后缀名为(1),(2)...等，因此需要进一步处理
                            处理方式为：()中的数字最大的文件为当天最新的文件，或者时间最新的文件为当天最新的文件？？？

        """
        file_path = os.path.join(WORKSPACE, instance.currentDate.yyyymm_str, instance.currentDate.yyyymmdd_str, "Finished points", f"{instance.mapsheetFileName}_finished_points_and_tracks_{instance.currentDate.yyyymmdd_str}.kmz")
        # print(f"开始查找{instance.currentDate.yyyymmdd_str}当天的文件")
        # Step 1: 在当天的工作文件夹中查找当天的文件
        if os.path.exists(file_path):
            instance.currentfilepath = file_path
        # Step 2: 在微信聊天记录文件夹中查找当天的文件，如果有更新的文件，则拷贝至工作文件夹，替换原有文件
        # 列出微信聊天记录文件夹中包含指定日期、图幅名称和finished_points的文件
        folder = os.path.join(WECHAT_FOLDER)
        searchedFile_list = list_fullpath_of_files_with_keywords(folder, [instance.currentDate.yyyymmdd_str, instance.mapsheetFileName, "finished_points_and_tracks", ".kmz"])
        # print(f"在微信记录中查找查找{instance.currentDate.yyyymmdd_str}当天的文件", searchedFile_list) 
        if searchedFile_list:
            #TODO: 用find_files_with_max_number函数二次验证获取的文件
            # 选择时间最新的文件(区别于用find_files_with_max_number函数，获取文件名中数字最大的文件)
            fetched_file = max(searchedFile_list, key=os.path.getctime)
            # 如果工作文件夹中的文件不存在
            if not os.path.exists(file_path):
                os.makedirs(os.path.dirname(file_path), exist_ok=True)
                shutil.copy(fetched_file, file_path)
                os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
                instance.currentfilepath = file_path
            else:
                if KMZFile(filepath=file_path).hashMD5 != KMZFile(filepath=fetched_file).hashMD5:
                    # 将获取的文件拷贝至工作文件夹，并进行了重命名
                    shutil.copy(fetched_file, file_path)
                    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
                    instance.currentfilepath = file_path
                else:
                    instance.currentfilepath = file_path
        if instance.currentfilepath:
            instance.currentfilename = os.path.basename(instance.currentfilepath)
            file = KMZFile(filepath=instance.currentfilepath)
            instance.currentPlacemarks = file.placemarks
            if file.errorMsg:
                instance.__errorMsg[instance.currentfilename] = file.errorMsg
        return cls
    
    @classmethod
    def findlastFinished(cls, instance):
        """
        """
        # print(f"开始查找{instance.currentDate.yyyymmdd_str}之前的文件")
        lastDate_datetime1 = instance.currentDate.date_datetime
        while lastDate_datetime1 > datetime.strptime(TRACEBACK_DATE, "%Y%m%d"):
            # Step 1: 在工作文件夹（当前日期）中直接查找，在对应的时间日期直接快速查找
            lastDate_datetime2 = instance.currentDate.date_datetime
            while lastDate_datetime2 > datetime.strptime(TRACEBACK_DATE, "%Y%m%d"):
                lastDate_datetime2 -= timedelta(days=1)
                search_date_str = lastDate_datetime2.strftime("%Y%m%d")
                file_path = os.path.join(WORKSPACE, lastDate_datetime1.strftime("%Y%m"), lastDate_datetime1.strftime("%Y%m%d"), "Finished points", f"{instance.mapsheetFileName}_finished_points_and_tracks_{search_date_str}.kmz")
                # print(f"在工作文件夹中查找{instance.mapsheetFileName}日期{file_path}的文件")
                if os.path.exists(file_path):
                    instance.lastDate = DateType(date_datetime=lastDate_datetime2)
                    #TODO: 将文件拷贝至当前日期文件夹中，需要补充
                    if instance.currentfilename is None:
                        dest = os.path.join(WORKSPACE, instance.currentDate.yyyymm_str, instance.currentDate.yyyymmdd_str, "Finished points", os.path.basename(file_path))
                        if file_path != dest:
                            os.makedirs(os.path.dirname(dest), exist_ok=True)
                            shutil.copy(file_path, dest)
                            os.chmod(dest, stat.S_IWRITE | stat.S_IREAD)
                            instance.lastfilepath = dest
                    else:
                        instance.lastfilepath = file_path
                    # print(f"在工作文件夹中找到{instance.lastDate.yyyymmdd_str}的文件: {instance.lastfilepath}")
                    break
            if instance.lastfilepath:
                break
            lastDate_datetime1 -= timedelta(days=1)
        else:
            # print(f"Step 1: 完成\n在工作文件夹中未找到上次{instance.mapsheetFileName}日期{search_date_str}的文件")
            # Step 2: 工作文件夹开启搜索
            lastDate_datetime = instance.currentDate.date_datetime
            folder = os.path.join(WORKSPACE)
            while lastDate_datetime > datetime.strptime(TRACEBACK_DATE, "%Y%m%d"):
                lastDate_datetime -= timedelta(days=1)
                # 定位到年月日文件夹中的Finished points文件夹
                date_folder = os.path.join(WORKSPACE, lastDate_datetime.strftime("%Y%m"), lastDate_datetime.strftime("%Y%m%d"), "Finished points")
                # 列出文件夹中包含图幅名称和finished_points的文件
                searchedFile_list = list_fullpath_of_files_with_keywords(date_folder, [instance.mapsheetFileName, "finished_points_and_tracks", ".kmz"]) 
                if searchedFile_list:
                    #TODO: 用find_files_with_max_number函数二次验证获取的文件（也许不用？一般是唯一值）
                    # 选择时间最新的文件(区别于用find_files_with_max_number函数，获取文件名中数字最大的文件)
                    fetched_file = max(searchedFile_list, key=os.path.getctime)
                    instance.lastDate = DateType(date_datetime=lastDate_datetime)
                    instance.lastfilepath = fetched_file
                    break
            else:
                # 如果在TRACEBACK_DATE之后未找到文件，在循环结束后在微信聊天记录文件夹中查找
                # 列出微信聊天记录文件夹中包含指定日期、图幅名称和finished_points的文件
                lastDate_datetime = instance.currentDate.date_datetime
                folder = os.path.join(WECHAT_FOLDER)
                while lastDate_datetime > datetime.strptime(TRACEBACK_DATE, "%Y%m%d"):
                    lastDate_datetime -= timedelta(days=1)
                    search_date_str = lastDate_datetime.strftime("%Y%m%d")
                    searchedFile_list = list_fullpath_of_files_with_keywords(folder, [search_date_str, instance.mapsheetFileName, "finished_points_and_tracks", ".kmz"]) 
                    if searchedFile_list:
                        #TODO: 用find_files_with_max_number函数二次验证获取的文件
                        # 选择时间最新的文件(区别于用find_files_with_max_number函数，获取文件名中数字最大的文件)
                        fetched_file = max(searchedFile_list, key=os.path.getctime)
                        instance.lastDate = DateType(date_datetime=lastDate_datetime)
                        instance.lastfilepath = fetched_file
                        break
        if instance.lastfilepath:
            instance.lastfilename = os.path.basename(instance.lastfilepath)
            file = KMZFile(filepath=instance.lastfilepath)
            instance.lastPlacemarks = file.placemarks
            if file.errorMsg:
                instance.__errorMsg[instance.lastfilename] = file.errorMsg
            # print(instance.lastPlacemarks)
        # print("正在获取", instance.lastfilepath)
        # print(instance.lastPlacemarks)
        return cls
    
    @classmethod
    def findNextPlan(cls, instance):
        """
        findNextPlan 通常应是查找第二天的计划文件：
                     或周五-周六休息，周四会查找周六/周日的计划文件，因此为了冗余日期，TRACEFORWARD_DAYS = 5，即向前查找5天，找到最近的一个计划文件
        """
        # print(f"开始查找{instance.currentDate.yyyymmdd_str}之后的计划文件")
        date = instance.currentDate.date_datetime
        endDate = instance.currentDate.date_datetime + timedelta(days=TRACEFORWARD_DAYS)
        while date < endDate:
            date += timedelta(days=1)
            search_date_str = date.strftime("%Y%m%d")
            # 列出微信聊天记录文件夹中包含指定日期、图幅名称和plan_tracks的文件
            #TODO: 搜索效率低，需要优化
            searchedFile_list = list_fullpath_of_files_with_keywords(WECHAT_FOLDER, [search_date_str, instance.mapsheetFileName, "plan_routes"])
            # print(searchedFile_list) 
            if searchedFile_list:
                # print(f"找到工作计划: {searchedFile_list}")
                instance.nextDate = DateType(date_datetime=date)
                file_path = os.path.join(WORKSPACE, instance.nextDate.yyyymm_str, instance.nextDate.yyyymmdd_str, "Planned routes", f"{instance.mapsheetFileName}_plan_routes_{instance.nextDate.yyyymmdd_str}.kmz")
                latestFile = max(searchedFile_list, key=os.path.getctime)
                if not os.path.exists(file_path):
                    # 选择时间最新的文件(区别于用find_files_with_max_number函数，获取文件名中数字最大的文件)
                    os.makedirs(os.path.dirname(file_path), exist_ok=True)
                    shutil.copy(latestFile, file_path)
                    os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
                    instance.nextfilepath = file_path
                    break
                else:
                    if KMZFile(filepath=file_path).hashMD5 != KMZFile(filepath=latestFile).hashMD5:
                        # 将获取的文件拷贝至工作文件夹，并进行重命名
                        shutil.copy(latestFile, file_path)
                        os.chmod(file_path, stat.S_IWRITE | stat.S_IREAD)
                        instance.nextfilepath = file_path
                        break
                    else:
                        instance.nextfilepath = file_path

                        break
        if instance.nextfilepath:
            instance.nextfilename = os.path.basename(instance.nextfilepath)
            file = KMZFile(filepath=instance.nextfilepath)
            instance.planPlacemarks = file.placemarks
            if file.errorMsg:
                instance.__errorMsg[instance.nextfilename] = file.errorMsg
        return cls
    
    def __mapsheetfiles(self):
        MapsheetDailyFile.getCurrentDateFile(self)
        MapsheetDailyFile.findlastFinished(self)
        MapsheetDailyFile.findNextPlan(self)
        MapsheetDailyFile.__dailyIncrease(self)
        MapsheetDailyFile.__Finished(self)
        return self
    
    def __Finished(self):
        """
        计算截止当天的总计点数和线路数
        """
        if self.currentPlacemarks is None and self.lastPlacemarks is None:
            self.currenttotalPointNum = '-'
            self.currenttotalRouteNum = '-'
        if self.currentPlacemarks is not None and self.lastPlacemarks is not None:
            self.currenttotalPointNum = len(self.currentPlacemarks.points)
            self.currenttotalRouteNum = len(self.currentPlacemarks.routes)
        if self.currentPlacemarks is None and self.lastPlacemarks is not None:
            self.currenttotalPointNum = len(self.lastPlacemarks.points)
            self.currenttotalRouteNum = len(self.lastPlacemarks.routes)
        return self
        
    
    def __dailyIncrease(self):
        """
        计算当天新增的点要素和线要素
        """
        if self.currentPlacemarks is None:
            dailyincreasePlacemarks = 0
            self.dailyincreasePointNum = 0
            self.dailyincreaseRouteNum = 0
        if self.currentPlacemarks and self.lastPlacemarks:
            dailyincreasePlacemarks = self.currentPlacemarks - self.lastPlacemarks
            self.dailyincreasePointNum = len(dailyincreasePlacemarks.points)
            self.dailyincreaseRouteNum = len(dailyincreasePlacemarks.routes)
        if self.currentPlacemarks is not None and self.lastPlacemarks is None:
            dailyincreasePlacemarks = self.currentPlacemarks
            self.dailyincreasePointNum = len(dailyincreasePlacemarks.points)
            self.dailyincreaseRouteNum = len(dailyincreasePlacemarks.routes)
            print(f"提示：{self.mapsheetFileName}是否为第一次提交？")
        return self
    
    @property
    def errorMsg(self):
        if self.__errorMsg:
            return self.__errorMsg
        else:
            return None
    
    def __str__(self):
        return f"图幅名称：{self.mapsheetFileName}\n当天文件: {self.currentfilename}\n上一次文件: {self.lastfilename}\n下一次文件: {self.nextfilename}\n当天新增点数: {self.dailyincreasePointNum}\n当天新增线路数: {self.dailyincreaseRouteNum}\n当天文件中存在的错误：{self.errorMsg}"
    

    
class CurrentDateFiles(object):
    """
    CurrentDateFiles类  容器类
                        用于存储制定日期的所有图幅的集合
    """
    maps_info: dict = {}

    def __new__(cls, currentdate: 'DateType', *args, **kwargs):
        if not hasattr(cls, 'instance'):
            cls.mapsInfo()
        cls.instance = super(CurrentDateFiles, cls).__new__(cls)
        # cls.instance.__init__(currentdate)
        return cls.instance

    def __init__(self, currentdate: 'DateType'):

        self.currentDate: DateType = currentdate
        self.currentDateFiles: dict = []
        # 本日新增点数、线路数、点要素和线要素
        self._totalDaiyIncreasePointNum: int = None
        # 本日总新增点
        self._totalDailyIncreasePoints: dict = {}
        # 本日新增线路数
        self._totalDaiyIncreaseRouteNum: int = None
        # 本日新增线路
        self._totalDailyIncreaseRoutes: list = []
        # 总计计划线路数
        self._totalPlanNum: int = None
        self._totalPlans: list = []
        # 截止本日总计点数
        self._totalPointNum: int = None
        # 截止本日总计线路数
        self._totalRouteNum: int = None
        # 截止本日所有的点
        self._allPoints: dict = None
        # 截止本日所有的线
        self._allRoutes: list = None
        # 本日各图幅完成的点数量
        self._dailyIncreasedPoints: dict = None
        # 截止本日，各图幅各自完成的点的总数
        self._dailyFinishedPoints: dict = None
        # 本日各图幅完成的线数量
        pass
        # 本日各图幅计划的线数量
        self._dailyPlanedRoutes: dict = None

        # 错误信息
        self.__errorMsg: list = None
        # 获取当天的文件
        self.__datacollect()
    
    @classmethod
    def mapsInfo(cls):
        """
        从100K图幅名称信息表中获取图幅的罗马名称和拉丁名称
        """
        df = pd.read_excel(SHEET_NAMES_FILE, sheet_name="Sheet1", header=0)
        # df = pd.read_excel(SHEET_NAMES_FILE, sheet_name="Sheet1", header=0, index_col=7)
        # 获取数据帧：筛选出 'Sequence' 列值在 SEQUENCE_MIN 和 SEQUENCE_MAX 之间的行
        filtered_df = df[(df['Sequence'] >= SEQUENCE_MIN) & (df['Sequence'] <= SEQUENCE_MAX)]
        # 如果Sequence列长度值不等于 SEQUENCE_MAX - SEQUENCE_MIN + 1，需要报错
        # 如果Sequence列有重复值，需要报错
        if len(filtered_df) != SEQUENCE_MAX - SEQUENCE_MIN+1 or filtered_df['Sequence'].duplicated().any():
            print("图幅信息有误，请检查图幅信息表(Sequence值错误：重复值/缺少定义)\n程序退出")
            exit()
        # 处理可能的 NaN 值，将其填充为一个默认整数值，例如 0
        # filtered_df.loc[:, 'Sequence'] = filtered_df['Sequence'].fillna(-1)
        # 确保 'Sequence' 列为 int 类型
        # filtered_df.loc[:, 'Sequence'] = filtered_df['Sequence'].astype(int)
        # filtered_df['Sequence'] = filtered_df['Sequence'].astype(str)
        # print(filtered_df['Sequence'].dtype)  # 检查数据类型
        # print(filtered_df['Sequence'].values)
        # 判断是否有重复值
        if filtered_df['Sequence'].duplicated().any():
            print("图幅信息表中存在重复的图幅序号，请检查图幅信息表\n程序退出")
            exit()
        # # 按照 'Sequence' 列值从小到大排序
        sorted_df = filtered_df.sort_values(by='Sequence')
        # # 获取图幅名称、罗马名称和拉丁名称，并存储为字典
        # 构建以 'Sequence' 为键的字典
        cls.maps_info = {
            row['Sequence']: {
                'Sheet ID': row['Alternative sheet ID'],
                'Group': row['Group'],
                'File Name': row['File Name'],
                'Arabic Name': row['Arabic'],
                'Roman Name': row['Roman Name'],
                'Latin Name': row['Latin Name'],
                'Team Number': row['Team Number'],
                'Leaders': row['Leaders'],
            }
            for _, row in sorted_df.iterrows()
        }
        # 输出字典
        # print(json.dumps(cls.maps_info, indent=4, ensure_ascii=False))
        # 检查是否获取了所有的图幅信息
        if len(cls.maps_info) != SEQUENCE_MAX - SEQUENCE_MIN + 1:
            print("图幅信息有误，请检查图幅信息表\n程序退出")
            exit()
        return cls.maps_info
    
    def __datacollect(self):
        """
        Purpose: 
        """
        # 每次实例化时，获取当天的所有文件
        # 如果调用了__datacollect()方法，说明需要重新获取当天的文件，因此需要清空当前的文件列表
        self.currentDateFiles = []
        for sequence in range(SEQUENCE_MIN, SEQUENCE_MAX+1):
            mapsheetFileName = self.__class__.maps_info[sequence]['File Name']
            mapsheet = MapsheetDailyFile(mapsheetFileName, self.currentDate)
            self.currentDateFiles.append(mapsheet)
        return self
    
    @property
    def totalDaiyIncreasePointNum(self):
        if self._totalDaiyIncreasePointNum is None:
            total = 0
            for mapsheet in self.currentDateFiles:
                total += mapsheet.dailyincreasePointNum
            self._totalDaiyIncreasePointNum = total
        return self._totalDaiyIncreasePointNum
    
    @property
    def dailyFinishedPoints(self):
        if not self._dailyFinishedPoints:
            sorted_mapsheets = sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)
            dailyPoints = {}
            for mapsheet in sorted_mapsheets:
                dailyPoints[mapsheet.romanName] = mapsheet.currenttotalPointNum
                # print(mapsheet.romanName)
                # print(mapsheet.currenttotalPointNum)
            self._dailyFinishedPoints = dailyPoints
        return self._dailyFinishedPoints
    
    @property
    def dailyIncreasedPoints(self):
        if not self._dailyIncreasedPoints:
            sorted_mapsheets = sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)
            dailyPoints = {}
            for mapsheet in sorted_mapsheets:
                dailyPoints[mapsheet.romanName] = mapsheet.dailyincreasePointNum
            self._dailyIncreasedPoints = dailyPoints
        return self._dailyIncreasedPoints
    
    @property
    def totalDaiyIncreaseRouteNum(self):
        if self._totalDaiyIncreaseRouteNum is None:
            total = 0
            for mapsheet in self.currentDateFiles:
                total += mapsheet.dailyincreaseRouteNum
            self._totalDaiyIncreaseRouteNum = total
        return self._totalDaiyIncreaseRouteNum
    
    @property
    def DailyPlans(self):
        if not self._dailyPlanedRoutes:
            sorted_mapsheets = sorted(self.currentDateFiles, key=lambda mapsheet: mapsheet.sequence)
            dailyPlaneds = {}
            for mapsheet in sorted_mapsheets:
                if mapsheet.nextfilename:
                    dailyPlaneds[mapsheet.romanName] = '#'
                else:
                    dailyPlaneds[mapsheet.romanName] = ''
            self._dailyPlanedRoutes = dailyPlaneds
        return self._dailyPlanedRoutes
    
    @property
    def totalDailyPlanNum(self):
        if self._totalPlanNum is None:
            total = 0
            for mapsheet in self.currentDateFiles:
                if mapsheet.nextfilename:
                    total += 1
            self._totalPlanNum = total
        return self._totalPlanNum
    
    @property
    def totalPointNum(self):
        """
        totalPointNum 截止当天所有文件的点要素总数
                        累加当天的文件的点要素数量
                        如果当天的文件为空，累加上一次提交的文件的点要素总数
        """
        if self._totalPointNum is None:
            totalNum = 0
            for mapsheet in self.currentDateFiles:
                if mapsheet.currentPlacemarks is not None:
                    totalNum += mapsheet.currentPlacemarks.pointsCount
                elif mapsheet.lastPlacemarks is not None:
                    totalNum += mapsheet.lastPlacemarks.pointsCount
            self._totalPointNum = totalNum
        return self._totalPointNum
    
    @property
    def allPoints(self):
        """
        allPoints 截止当天所有文件的点要素
        """
        if self._allPoints is None:
            allPoints = {}
            for mapsheet in self.currentDateFiles:
                # print(mapsheet.currentPlacemarks)
                if mapsheet.currentPlacemarks is not None:
                    allPoints.update(mapsheet.currentPlacemarks.points)
                elif mapsheet.lastPlacemarks is not None:
                    allPoints.update(mapsheet.lastPlacemarks.points)
            self._allPoints = allPoints
        return self._allPoints
    
    @property
    def totalRoutesNum(self):
        """
        totalRoutes 截止当天所有文件的线要素总数
        """
        if self._totalRouteNum is None:
            totalNum = 0
            for mapsheet in self.currentDateFiles:
                if mapsheet.currentPlacemarks is not None:
                    totalNum += mapsheet.currentPlacemarks.routesCount
                elif mapsheet.lastPlacemarks is not None:
                    totalNum += mapsheet.lastPlacemarks.routesCount
            self._totalRouteNum = totalNum
        return self._totalRouteNum

    @property
    def allRoutes(self):
        """
        allRoutes 截止当天所有文件的线要素
        """
        if self._allRoutes is None:
            allRoutes = []
            for mapsheet in self.currentDateFiles:
                if mapsheet.currentPlacemarks is not None:
                    allRoutes.extend(mapsheet.currentPlacemarks.routes)
                elif mapsheet.lastPlacemarks is not None:
                    allRoutes.extend(mapsheet.lastPlacemarks.routes)
            self._allRoutes = allRoutes
        return self._allRoutes

    @property
    def totalFiles(self):
        pass

    @property
    def errorMsg(self):
        if self.__errorMsg is None:
            self.__errorMsg = []
            for mapsheet in self.currentDateFiles:
                self.__errorMsg.append(mapsheet.errorMsg)
                # print(mapsheet.errorMsg)
        return self.__errorMsg

    def __contains__(self, key):
        #! 重写__contains__方法，用于判断图幅文件是否存在
        return key in self.currentDateFiles
    
    def dailyKMZReport(self):
        dailykmz =KMZFile(placemarks=ObservationData(points=self.allPoints, pointsCount=len(self.allPoints), routes=self.allRoutes, routesCount=len(self.allRoutes)))
        dailykmz.write_as(newpath=os.path.join(WORKSPACE, self.currentDate.yyyymm_str, self.currentDate.yyyymmdd_str, f"GMAS_Points_and_tracks_until_{self.currentDate.yyyymmdd_str}.kmz") )
    
    def dailyExcelReport(self):
        """_summary_

        Args:
            date_str (str): 日期字符串，格式为“YYYY/MM/DD”
            excel_filename (str): Excel文件路径
            minSequenceValue (int): 最小填图序列号
            maxSequenceValue (int): 最大填图序列号
        """
        dailyExcel = os.path.join(WORKSPACE, self.currentDate.yyyymm_str, self.currentDate.yyyymmdd_str, f"{self.currentDate.yyyymmdd_str}_Daily_Statistics.xlsx")
        #! 删除已存在的文件
        if os.path.exists(dailyExcel):
            os.remove(dailyExcel)
        romanNames_list = [self.__class__.maps_info[sequence]['Roman Name'] for sequence in range(SEQUENCE_MIN, SEQUENCE_MAX+1)]
        # 计算输出表格的行数和列数
        maxTableRows, maxTableColumns = len(romanNames_list) + 4, 4
        # 每日统计点文件的表头（前三行）
        daily_stat_header1 = ['Date', self.currentDate.yyyy_str + "/" + self.currentDate.mm_str + "/" + self.currentDate.dd_str]
        daily_stat_header2 = ['Map sheet name',
                            'Regular observation points finished',
                            'Field points on revised route'
                            ]
        daily_stat_header3 = ['', '', 'Added observation points',
                            'Added Structure points, photo points, mineralization points'
                            ]
        # 每日统计点文件的合计行（最后一行）
        daily_stat_footer = ['Today', '', '', '']

        # 创建一个新的 Excel 文件
        try:
            book = load_workbook(dailyExcel)
        except FileNotFoundError:
            book = Workbook()

        sheet = book.active
        sheet.title = "Daily Statistics"
        # 写入表头到前三行
        for col_num, value in enumerate(daily_stat_header1, start=1):
            sheet.cell(row=1, column=col_num, value=value)
        for col_num, value in enumerate(daily_stat_header2, start=1):
            sheet.cell(row=2, column=col_num, value=value)
        for col_num, value in enumerate(daily_stat_header3, start=1):
            sheet.cell(row=3, column=col_num, value=value)
        # 写入表尾到最后一行
        for col_num, value in enumerate(daily_stat_footer, start=1):
            sheet.cell(row=maxTableRows, column=col_num, value=value)
        # 从第四行起，写入图幅罗马名称
        for i, value in enumerate(romanNames_list, start=4):
            sheet.cell(row=i, column=1, value=value)
        # 设置字体样式
        # 设置表头字体样式
        font_header = Font(
            name='Calibri',
            size=12,
            bold=True,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00000000'
        )
        # 设置正文字体样式
        font = Font(
            name='Calibri',
            size=11,
            bold=False,
            italic=False,
            vertAlign=None,
            underline='none',
            strike=False,
            color='00000000'
        )

        border = Border(
            left=Side(border_style='thin', color='FF000000'),
            right=Side(border_style='thin', color='FF000000'),
            top=Side(border_style='thin', color='FF000000'),
            bottom=Side(border_style='thin', color='FF000000'),
            diagonal=Side(border_style='thin', color='FF000000'),
            diagonal_direction=0,
            outline=Side(border_style='thin', color='FF000000'),
            vertical=Side(border_style='thin', color='FF000000'),
            horizontal=Side(border_style='thin', color='FF000000')
        )

        # 将字体样式应用到指定单元格
        # 前三行字体为表头字体
        for row in range(1, 4):
            for col in range(1, maxTableColumns + 1):
                cell = sheet.cell(row, column=col)
                cell.font = font_header
        # 最后一行字体为表头字体
        for col in range(1, maxTableColumns + 1):
            cell = sheet.cell(maxTableRows, column=col)
            cell.font = font_header
        # 其他字体为正文字体
        for row in range(4, maxTableRows):
            for col in range(1, maxTableColumns + 1):
                cell = sheet.cell(maxTableRows-1, column=col)
                cell.font = font

        # 将单元格格式应用到指定单元格
        for row in range(maxTableRows):
            for col in range(maxTableColumns):
                cell = sheet.cell(row=row+1, column=col+1)
                cell.border = border

        # 设置多列的宽度，adjusted_width是单元格宽度，比最大长度多2
        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception:
                    pass
            adjusted_width = (max_length + 2)  # 设置单元格宽度，比最大长度多2
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        # 创建一个居中对齐的对象
        center_aligned = Alignment(horizontal='center', vertical='center')
        # 将居中对齐应用到指定单元格
        for row in sheet['A1:D3']:  # 修改这个范围以适应你的需要
            for cell in row:
                cell.alignment = center_aligned
        for row in range(4, maxTableRows):
            for col in range(2, maxTableColumns):
                cell = sheet.cell(row, column=col)
                cell.alignment = center_aligned
        for row in range(maxTableRows + 1):
            for col in range(1, maxTableColumns + 1):
                cell = sheet.cell(row+1, column=col)
                cell.alignment = center_aligned

        # 设置合计行的公式
        sheet.cell(row=maxTableRows, column=2).value = f"=SUM(B4:B{maxTableRows-1})"
        sheet.cell(row=maxTableRows, column=3).value = f"=SUM(C4:C{maxTableRows-1})"
        sheet.cell(row=maxTableRows, column=4).value = f"=SUM(D4:D{maxTableRows-1})"

        # 先设置值，再合并单元格
        sheet.cell(row=1, column=2).value = self.currentDate.yyyy_str + "/" + self.currentDate.mm_str + "/" + self.currentDate.dd_str

        sheet.merge_cells('B1:D1')
        sheet.merge_cells('C2:D2')
        sheet.merge_cells('A2:A3')
        sheet.merge_cells('B2:B3')

        # 保存工作簿
        book.save(dailyExcel)
        print(f"创建每日统计点 {dailyExcel} 空表成功。")
        return True
    
    def onScreenDisplay(self):
        # 填图组号
        team_list = []
        # 图幅罗马名称
        map_name_list = []
        # 当天新增点数
        daily_collection_list = []
        # 截止当天，图幅完成的总点数
        daily_Finished_list = []
        # 图幅第二天的野外计划
        daily_plan_list = []
        for key, value in self.dailyIncreasedPoints.items():
            map_name_list.append(key)
            x = ''  # Define x with an appropriate value
            # 各个图幅当天的完成点数，如果完成点数为0，则显示空字符串，否则显示完成点数
            daily_collection_list.append(x if self.dailyIncreasedPoints[key] == 0 else self.dailyIncreasedPoints[key])
            # 各个图幅截止当天的完成点数，如果完成点数为0，则显示空字符串，否则显示完成点数
            daily_Finished_list.append(x if self.dailyFinishedPoints[key] == 0 else self.dailyFinishedPoints[key])
            daily_plan_list.append(self.DailyPlans[key])
        table_data = []
        # 调整显示的每行顺序
        for i in range(len(map_name_list)):
            table_data.append([i + 1, map_name_list[i], daily_collection_list[i], daily_plan_list[i], daily_Finished_list[i]])
        # 添加总计行
        table_data.append(["TOTAL", "", self.totalDaiyIncreasePointNum, self.totalDailyPlanNum, self.totalPointNum])
        print('\n'*2)
        headers = ["Seq", "Name", "Increase", "Plan", "Finished"]
        print(tabulate(table_data, headers, tablefmt="grid"))
        # print('\n'*1)

    
    def dailyExcelReportUpdate(self):
        dailyExcel = os.path.join(WORKSPACE, self.currentDate.yyyymm_str, self.currentDate.yyyymmdd_str, f"{self.currentDate.yyyymmdd_str}_Daily_Statistics.xlsx")
        completed_points = [value for key, value in self.dailyIncreasedPoints.items()]
        # 将列表中的0值替换为np.nan
        # 如果注释掉下面这行代码，那么0值将会被替换为np.nan，列表中的0值将不会被写入到Excel文件中
        # completed_points = [value if value != 0 else np.nan for value in completed_points]
        completed_points_series = pd.Series(completed_points)
        book = load_workbook(dailyExcel)
        sheet = book['Daily Statistics']
        # 将数据写入到指定的单元格
        for i, value in enumerate(completed_points_series, start=4):
            sheet.cell(row=i, column=2, value=value)
        # 保存工作簿
        book.save(dailyExcel)
        # os.startfile(dailyExcel)
        print(f"每日统计点写入 {dailyExcel} 成功。")
        return True

    
class DataSubmition(object):
    """
    根据COLLECTION_WEEKDAYS列表，在每周的某一天：
    1. 输出shp文件
    2. 在制图工程文件夹下创建目录，并拷贝截止当天的shp
    3. 拷贝一周前的shp文件（或kmz文件）
    4. 生成一周新增的shp文件
    5. 调用ArcPy，生成一周报告
    """
    SHP_EXTENSIONS = ['.shp', '.shx', '.dbf', '.prj']

    def __init__(self, date: 'DateType', points_dict: dict):

        self.date = date
        self.pointDict: dict = points_dict

    def __pointDictValidation(self):
        """
        检查点要素字典的有效性
        """
        pass

    def weeklyPointToShp(self):
        # 在制图工程文件夹下建立新文件夹，如果已存在则删除清理文件树
        week_report_folder = os.path.join(WORKSPACE, MAP_PROJECT_FOLDER, f"Finished_ObsPoints_Until_{self.date.yyyymmdd_str}")
        weekdayDir = os.path.join(WORKSPACE, MAP_PROJECT_FOLDER, f"Finished_ObsPoints_Until_{self.date.yyyymmdd_str}")
        if os.path.exists(weekdayDir):
            shutil.rmtree(weekdayDir)
        os.makedirs(weekdayDir, exist_ok=True)

        # 输出shp文件
        output_shp_file = os.path.join(week_report_folder, f"GMAS_points_until_{self.date.yyyymmdd_str}.shp")
        self.pointDictToShp(self.pointDict, output_shp_file)


        # 将shp文件压缩为zip文件，如果zip文件已存在，则删除
        zip_file = os.path.join(WORKSPACE, self.date.yyyymm_str, self.date.yyyymmdd_str, f"GMAS_points_until_{self.date.yyyymmdd_str}.zip")
        if os.path.exists(zip_file):
            os.remove(zip_file)
        with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
            for ext in DataSubmition.SHP_EXTENSIONS:
                zipf.write(output_shp_file.replace('.shp', ext), os.path.basename(output_shp_file).replace('.shp', ext))
        logger.info(f"截止今日的完成点已生成SHP文件，并已压缩为ZIP文件: {zip_file}")

        # 尝试从多种路径读取一周前的文件(zip，shp)
        # 工作目录下的存档文件夹中的zip文件
        one_week_ago = DateType(date_datetime=(self.date.date_datetime - timedelta(days=7)))
        # 从制图工程文件夹中的shp文件
        one_week_ago_shpfile = os.path.join(WORKSPACE, MAP_PROJECT_FOLDER, f"Finished_ObsPoints_Until_{one_week_ago.yyyymmdd_str}", f"GMAS_points_until_{one_week_ago.yyyymmdd_str}.shp")
        # 从存档文件夹中读取一周前的zip文件
        one_week_ago_zipfile = os.path.join(WORKSPACE, one_week_ago.yyyymm_str, one_week_ago.yyyymmdd_str, f"GMAS_points_until_{one_week_ago.yyyymmdd_str}.zip")
        # 从存档文件夹中读取一周前的kmz文件
        # one_week_ago_kmzfile = os.path.join(WORKSPACE, one_week_ago.yyyymm_str, one_week_ago.yyyymmdd_str, f"GMAS_Points_and_tracks_until_{one_week_ago.yyyymmdd_str}.kmz")

        #TODO: 也可以尝试从kmz文件中读取一周前的点要素
        match (os.path.exists(one_week_ago_shpfile), os.path.exists(one_week_ago_zipfile)):
            case (True, _):
                # 拷贝一周前的shp文件至制图文件夹中
                for ext in DataSubmition.SHP_EXTENSIONS:
                    shutil.copy(one_week_ago_shpfile.replace('.shp', ext), week_report_folder)
                    os.chmod(week_report_folder, stat.S_IWRITE | stat.S_IREAD)
                # logger.info(f"已拷贝一周前的 SHP 文件及相关文件至: {week_report_folder}")
            case (False, True):
                # 解压一周前的zip文件至制图文件夹中
                with zipfile.ZipFile(one_week_ago_zipfile, 'r') as zip_ref:
                    zip_ref.extractall(week_report_folder)
                # logger.info(f"已解压一周前的 ZIP 文件到: {week_report_folder}")
            case (False, False):
                logger.warning(f"一周前的文件不存在: {one_week_ago_zipfile}\n{one_week_ago_shpfile}")
                return False

        one_week_ago_points_dict = self.readShpToPointDict(os.path.join(week_report_folder, f"GMAS_points_until_{one_week_ago.yyyymmdd_str}.shp"))
        count_one_week_ago_points_dict = len(one_week_ago_points_dict)
        logger.info(f"一周前{one_week_ago.yyyymmdd_str}的点要素总数: {count_one_week_ago_points_dict}")
        # 计算本周新增的点要素
        diff_dict = {k: v for k, v in self.pointDict.items() if k not in one_week_ago_points_dict}

        one_week_ago_nextday = DateType(date_datetime=(self.date.date_datetime - timedelta(days=6)))
        logger.info(f"{one_week_ago_nextday.yyyymmdd_str}至{self.date.yyyymmdd_str}，本周新增点要素: {len(diff_dict)}")

        weekly_increase_shp_file = os.path.join(week_report_folder, f"GMAS_points_{one_week_ago_nextday.yyyymmdd_str}_{self.date.yyyymmdd_str}.shp")
        self.pointDictToShp(diff_dict, weekly_increase_shp_file)

        logger.info(f"GMAS_points_{one_week_ago_nextday.yyyymmdd_str}_{self.date.yyyymmdd_str}.shp创建成功")

    def pointDictToShp(self, pointDict, output_shp_file):
        # 创建 SHP 文件驱动
        shp_driver = ogr.GetDriverByName('ESRI Shapefile')
        if os.path.exists(output_shp_file):
            shp_driver.DeleteDataSource(output_shp_file)
        shp_ds = shp_driver.CreateDataSource(output_shp_file)
        if shp_ds is None:
            raise RuntimeError("Failed to create SHP file")
        # 创建 SHP 图层
        srs = osr.SpatialReference()
        srs.ImportFromEPSG(4326)  # WGS84
        shp_layer = shp_ds.CreateLayer('points', srs, ogr.wkbPoint)
        # 创建字段
        field_name = ogr.FieldDefn('Name', ogr.OFTString)
        field_name.SetWidth(24)
        shp_layer.CreateField(field_name)
        field_longitude = ogr.FieldDefn('Longitude', ogr.OFTReal)
        shp_layer.CreateField(field_longitude)
        field_latitude = ogr.FieldDefn('Latitude', ogr.OFTReal)
        shp_layer.CreateField(field_latitude)
        # 创建要素并添加到图层
        for obspid, coords in pointDict.items():
            feature = ogr.Feature(shp_layer.GetLayerDefn())
            feature.SetField('Name', obspid)
            feature.SetField('Longitude', float(coords['longitude']))
            feature.SetField('Latitude', float(coords['latitude']))
            point = ogr.Geometry(ogr.wkbPoint)
            point.AddPoint(float(coords['longitude']), float(coords['latitude']))
            feature.SetGeometry(point)
            shp_layer.CreateFeature(feature)
            feature = None  # 清理要素
        # 清理
        shp_ds = None
        print(f"点要素已成功生成 SHP 文件: {output_shp_file}")

    def readShpToPointDict(self, shp_file):
        points_dict = {}
        # 打开 SHP 文件
        shp_driver = ogr.GetDriverByName('ESRI Shapefile')
        shp_ds = shp_driver.Open(shp_file, 0)  # 0 表示只读模式
        if shp_ds is None:
            raise RuntimeError(f"Failed to open SHP file: {shp_file}")
        # 获取图层
        shp_layer = shp_ds.GetLayer()
        # 遍历要素
        for feature in shp_layer:
            obspid = feature.GetField('Name')
            longitude = feature.GetField('Longitude')
            latitude = feature.GetField('Latitude')
            points_dict[obspid] = {
                'longitude': longitude,
                'latitude': latitude
            }
        # 清理
        shp_ds = None
        # print(f"已成功读取 SHP 文件: {shp_file}")
        return points_dict


if __name__ == "__main__":

    #KMZFile类测试
    #错误路径
    # kmz = KMZFile(filepath=r"D:\Workspace\2024\20241224\Ad_Dawadami_finished_points_and_tracks_20241224.kmz")
    # 正确路径
    # kmz = KMZFile(filepath=r"D:\RouteDesigen\202412\20241224\Finished points\Ad_Dawadami_finished_points_and_tracks_20241224.kmz")
    # print(kmz.placemarks)
    # exit()

    # MapsheetDailyFile类测试
    # Ad_Dawadami = MapsheetDailyFile("Ad_Dawadami", DateType(yyyymmdd_str="20241226"))
    # print(Ad_Dawadami.__dict__)
    # exit()

    date = DateType(date_datetime=datetime.now())
    # 回溯日期测试
    while date.date_datetime > datetime.strptime(TRACEBACK_DATE, "%Y%m%d"):
        collection = CurrentDateFiles(date)
        print(f"{date}新增点数：", collection.totalDaiyIncreasePointNum, 3*"\n")
        date = DateType(date_datetime=date.date_datetime - timedelta(days=1))
    exit()
